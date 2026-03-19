"""
PDF → Excel  ·  ULTIMATE AI-POWERED CONVERTER
================================================
Combines two approaches for maximum accuracy:

  LAYER 1 — Accurate structural extraction (no AI needed):
    • Uses PDF's actual vertical lines as column boundaries
    • Character-level x/y positioning for exact cell placement
    • Dot-leader absorption, bold/size detection
    • Image extraction via PyMuPDF
    • Works offline, handles ANY text-based PDF

  LAYER 2 — Azure GPT-4.1 AI (optional, for hard cases):
    • Scanned/image PDFs → Vision AI extracts tables
    • Complex/broken layouts → AI cleans and reconstructs
    • Confidential data → AI normalizes numbers correctly
    • Multi-language PDFs → AI handles Unicode

  DECISION LOGIC (automatic, per page):
    ┌─────────────────────────────────────────────────────┐
    │ chars > 50  →  Layer 1 (structural extraction)      │
    │   + weak structure (few lines) → AI cleans result   │
    │ chars ≤ 50  →  Layer 2 (Vision AI, scanned PDF)     │
    └─────────────────────────────────────────────────────┘

INSTALL:
    pip install pdfplumber pymupdf pillow openpyxl openai tenacity

CONFIGURE:
    Set AZURE_API_KEY, AZURE_ENDPOINT, AZURE_MODEL below
    Set USE_AI = False to run offline (structural only)

RUN:
    python pdf_to_excel.py  input.pdf  output.xlsx
"""

import sys, os, io, re, json, base64
from collections import defaultdict

# ── third-party ────────────────────────────────────────────────────────────────
try:
    import fitz; HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import pdfplumber
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

try:
    from openai import AzureOpenAI
    from tenacity import retry, stop_after_attempt, wait_fixed
    HAS_OPENAI = True
except ImportError:
    HAS_OPENAI = False

# ══════════════════════════════════════════════════════════════════════════════
# ▶ CONFIGURATION  —  edit these
# ══════════════════════════════════════════════════════════════════════════════
USE_AI         = True                              # set False to skip AI entirely
AZURE_API_KEY  = "YOUR_AZURE_KEY"                 # ← paste your key
AZURE_ENDPOINT = "https://YOUR-RESOURCE.openai.azure.com/"
AZURE_API_VER  = "2024-02-15-preview"
AZURE_MODEL    = "gpt-4.1"                        # your deployed model name

# Thresholds
MIN_CHARS_FOR_TEXT_MODE = 50   # pages with fewer chars → vision AI
MIN_LINES_FOR_STRUCTURE = 3    # pages with fewer h/v lines → AI clean

# ══════════════════════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════════════════════
def solid(h): return PatternFill("solid", fgColor=h)
def mk_bdr(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)
THIN  = mk_bdr("thin",   "CCCCCC")
THICK = mk_bdr("medium", "333333")
def fnt(bold=False, size=9, color="111111"):
    return Font(name="Arial", bold=bold, size=int(max(size, 7)), color=color)
def aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

# ══════════════════════════════════════════════════════════════════════════════
# AZURE AI CLIENT
# ══════════════════════════════════════════════════════════════════════════════
_ai_client = None

def get_ai_client():
    global _ai_client
    if _ai_client is None:
        if not HAS_OPENAI:
            raise ImportError("pip install openai tenacity")
        _ai_client = AzureOpenAI(
            api_key      = AZURE_API_KEY,
            api_version  = AZURE_API_VER,
            azure_endpoint = AZURE_ENDPOINT,
        )
    return _ai_client

# ══════════════════════════════════════════════════════════════════════════════
# LAYER 2A — AI TABLE CLEANER  (text PDF with complex layout)
# ══════════════════════════════════════════════════════════════════════════════
def ai_clean_rows(raw_rows: list) -> list:
    """
    Send raw extracted rows to GPT-4.1.
    Returns a normalized list-of-lists with consistent column count.
    """
    if not USE_AI or not HAS_OPENAI:
        return raw_rows

    client = get_ai_client()
    prompt = f"""You are a precise table reconstruction system.
Given raw rows extracted from a PDF (may have broken cells, merged content, noise):

RULES:
1. All rows MUST have the SAME number of columns
2. Merge any broken cells that belong together
3. Preserve ALL numbers exactly as-is (do not round or change)
4. Keep original meaning, remove only noise characters
5. Return ONLY a valid JSON array of arrays — no explanation, no markdown

Raw rows:
{json.dumps(raw_rows, ensure_ascii=False)}
"""
    try:
        resp = client.chat.completions.create(
            model       = AZURE_MODEL,
            messages    = [{"role": "user", "content": prompt}],
            temperature = 0,
        )
        content = resp.choices[0].message.content.strip()
        # strip markdown fences if present
        content = re.sub(r"^```(?:json)?|```$", "", content, flags=re.M).strip()
        return json.loads(content)
    except Exception as e:
        print(f"    [AI clean error] {e} — using raw rows")
        return raw_rows

# ══════════════════════════════════════════════════════════════════════════════
# LAYER 2B — VISION AI  (scanned / image PDF)
# ══════════════════════════════════════════════════════════════════════════════
def fitz_page_to_b64(fitz_doc, page_num: int) -> str:
    """Render a PDF page to PNG and return base64 string."""
    pg  = fitz_doc[page_num]
    pix = pg.get_pixmap(dpi=150)
    return base64.b64encode(pix.tobytes("png")).decode()

def ai_vision_extract(fitz_doc, page_num: int) -> list:
    """
    Use GPT-4.1 Vision to extract table from a scanned/image PDF page.
    Returns list-of-lists.
    """
    if not USE_AI or not HAS_OPENAI:
        return []

    client = get_ai_client()
    img_b64 = fitz_page_to_b64(fitz_doc, page_num)

    prompt = """Extract ALL content from this PDF page into a structured table.
Rules:
- Return a JSON array of arrays (rows × columns)
- Preserve every number exactly as printed
- Include all text, labels, headers, values
- Return ONLY valid JSON — no explanation"""

    try:
        resp = client.chat.completions.create(
            model    = AZURE_MODEL,
            messages = [{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url",
                     "image_url": {"url": f"data:image/png;base64,{img_b64}"}},
                ],
            }],
            temperature = 0,
        )
        content = resp.choices[0].message.content.strip()
        content = re.sub(r"^```(?:json)?|```$", "", content, flags=re.M).strip()
        return json.loads(content)
    except Exception as e:
        print(f"    [Vision AI error] {e}")
        return []

# ══════════════════════════════════════════════════════════════════════════════
# LAYER 1 — STRUCTURAL EXTRACTION ENGINE
# ══════════════════════════════════════════════════════════════════════════════

# ── images from PDF ────────────────────────────────────────────────────────────
def get_embedded_images(fitz_doc, page_num):
    out = []
    pg  = fitz_doc[page_num]
    for info in pg.get_images(full=True):
        xref = info[0]
        try:
            b   = fitz_doc.extract_image(xref)
            pil = PILImage.open(io.BytesIO(b["image"])).convert("RGBA")
            rs  = pg.get_image_rects(xref)
            r   = rs[0] if rs else fitz.Rect(0, 0, pil.width, pil.height)
            out.append(dict(x0=r.x0, y0=r.y0, x1=r.x1, y1=r.y1, pil=pil))
        except: pass
    return out

def pil_to_xl(pil, wp, hp):
    wp, hp = max(wp, 20), max(hp, 20)
    bg = PILImage.new("RGB", pil.size, (255, 255, 255))
    bg.paste(pil.convert("RGB"),
             mask=pil.split()[3] if pil.mode == "RGBA" else None)
    bg = bg.resize((wp, hp), PILImage.LANCZOS)
    buf = io.BytesIO(); bg.save(buf, "PNG"); buf.seek(0)
    xl = XLImage(buf); xl.width = wp; xl.height = hp; return xl

# ── column boundaries ──────────────────────────────────────────────────────────
def get_col_slots(page, snap=4) -> list:
    """
    Column boundaries = PDF's actual vertical lines + page edges.
    This gives exactly the same column count as the form designer intended.
    """
    raw = [float(l["x0"]) for l in page.lines if abs(l["x0"] - l["x1"]) < 1]
    raw += [0.0, float(page.width)]
    out = []
    for v in sorted(set(round(x, 1) for x in raw)):
        if out and v - out[-1] <= snap: out[-1] = (out[-1] + v) / 2
        else: out.append(v)
    return sorted(out)

# ── visual lines from characters ───────────────────────────────────────────────
def chars_to_lines(chars, y_snap=1.5):
    bkts = defaultdict(list)
    for ch in chars:
        key = round(float(ch["top"]) / y_snap) * y_snap
        bkts[key].append(ch)
    lines = []
    for k in sorted(bkts):
        chs = sorted(bkts[k], key=lambda c: float(c["x0"]))
        lines.append((sum(float(c["top"]) for c in chs) / len(chs), chs))
    return lines

def is_dot_leader(chs):
    return all(c["text"].strip() in (".", "", " ") for c in chs)

def chars_to_segs(chs, gap=8) -> list:
    """Split characters into text segments by x-gap. Absorb dot leaders."""
    if not chs: return []
    chs  = sorted(chs, key=lambda c: float(c["x0"]))
    segs = []; seg = [chs[0]]
    for i in range(1, len(chs)):
        if float(chs[i]["x0"]) - float(chs[i-1]["x1"]) > gap:
            segs.append(seg); seg = []
        seg.append(chs[i])
    segs.append(seg)

    result = []
    for s in segs:
        txt = "".join(c["text"] for c in s).strip()
        if not txt: continue
        if is_dot_leader(s) and result:
            result[-1]["x1"] = float(s[-1]["x1"])
            continue
        result.append(dict(
            x0   = float(s[0]["x0"]),
            x1   = float(s[-1]["x1"]),
            text = txt,
            bold = any("Bold" in str(c.get("fontname", "")) for c in s),
            size = float(s[0].get("size") or 9),
        ))
    return result

# ── filled rectangle detection ─────────────────────────────────────────────────
def get_fills(page) -> list:
    out = []
    for r in page.rects:
        c = r.get("non_stroking_color")
        if c is None: continue
        if isinstance(c, (int, float)): c = (float(c),) * 3
        if not (isinstance(c, (list, tuple)) and len(c) >= 3): continue
        r0, g0, b0 = float(c[0]), float(c[1]), float(c[2])
        if   r0 < 0.15 and g0 < 0.15 and b0 < 0.15: kind = "black"
        elif r0 > 0.50 and g0 > 0.50 and b0 > 0.50: kind = "grey"
        else: continue
        out.append(dict(top=float(r["top"]), bottom=float(r["bottom"]),
                        x0=float(r["x0"]),   x1=float(r["x1"]),
                        width=float(r["x1"] - r["x0"]), kind=kind))
    return out

def fill_at(y, fills, page_width, tol=4):
    for f in fills:
        if f["top"] - tol <= y <= f["bottom"] + tol:
            if f["width"] > page_width * 0.25:   # only wide rects are headers
                return f["kind"]
    return None

# ── slot helpers ───────────────────────────────────────────────────────────────
def slot_of(x, slots):
    best = 0
    for i in range(len(slots) - 1):
        if slots[i] - 3 <= x: best = i
    return best

def end_slot_of(x1, slots):
    best = 0
    for i in range(len(slots) - 1):
        if slots[i] <= x1 + 4: best = i
    return best

def detect_align(seg, sx0, sx1):
    mt = (seg["x0"] + seg["x1"]) / 2
    mc = (sx0 + sx1) / 2
    cw = sx1 - sx0
    if cw < 5: return "left"
    if abs(mt - mc) <= cw * 0.20: return "center"
    if seg["x0"] > sx0 + cw * 0.55: return "right"
    return "left"

def safe_merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2: return
    try: ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    except: pass

# ── extract raw rows for AI cleaning ──────────────────────────────────────────
def extract_raw_rows(page) -> list:
    """Extract text rows using column slots — for passing to AI cleaner."""
    slots = get_col_slots(page)
    vis   = chars_to_lines(page.chars)
    rows  = []
    for (y, chs) in vis:
        segs = chars_to_segs(chs)
        row  = [""] * (len(slots) - 1)
        for seg in segs:
            ci = slot_of(seg["x0"], slots)
            ci = min(ci, len(row) - 1)
            row[ci] = (row[ci] + " " + seg["text"]).strip()
        if any(r.strip() for r in row):
            rows.append(row)
    return rows

# ── write AI rows to worksheet ─────────────────────────────────────────────────
def write_ai_rows(ws, rows: list):
    """Write AI-cleaned rows (list-of-lists) to worksheet with basic styling."""
    if not rows: return

    # Normalize column count
    max_cols = max(len(r) for r in rows)
    for row in rows:
        while len(row) < max_cols: row.append("")

    # Column widths
    for ci in range(max_cols):
        ws.column_dimensions[get_column_letter(ci + 1)].width = 18

    for ri, row in enumerate(rows, start=1):
        bg = "F5F5F5" if ri % 2 == 0 else "FFFFFF"
        ws.row_dimensions[ri].height = 14
        for ci, val in enumerate(row, start=1):
            v    = str(val).strip() if val else ""
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font      = fnt(size=9)
            cell.fill      = solid(bg)
            cell.alignment = aln("left")
            cell.border    = THIN

# ── write structural page to worksheet ────────────────────────────────────────
def write_structural_page(ws, page, page_images):
    """Layer 1: accurate structural extraction → Excel."""
    pw    = float(page.width)
    chars = page.chars
    fills = get_fills(page)
    slots = get_col_slots(page)
    n_sl  = len(slots) - 1

    vis_lines = chars_to_lines(chars)
    line_ys   = [y for y, _ in vis_lines]

    for ci in range(n_sl):
        cw = slots[ci + 1] - slots[ci]
        ws.column_dimensions[get_column_letter(ci + 1)].width = max(3, round(cw / 5.0))

    for xl_row, (y, chs) in enumerate(vis_lines, start=1):
        segs = chars_to_segs(chs)
        kind = fill_at(y, fills, pw)
        sz   = segs[0]["size"] if segs else 9
        ws.row_dimensions[xl_row].height = max(10, round(sz * 1.65))
        if not segs: continue

        # ── Black section header ─────────────────────────────────────────────
        if kind == "black":
            has_kw = any(re.search(r'\bPart\b|Summary|Recapture', s["text"], re.I)
                         for s in segs)
            if has_kw:
                txt  = "  " + "  ".join(s["text"] for s in segs)
                cell = ws.cell(row=xl_row, column=1, value=txt)
                cell.font      = fnt(bold=True, size=max(int(sz), 9), color="FFFFFF")
                cell.fill      = solid("111111")
                cell.alignment = aln("left")
                cell.border    = THICK
                ws.row_dimensions[xl_row].height = 15
                if n_sl > 1: safe_merge(ws, xl_row, 1, xl_row, n_sl)
                continue

        # ── Grey column header ───────────────────────────────────────────────
        if kind == "grey":
            used = set()
            for seg in segs:
                ci_s = slot_of(seg["x0"], slots)
                ci_e = end_slot_of(seg["x1"], slots)
                if ci_s + 1 in used: continue
                al   = detect_align(seg, slots[ci_s], slots[min(ci_s+1, len(slots)-1)])
                cell = ws.cell(row=xl_row, column=ci_s+1, value=seg["text"])
                cell.font      = fnt(bold=True, size=max(int(seg["size"]), 8))
                cell.fill      = solid("DCDCDC")
                cell.alignment = aln(al)
                cell.border    = THIN
                if ci_e > ci_s:
                    safe_merge(ws, xl_row, ci_s+1, xl_row, ci_e+1)
                    for c in range(ci_s+1, ci_e+2): used.add(c)
                else: used.add(ci_s+1)
            ws.row_dimensions[xl_row].height = 13
            continue

        # ── Normal row ───────────────────────────────────────────────────────
        bg   = "F5F5F5" if xl_row % 2 == 0 else "FFFFFF"
        used = set()
        for seg in segs:
            ci_s = slot_of(seg["x0"], slots)
            ci_e = end_slot_of(seg["x1"], slots)
            if ci_s + 1 in used: continue
            al   = detect_align(seg, slots[ci_s], slots[min(ci_s+1, len(slots)-1)])
            cell = ws.cell(row=xl_row, column=ci_s+1, value=seg["text"])
            cell.font      = fnt(bold=seg["bold"], size=max(int(seg["size"]), 8))
            cell.fill      = solid(bg)
            cell.alignment = aln(al)
            cell.border    = THIN
            if ci_e > ci_s:
                safe_merge(ws, xl_row, ci_s+1, xl_row, ci_e+1)
                for c in range(ci_s+1, ci_e+2): used.add(c)
            else: used.add(ci_s+1)

    # ── Embedded images ───────────────────────────────────────────────────────
    PX = 96 / 72
    for img in page_images:
        wp = max(int((img["x1"] - img["x0"]) * PX), 30)
        hp = max(int((img["y1"] - img["y0"]) * PX), 20)
        try:
            xl_img = pil_to_xl(img["pil"], wp, hp)
            ar = 1
            for i, ly in enumerate(line_ys):
                if ly <= img["y0"]: ar = i + 1
            ac = slot_of(img["x0"], slots) + 1
            ws.add_image(xl_img, f"{get_column_letter(ac)}{ar}")
        except Exception as e:
            print(f"    img: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ══════════════════════════════════════════════════════════════════════════════
def is_weak_structure(page) -> bool:
    """True if the page layout is too complex/sparse for pure structural extraction."""
    n_lines = len(page.lines)
    n_chars = len(page.chars)
    # Few drawn lines + sparse text = complex layout needing AI
    return n_lines < MIN_LINES_FOR_STRUCTURE and n_chars < 300

def convert(pdf_path: str, out_path: str):
    print(f"\n{'='*60}")
    print(f"  PDF → Excel  (AI={'ON' if USE_AI and HAS_OPENAI else 'OFF'})")
    print(f"  Input:  {pdf_path}")
    print(f"  Output: {out_path}")
    print(f"{'='*60}\n")

    wb = Workbook()
    fitz_doc = fitz.open(pdf_path) if HAS_FITZ else None

    with pdfplumber.open(pdf_path) as pdf:
        for pn, page in enumerate(pdf.pages):
            ws = wb.active if pn == 0 else wb.create_sheet()
            ws.title = f"Page {pn+1}"

            n_chars = len(page.chars)
            n_lines = len(page.lines)
            slots   = get_col_slots(page)
            mode    = "?"

            print(f"  [Page {pn+1}]  {page.width:.0f}×{page.height:.0f} pts  "
                  f"{n_chars} chars  {n_lines} lines  {len(slots)-1} cols")

            # ── SCANNED / IMAGE PAGE → Vision AI ─────────────────────────────
            if n_chars < MIN_CHARS_FOR_TEXT_MODE:
                if USE_AI and HAS_OPENAI and fitz_doc:
                    mode = "VISION-AI"
                    print(f"    → Mode: {mode} (scanned PDF)")
                    rows = ai_vision_extract(fitz_doc, pn)
                    write_ai_rows(ws, rows)
                else:
                    mode = "SKIP (scanned, no AI)"
                    print(f"    → Mode: {mode}")

            # ── TEXT PAGE WITH WEAK STRUCTURE → Structural + AI clean ─────────
            elif is_weak_structure(page) and USE_AI and HAS_OPENAI:
                mode = "STRUCTURAL+AI-CLEAN"
                print(f"    → Mode: {mode}")
                raw_rows = extract_raw_rows(page)
                cleaned  = ai_clean_rows(raw_rows)
                write_ai_rows(ws, cleaned)

            # ── TEXT PAGE WITH GOOD STRUCTURE → Pure structural ───────────────
            else:
                mode = "STRUCTURAL"
                print(f"    → Mode: {mode}")
                pg_imgs = get_embedded_images(fitz_doc, pn) if fitz_doc else []
                write_structural_page(ws, page, pg_imgs)

            print(f"    ✓ Done  [{mode}]")

    if fitz_doc: fitz_doc.close()
    wb.save(out_path)
    print(f"\n  ✅  Saved → {out_path}\n")

# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        print("Usage: python pdf_to_excel.py  input.pdf  output.xlsx")
        sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"File not found: {sys.argv[1]}"); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
