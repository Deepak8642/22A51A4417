"""
PDF → Excel  DYNAMIC CONVERTER  —  Works on ANY PDF
====================================================
No hardcoding. No API. 100% offline.
Replicates ilovepdf.com quality for any PDF type.

MODES (auto-selected per page):
  GRID  — PDF has drawn vertical lines → exact column positions
  TABLE — pdfplumber detects structured tables
  TEXT  — Free-form text → cluster by x-position

pip install pdfplumber openpyxl
python pdf_to_excel.py  input.pdf  [output.xlsx]
"""
import sys, os, re, glob
from collections import defaultdict
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── constants ─────────────────────────────────────────────────────────────────
Y_SNAP    = 1.8
SEG_GAP   = 6.0
VLINE_MIN = 2
TABLE_COLS= 2
LINE_MIN  = 5.0

# ── styles ────────────────────────────────────────────────────────────────────
def _sd(s="thin",c="000000"): return Side(style=s,color=c)
def _bfull():
    s=_sd(); return Border(top=s,bottom=s,left=s,right=s)
def _bnone(): return Border()
def _fnone(): return PatternFill(fill_type=None)
def _fill(rgb): return PatternFill("solid",fgColor=rgb)
def _fnt(bold=False,sz=9,color="000000"):
    return Font(name="Calibri",bold=bold,size=max(int(sz),7),color=color)
def _aln(h="left",wrap=True):
    return Alignment(horizontal=h,vertical="center",wrap_text=wrap)

def wc(cell,val,bold=False,sz=9,align="left",wrap=True,border=True,fill=None,color="000000"):
    cell.value=str(val) if val is not None else ""
    cell.number_format="@"
    cell.font=_fnt(bold=bold,sz=sz,color=color)
    cell.fill=_fill(fill) if fill else _fnone()
    cell.alignment=_aln(align,wrap=wrap)
    cell.border=_bfull() if border else _bnone()

def mg(ws,r1,c1,r2,c2):
    if r1==r2 and c1==c2: return
    try: ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    except: pass

def set_cws(ws,slots):
    for ci in range(len(slots)-1):
        w=(slots[ci+1]-slots[ci])/5.5
        ws.column_dimensions[get_column_letter(ci+1)].width=max(3.5,w)

def rh(ws,r,h): ws.row_dimensions[r].height=h

# ── pdf parsing ───────────────────────────────────────────────────────────────
def get_rows(chars,y_snap=Y_SNAP):
    b=defaultdict(list)
    for c in chars: b[round(float(c["top"])/y_snap)*y_snap].append(c)
    return [(sum(float(c["top"]) for c in v)/len(v),
             sorted(v,key=lambda c:float(c["x0"])))
            for k in sorted(b) for v in [b[k]]]

def get_segs(chs,gap=SEG_GAP):
    if not chs: return []
    chs=sorted(chs,key=lambda c:float(c["x0"]))
    groups=[[chs[0]]]
    for i in range(1,len(chs)):
        if float(chs[i]["x0"])-float(chs[i-1]["x1"])>gap: groups.append([])
        groups[-1].append(chs[i])
    result=[]
    for g in groups:
        txt="".join(c["text"] for c in g).strip()
        if not txt: continue
        if all(c in ". " for c in txt) and result:
            result[-1]["x1"]=float(g[-1]["x1"]); continue
        result.append(dict(x0=float(g[0]["x0"]),x1=float(g[-1]["x1"]),
            text=txt,bold=any("Bold" in str(c.get("fontname","")) or
            "bold" in str(c.get("fontname","")).lower() for c in g),
            size=float(g[0].get("size") or 9)))
    return result

def snap(vals,tol):
    out=[]
    for v in sorted(set(round(float(x),2) for x in vals)):
        if out and v-out[-1]<=tol: out[-1]=(out[-1]+v)/2
        else: out.append(v)
    return sorted(out)

def vline_slots(page):
    xs=[float(l["x0"]) for l in page.lines
        if abs(l["x0"]-l["x1"])<1.5 and abs(l["y0"]-l["y1"])>LINE_MIN]
    xs+=[0.0,float(page.width)]
    return snap(xs,4)

def text_slots(all_segs,pw):
    xs={0.0,float(pw)}
    for segs in all_segs:
        for s in segs: xs.add(s["x0"])
    return snap(list(xs),6)

def slot_of(x,slots):
    best=0
    for i in range(len(slots)-1):
        if slots[i]-3<=x: best=i
    return best

def end_slot(si,segs,slots):
    cs=slot_of(segs[si]["x0"],slots)
    if si+1<len(segs): return max(cs,slot_of(segs[si+1]["x0"],slots)-1)
    return max(cs,slot_of(segs[si]["x1"],slots))

def detect_align(seg,sx0,sx1):
    cw=sx1-sx0
    if cw<5: return "left"
    if abs((seg["x0"]+seg["x1"])/2-(sx0+sx1)/2)<=cw*0.22: return "center"
    if seg["x0"]>sx0+cw*0.52: return "right"
    return "left"

def get_fills(page):
    pw=float(page.width); out=[]
    for r in page.rects:
        c=r.get("non_stroking_color")
        if c is None: continue
        if isinstance(c,(int,float)): c=(float(c),)*3
        if not(isinstance(c,(list,tuple)) and len(c)>=3): continue
        r0,g0,b0=float(c[0]),float(c[1]),float(c[2])
        if   r0<0.20 and g0<0.20 and b0<0.20: kind="black"
        elif r0>0.40 and g0>0.40 and b0>0.40 and not(r0>0.94 and g0>0.94 and b0>0.94): kind="grey"
        else: continue
        if float(r["x1"]-r["x0"])>pw*0.18:
            out.append(dict(top=float(r["top"]),bottom=float(r["bottom"]),kind=kind))
    return out

def fill_at(y,fills,tol=4):
    for f in fills:
        if f["top"]-tol<=y<=f["bottom"]+tol: return f["kind"]
    return None

# ── core line writer ──────────────────────────────────────────────────────────
def write_line(ws,xl_row,segs,slots,n_slots,fk):
    if not segs: rh(ws,xl_row,5); return 1
    sz=segs[0]["size"] if segs else 9

    if fk=="black":
        txt="  "+"  ".join(s["text"] for s in segs)
        c=ws.cell(xl_row,1); wc(c,txt,bold=True,sz=max(int(sz),9),color="FFFFFF",fill="1F1F1F")
        rh(ws,xl_row,max(14,round(sz*1.6)))
        if n_slots>1: mg(ws,xl_row,1,xl_row,n_slots)
        return 1

    if fk=="grey":
        used=set()
        for si,seg in enumerate(segs):
            cs=slot_of(seg["x0"],slots); ce=end_slot(si,segs,slots); col=cs+1
            if col in used: continue
            al=detect_align(seg,slots[cs],slots[min(cs+1,len(slots)-1)])
            c=ws.cell(xl_row,col); wc(c,seg["text"],bold=True,sz=max(int(seg["size"]),8),fill="DCDCDC",align=al)
            if ce>cs: mg(ws,xl_row,col,xl_row,ce+1); [used.add(x) for x in range(col,ce+2)]
            else: used.add(col)
        rh(ws,xl_row,13); return 1

    rh(ws,xl_row,max(11,round(sz*1.65)))
    su=defaultdict(set); pl=[]
    for si,seg in enumerate(segs):
        cs=slot_of(seg["x0"],slots); ce=end_slot(si,segs,slots); sub=0
        while any((cs+1+k) in su[sub] for k in range(ce-cs+1)): sub+=1
        pl.append((sub,cs,ce,seg))
        for k in range(ce-cs+1): su[sub].add(cs+1+k)
    ns=max(p[0] for p in pl)+1
    for sub,cs,ce,seg in pl:
        row=xl_row+sub
        al=detect_align(seg,slots[cs],slots[min(cs+1,len(slots)-1)])
        c=ws.cell(row,cs+1); wc(c,seg["text"],bold=seg["bold"],sz=max(int(seg["size"]),8),align=al)
        rh(ws,row,max(11,round(sz*1.65)))
        if ce>cs: mg(ws,row,cs+1,row,ce+1)
    return ns

# ══════════════════════════════════════════════════════════════════════════════
#  MODE 1 — GRID
# ══════════════════════════════════════════════════════════════════════════════
def write_grid(ws,page):
    slots=vline_slots(page); n=max(len(slots)-1,1)
    fills=get_fills(page); vis=get_rows(page.chars)
    all_segs=[get_segs(chs) for _,chs in vis]
    set_cws(ws,slots)
    xl=1
    for li,(_,__) in enumerate(vis):
        segs=all_segs[li]; fk=fill_at(vis[li][0],fills)
        if not segs: rh(ws,xl,5); xl+=1; continue
        xl+=write_line(ws,xl,segs,slots,n,fk)

# ══════════════════════════════════════════════════════════════════════════════
#  MODE 2 — TABLE
# ══════════════════════════════════════════════════════════════════════════════
def write_table(ws,page,tables):
    pw=float(page.width); fills=get_fills(page)
    vis=get_rows(page.chars); all_segs=[get_segs(chs) for _,chs in vis]
    spans=[(t.bbox[1],t.bbox[3],t) for t in tables]

    def in_tbl(y):
        for tp,bt,_ in spans:
            if tp-5<=y<=bt+5: return True
        return False

    xl=1; done=set()
    for li,(y,_) in enumerate(vis):
        for ti,(tp,bt,t) in enumerate(spans):
            if ti not in done and abs(y-tp)<12:
                done.add(ti)
                data=t.extract()
                if not data: continue
                nc=max((len(r) for r in data if r),default=1)
                tw=t.bbox[2]-t.bbox[0]
                for ci in range(nc):
                    ltr=get_column_letter(ci+1)
                    cw=max(4,round((tw/nc)/5.5))
                    if ws.column_dimensions[ltr].width<cw:
                        ws.column_dimensions[ltr].width=cw
                hdr=True
                for trow in data:
                    if trow is None: xl+=1; continue
                    cells=[str(v).strip() if v is not None else "" for v in trow]
                    if all(c=="" for c in cells): xl+=1; continue
                    rh(ws,xl,14)
                    if hdr:
                        for ci,val in enumerate(cells):
                            c=ws.cell(xl,ci+1); wc(c,val,bold=True,sz=9,align="center",fill="BDD7EE")
                        hdr=False
                    else:
                        for ci,val in enumerate(cells):
                            is_num=bool(re.match(r'^-?[\d,]+\.?\d*$',val.replace(" ","")))
                            c=ws.cell(xl,ci+1); wc(c,val,sz=9,align="right" if is_num else "left")
                    xl+=1
                continue
        if in_tbl(y): continue
        segs=all_segs[li]; fk=fill_at(y,fills); sz=segs[0]["size"] if segs else 9
        if not segs: xl+=1; continue
        rh(ws,xl,max(11,round(sz*1.6)))
        if fk=="black":
            c=ws.cell(xl,1); wc(c,"  "+"  ".join(s["text"] for s in segs),bold=True,sz=max(int(sz),9),color="FFFFFF",fill="1F1F1F"); rh(ws,xl,14)
        elif fk=="grey":
            for ci2,seg in enumerate(segs): c=ws.cell(xl,ci2+1); wc(c,seg["text"],bold=True,sz=8,fill="DCDCDC")
        else:
            nz=max(2,round(pw/150))
            for seg in segs:
                col=max(1,round(seg["x0"]/(pw/nz)))
                c=ws.cell(xl,col)
                while c.value: col+=1; c=ws.cell(xl,col)
                wc(c,seg["text"],bold=seg["bold"],sz=max(int(seg["size"]),8),
                   align="right" if seg["x0"]>pw*0.55 else "left",border=False)
        xl+=1

# ══════════════════════════════════════════════════════════════════════════════
#  MODE 3 — TEXT
# ══════════════════════════════════════════════════════════════════════════════
def write_text(ws,page):
    pw=float(page.width); fills=get_fills(page)
    vis=get_rows(page.chars); all_segs=[get_segs(chs) for _,chs in vis]
    slots=text_slots(all_segs,pw); n=max(len(slots)-1,1)
    set_cws(ws,slots)
    xl=1
    for li,(y,_) in enumerate(vis):
        segs=all_segs[li]; fk=fill_at(y,fills)
        if not segs: rh(ws,xl,5); xl+=1; continue
        xl+=write_line(ws,xl,segs,slots,n,fk)

# ══════════════════════════════════════════════════════════════════════════════
#  PAGE DISPATCHER
# ══════════════════════════════════════════════════════════════════════════════
def process_page(ws,page):
    vlines=[l for l in page.lines if abs(l["x0"]-l["x1"])<1.5 and abs(l["y0"]-l["y1"])>LINE_MIN]
    good=[t for t in page.find_tables()
          if t.extract() and max((len(r) for r in t.extract() if r),default=0)>=TABLE_COLS]
    if len(vlines)>=VLINE_MIN:
        write_grid(ws,page); return "GRID"
    elif good:
        write_table(ws,page,good); return "TABLE"
    else:
        write_text(ws,page); return "TEXT"

# ══════════════════════════════════════════════════════════════════════════════
#  CONVERT
# ══════════════════════════════════════════════════════════════════════════════
def convert(pdf_path:str, out_path:str=None)->str:
    if out_path is None:
        out_path=os.path.splitext(pdf_path)[0]+".xlsx"
    print(f"\n{'═'*58}\n  PDF → Excel  (dynamic — any PDF)\n  In : {os.path.basename(pdf_path)}\n  Out: {out_path}\n{'═'*58}")
    wb=Workbook()
    with pdfplumber.open(pdf_path) as pdf:
        n=len(pdf.pages)
        for pn,page in enumerate(pdf.pages):
            ws=wb.active if pn==0 else wb.create_sheet()
            ws.title=f"Page {pn+1}"
            nv=len([l for l in page.lines if abs(l["x0"]-l["x1"])<1.5 and abs(l["y0"]-l["y1"])>LINE_MIN])
            print(f"  Page {pn+1}/{n} ({page.width:.0f}x{page.height:.0f} chars={len(page.chars)} vlines={nv})",end="  ",flush=True)
            mode=process_page(ws,page)
            ws.freeze_panes="A2"
            print(f"[{mode}] ✓")
    wb.save(out_path)
    print(f"\n  ✓  {n} page(s) → {out_path}\n")
    return out_path

# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════
if __name__=="__main__":
    if len(sys.argv)<2:
        print("Usage: python pdf_to_excel.py  input.pdf  [output.xlsx]")
        print("       python pdf_to_excel.py  *.pdf")
        sys.exit(0)
    inputs=[]; out=None
    for arg in sys.argv[1:]:
        if arg.lower().endswith(".xlsx"): out=arg
        else:
            exp=glob.glob(arg)
            if exp: inputs.extend(sorted(exp))
            elif os.path.exists(arg): inputs.append(arg)
            else: print(f"  Not found: {arg}")
    if not inputs: print("No PDF files."); sys.exit(1)
    if len(inputs)>1: out=None
    ok=fail=0
    for p in inputs:
        try: convert(p,out if len(inputs)==1 else None); ok+=1
        except Exception as e:
            print(f"  ERROR {p}: {e}"); import traceback; traceback.print_exc(); fail+=1
    if len(inputs)>1: print(f"\n  Done: {ok} ok, {fail} failed.")
