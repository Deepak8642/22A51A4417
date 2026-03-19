"""
PDF -> Excel Converter  (Zero System Dependencies)
===================================================
Handles:
  * Structured FORMS  (IRS, invoices, applications — multi-column label/field layouts)
  * Regular DOCUMENTS (reports, articles — headings, paragraphs, data tables)
  * TEXT ALIGNMENT     (left / center / right detected from x-coordinates)
  * IMAGES             (embedded images placed at accurate PDF positions via PyMuPDF)
  * Each PDF page      -> one Excel sheet

INSTALL (one-time, no admin needed):
    pip install pdfplumber pymupdf pillow openpyxl

USAGE:
    python pdf_to_excel.py  input.pdf  output.xlsx
"""

import sys, os, io, re
from collections import defaultdict

# ── pip-only dependencies ──────────────────────────────────────────────────────
try:
    import fitz                              # PyMuPDF -> pip install pymupdf
    PYMUPDF = True
except ImportError:
    PYMUPDF = False

import pdfplumber                            # pip install pdfplumber
from PIL import Image as PILImage            # pip install pillow
from openpyxl import Workbook                # pip install openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
EXCEL_COLS    = 12        # number of Excel columns to map page width into
COL_WIDTH     = 13        # default column width (chars)
ROW_H_DEFAULT = 15        # pts
ROW_H_TITLE   = 30
ROW_H_HEADING = 22
ROW_H_TABLE   = 18
ROW_H_FORM    = 20        # form field rows

# Colour palette
C = dict(
    title_fg    = "1F3864",   title_bg   = "D9E2F3",
    heading_fg  = "2F5597",
    body_fg     = "222222",
    section_bg  = "1F3864",   section_fg = "FFFFFF",   # dark Part I header
    hdr_bg      = "2F5597",   hdr_fg     = "FFFFFF",   # table column headers
    total_bg    = "2F5597",   total_fg   = "FFFFFF",
    row_even    = "EBF3FB",   row_odd    = "FFFFFF",
    label_bg    = "F2F2F2",   # form field label background
    field_bg    = "FFFFFF",
    border_col  = "AAAAAA",   border_hdr = "2F5597",
)

def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

THIN = Side(style="thin",   color=C["border_col"])
MED  = Side(style="medium", color=C["border_hdr"])
NO   = Side(style=None)

def bdr(t=THIN, b=THIN, l=THIN, r=THIN):
    return Border(top=t, bottom=b, left=l, right=r)

HDR_BDR  = bdr(MED, MED, MED, MED)
DATA_BDR = bdr()
FORM_BDR = bdr(THIN, THIN, THIN, THIN)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION HEADER DETECTION
# ══════════════════════════════════════════════════════════════════════════════
SECTION_RE = re.compile(
    r'^(Part\s+[IVXLCDM\d]+|Section\s+\d+|PART\s+\d+)',
    re.IGNORECASE
)

def is_section_header(text):
    """Detect 'Part I', 'Part II', 'Section 1' style headers."""
    return bool(SECTION_RE.match(text.strip()))

# ══════════════════════════════════════════════════════════════════════════════
# TEXT ALIGNMENT DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def detect_alignment(words, page_width, margin_frac=0.10):
    """
    Reads x-coordinates of words on a line and returns 'left'/'center'/'right'.
      - If text midpoint is within ±10% of page centre  -> center
      - If text starts past 55% of page width           -> right
      - Otherwise                                        -> left
    """
    if not words:
        return "left"
    x_min   = min(float(w["x0"]) for w in words)
    x_max   = max(float(w["x1"]) for w in words)
    mid_txt = (x_min + x_max) / 2.0
    mid_pg  = page_width / 2.0
    if abs(mid_txt - mid_pg) <= page_width * margin_frac:
        return "center"
    if x_min > page_width * 0.55:
        return "right"
    return "left"

def avg_font_size(words):
    sizes = [float(w.get("size") or 10) for w in words if w.get("size")]
    return sum(sizes) / len(sizes) if sizes else 10.0

# ══════════════════════════════════════════════════════════════════════════════
# FORM vs DOCUMENT DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def is_form_page(page):
    """
    Heuristic: if the page has many narrow columns (>6) and many rows,
    it's likely a structured form rather than a document.
    Also triggered if many words start with digit+letter (field labels like '1a', '3b').
    """
    tables = page.find_tables()
    for tbl in tables:
        data = tbl.extract()
        if data and len(data[0]) > 6:          # wide multi-column table
            return True

    words = page.extract_words()
    field_labels = sum(1 for w in words
                       if re.match(r'^\d[a-z]$', w['text'].strip().lower()))
    return field_labels >= 3

# ══════════════════════════════════════════════════════════════════════════════
# IMAGE EXTRACTION  (PyMuPDF — pure pip, no system tools)
# ══════════════════════════════════════════════════════════════════════════════
def extract_images(pdf_path):
    """
    Returns list of dicts: {page, index, x0, y0, x1, y1, pil_image}
    Uses PyMuPDF (fitz). Falls back gracefully if not installed.
    """
    if not PYMUPDF:
        print("  [INFO] PyMuPDF not found — images skipped. Run: pip install pymupdf")
        return []

    results = []
    doc = fitz.open(pdf_path)
    for page_num in range(len(doc)):
        page = doc[page_num]
        for img_idx, img_info in enumerate(page.get_images(full=True)):
            xref = img_info[0]
            try:
                base   = doc.extract_image(xref)
                pil    = PILImage.open(io.BytesIO(base["image"])).convert("RGBA")
                rects  = page.get_image_rects(xref)
                r      = rects[0] if rects else fitz.Rect(0, 0, pil.width, pil.height)
                results.append(dict(
                    page=page_num, index=img_idx,
                    x0=r.x0, y0=r.y0, x1=r.x1, y1=r.y1,
                    pil_image=pil,
                ))
            except Exception as e:
                print(f"  [WARN] Image xref={xref} skipped: {e}")
    doc.close()
    return results

def pil_to_xl(pil_img, w_px, h_px):
    """Flatten + resize PIL image -> openpyxl Image object."""
    w_px, h_px = max(w_px, 20), max(h_px, 20)
    bg = PILImage.new("RGB", pil_img.size, (255, 255, 255))
    mask = pil_img.split()[3] if pil_img.mode == "RGBA" else None
    bg.paste(pil_img.convert("RGB"), mask=mask)
    bg = bg.resize((w_px, h_px), PILImage.LANCZOS)
    buf = io.BytesIO()
    bg.save(buf, "PNG")
    buf.seek(0)
    xl = XLImage(buf)
    xl.width, xl.height = w_px, h_px
    return xl

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def in_table_area(y, bboxes, tol=2):
    return any(top - tol <= y <= bot + tol for (_, top, _, bot) in bboxes)

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, height=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border
    if height: ws.row_dimensions[row].height = height
    return cell

def merge(ws, r1, c1, r2, c2):
    if r1 != r2 or c1 != c2:
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)

# ══════════════════════════════════════════════════════════════════════════════
# FORM PAGE RENDERER
# Treats the pdfplumber table as a grid; maps cells to Excel proportionally.
# ══════════════════════════════════════════════════════════════════════════════
def render_form_page(ws, page, page_images, EXCEL_COLS, page_w, page_h):
    """Write a form-style page to the worksheet."""

    current_row = 1

    tables = page.find_tables()
    if not tables:
        # Fallback to document rendering if no tables found
        return render_doc_page(ws, page, page_images, EXCEL_COLS, page_w, page_h)

    for tbl_obj in tables:
        data      = tbl_obj.extract()
        tbl_bbox  = tbl_obj.bbox          # (x0, top, x1, bottom)
        tbl_cols  = max(len(r) for r in data) if data else 1

        # Map PDF table columns -> Excel columns proportionally
        # We give each logical PDF column an equal slice of EXCEL_COLS
        xl_col_count = EXCEL_COLS
        col_span = max(1, xl_col_count // tbl_cols)

        for r_idx, row in enumerate(data):
            # Gather non-None cells
            filled = [(c_idx, str(v).strip()) for c_idx, v in enumerate(row) if v and str(v).strip()]

            if not filled:
                current_row += 1
                continue

            # ── Detect row type ──────────────────────────────────────────────
            first_val   = filled[0][1] if filled else ""
            one_cell    = len(filled) == 1
            full_span   = one_cell and filled[0][0] == 0

            is_section  = full_span and is_section_header(first_val)
            is_hdr_row  = full_span and not is_section and r_idx == 0
            is_data_row = not is_section and not full_span

            # ── Section header (Part I, Part II …) ──────────────────────────
            if is_section:
                cell = set_cell(
                    ws, current_row, 1, first_val,
                    font   = Font(name="Arial", bold=True, size=10,
                                  color=C["section_fg"]),
                    fill   = solid(C["section_bg"]),
                    align  = Alignment(horizontal="left", vertical="center"),
                    border = bdr(MED, MED, MED, MED),
                    height = ROW_H_HEADING,
                )
                merge(ws, current_row, 1, current_row, EXCEL_COLS)
                current_row += 1
                continue

            # ── Full-width non-section row (long instructions, etc.) ─────────
            if full_span and not is_section:
                text = first_val
                # Is it a column-header row?
                bold_it = (r_idx == 0 and len(data) > 3)
                cell = set_cell(
                    ws, current_row, 1, text,
                    font   = Font(name="Arial", bold=bold_it, size=9,
                                  color=C["hdr_fg"] if bold_it else C["body_fg"]),
                    fill   = solid(C["hdr_bg"]) if bold_it else solid(C["label_bg"]),
                    align  = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True),
                    border = FORM_BDR,
                    height = ROW_H_FORM,
                )
                merge(ws, current_row, 1, current_row, EXCEL_COLS)
                current_row += 1
                continue

            # ── Multi-column form row ────────────────────────────────────────
            # Place each filled cell at its proportional Excel column position
            placed = []  # (xl_col_start, xl_col_end, text)

            for i, (c_idx, text) in enumerate(filled):
                xl_start = 1 + c_idx * col_span
                # End column = start of next filled cell - 1  (or end of row)
                if i + 1 < len(filled):
                    next_c   = filled[i + 1][0]
                    xl_end   = next_c * col_span   # one before next start
                else:
                    xl_end = EXCEL_COLS
                xl_end = max(xl_start, xl_end)
                placed.append((xl_start, xl_end, text))

            for xl_start, xl_end, text in placed:
                is_lbl = bool(re.match(r'^\d+[a-z]?\s', text) or
                              text.startswith("$"))
                bg     = solid(C["label_bg"]) if is_lbl else solid(C["field_bg"])

                # Detect if this sub-cell text is itself a column header
                is_col_hdr = (r_idx == 0 and len(data) > 5)

                cell = set_cell(
                    ws, current_row, xl_start, text,
                    font   = Font(name="Arial",
                                  bold = is_col_hdr or is_lbl,
                                  size = 9,
                                  color = C["hdr_fg"] if is_col_hdr else C["body_fg"]),
                    fill   = solid(C["hdr_bg"]) if is_col_hdr else bg,
                    align  = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True),
                    border = FORM_BDR,
                    height = ROW_H_FORM,
                )
                merge(ws, current_row, xl_start, current_row, xl_end)

            current_row += 1

    # ── Place images ─────────────────────────────────────────────────────────
    _place_images(ws, page_images, page_w, page_h, current_row, EXCEL_COLS)


# ══════════════════════════════════════════════════════════════════════════════
# DOCUMENT PAGE RENDERER
# Handles headings, body text, and data tables (reports, articles)
# ══════════════════════════════════════════════════════════════════════════════
def render_doc_page(ws, page, page_images, EXCEL_COLS, page_w, page_h):
    """Write a document-style page to the worksheet."""

    current_row = 1

    tbl_objects  = page.find_tables()
    table_bboxes = [tuple(t.bbox) for t in tbl_objects]

    # Group words into text lines, skipping inside-table words
    words    = page.extract_words(keep_blank_chars=True, x_tolerance=3, y_tolerance=3)
    line_map = defaultdict(list)
    for w in words:
        y = float(w["top"])
        if not in_table_area(y, table_bboxes):
            line_map[round(y / 4) * 4].append(w)

    # Build sorted event list
    events = []
    for y_key, wds in line_map.items():
        events.append((y_key, "text", wds))
    for ti, tbl_obj in enumerate(tbl_objects):
        events.append((tbl_obj.bbox[1], "table", (ti, tbl_obj)))
    for ii, img in enumerate(page_images):
        events.append((img["y0"], "image", (ii, img)))
    events.sort(key=lambda e: e[0])

    written_tables = set()
    written_images = set()

    for pdf_y, etype, edata in events:

        # ── TEXT LINE ────────────────────────────────────────────────────────
        if etype == "text":
            wds = sorted(edata, key=lambda w: float(w["x0"]))
            text = " ".join(w["text"] for w in wds).strip()
            if not text:
                continue

            alignment  = detect_alignment(wds, page_w)
            size       = avg_font_size(wds)
            is_title   = size >= 16
            is_heading = 11 <= size < 16
            is_section = is_section_header(text)

            cell = ws.cell(row=current_row, column=1, value=text)
            merge(ws, current_row, 1, current_row, EXCEL_COLS)
            al = Alignment(horizontal=alignment, vertical="center", wrap_text=True)

            if is_section:
                cell.font      = Font(name="Arial", bold=True, size=10, color=C["section_fg"])
                cell.fill      = solid(C["section_bg"])
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = bdr(MED, MED, MED, MED)
                ws.row_dimensions[current_row].height = ROW_H_HEADING
            elif is_title:
                cell.font      = Font(name="Arial", bold=True, size=16, color=C["title_fg"])
                cell.fill      = solid(C["title_bg"])
                cell.alignment = al
                ws.row_dimensions[current_row].height = ROW_H_TITLE
            elif is_heading:
                cell.font      = Font(name="Arial", bold=True, size=12, color=C["heading_fg"])
                cell.alignment = al
                ws.row_dimensions[current_row].height = ROW_H_HEADING
            else:
                cell.font      = Font(name="Arial", size=10, color=C["body_fg"])
                cell.alignment = al
                ws.row_dimensions[current_row].height = ROW_H_DEFAULT

            current_row += 1

        # ── DATA TABLE ───────────────────────────────────────────────────────
        elif etype == "table":
            ti, tbl_obj = edata
            if ti in written_tables:
                continue
            written_tables.add(ti)

            tbl_data = tbl_obj.extract()
            if not tbl_data:
                continue

            current_row += 1
            num_cols = max(len(r) for r in tbl_data)

            for r_i, row in enumerate(tbl_data):
                is_hdr   = (r_i == 0)
                is_total = (r_i == len(tbl_data) - 1 and
                            any(str(v or "").strip().lower() in ("total","totals","sum")
                                for v in row))
                for c_i in range(num_cols):
                    val  = row[c_i] if c_i < len(row) else ""
                    cell = ws.cell(row=current_row, column=c_i + 1,
                                   value=str(val).strip() if val else "")

                    if is_hdr:
                        cell.font   = Font(name="Arial", bold=True, size=10, color=C["hdr_fg"])
                        cell.fill   = solid(C["hdr_bg"])
                        cell.border = HDR_BDR
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    elif is_total:
                        cell.font   = Font(name="Arial", bold=True, size=10, color=C["total_fg"])
                        cell.fill   = solid(C["total_bg"])
                        cell.border = HDR_BDR
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        bg = C["row_even"] if r_i % 2 == 0 else C["row_odd"]
                        cell.font   = Font(name="Arial", size=10, color="1A1A1A")
                        cell.fill   = solid(bg)
                        cell.border = DATA_BDR
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                ws.row_dimensions[current_row].height = ROW_H_TABLE
                current_row += 1

            current_row += 1

        # ── IMAGE ────────────────────────────────────────────────────────────
        elif etype == "image":
            ii, img = edata
            if ii in written_images:
                continue
            written_images.add(ii)
            current_row = _place_single_image(ws, img, page_w, page_h,
                                               current_row, EXCEL_COLS)

    return current_row


# ══════════════════════════════════════════════════════════════════════════════
# IMAGE PLACEMENT HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def _place_single_image(ws, img, page_w, page_h, current_row, EXCEL_COLS):
    """Place one image into the worksheet at the correct anchor cell."""
    PX_PER_PT = 96 / 72
    w_px = max(int((img["x1"] - img["x0"]) * PX_PER_PT), 30)
    h_px = max(int((img["y1"] - img["y0"]) * PX_PER_PT), 20)
    try:
        xl_img = pil_to_xl(img["pil_image"], w_px, h_px)
        anchor_row = max(1, round((img["y0"] / page_h) * current_row) + 1)
        anchor_col = max(1, round((img["x0"] / page_w) * EXCEL_COLS) + 1)
        ws.add_image(xl_img, f"{get_column_letter(anchor_col)}{anchor_row}")
        rows_needed = round(h_px / ROW_H_DEFAULT) + 2
        end = anchor_row + rows_needed
        while current_row < end:
            ws.row_dimensions[current_row].height = ROW_H_DEFAULT
            current_row += 1
        print(f"    Image -> {get_column_letter(anchor_col)}{anchor_row}  "
              f"({w_px}x{h_px}px)")
    except Exception as e:
        print(f"    Image skipped: {e}")
    return current_row

def _place_images(ws, page_images, page_w, page_h, current_row, EXCEL_COLS):
    """Place all images for a page (used after form rendering)."""
    for img in page_images:
        current_row = _place_single_image(ws, img, page_w, page_h,
                                           current_row, EXCEL_COLS)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def convert(pdf_path, output_path):
    print(f"\nConverting: {pdf_path}")

    all_images = extract_images(pdf_path)
    print(f"  {len(all_images)} image(s) found")

    wb = Workbook()

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):

            # Sheet setup
            ws = wb.active if page_num == 0 else wb.create_sheet()
            ws.title = f"Page {page_num + 1}"

            page_w = float(page.width)
            page_h = float(page.height)

            for c in range(1, EXCEL_COLS + 2):
                ws.column_dimensions[get_column_letter(c)].width = COL_WIDTH

            page_imgs = [im for im in all_images if im["page"] == page_num]

            # Auto-detect: form or document?
            if is_form_page(page):
                print(f"  [Page {page_num+1}] Detected: FORM layout")
                render_form_page(ws, page, page_imgs, EXCEL_COLS, page_w, page_h)
            else:
                print(f"  [Page {page_num+1}] Detected: DOCUMENT layout")
                render_doc_page(ws, page, page_imgs, EXCEL_COLS, page_w, page_h)

    wb.save(output_path)
    print(f"\n  Saved -> {output_path}\n")


# ── CLI ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"ERROR: File not found -> {sys.argv[1]}")
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
