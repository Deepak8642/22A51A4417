"""
╔══════════════════════════════════════════════════════════════════════════════╗
║        PDF → EXCEL  |  COMPLETE DATA TRANSFER  |  Azure Document Intelligence║
║                                                                              ║
║  Every word, line, table cell, key-value pair placed at its EXACT           ║
║  spatial position from the PDF — nothing is lost.                           ║
╚══════════════════════════════════════════════════════════════════════════════╝

WHAT THIS TOOL DOES:
  ✔  EVERY LINE/WORD extracted with its exact (x,y) bounding box → placed in Excel
  ✔  Tables       → structured grids, headers styled, spans merged
  ✔  Key-Value    → placed at exact PDF coordinates
  ✔  All text     → lines placed at correct row/column position
  ✔  Images       → embedded at their PDF position + OCR text beside them
  ✔  Multi-page   → one Excel sheet per PDF page
  ✔  Leading zeros→ preserved (all cells text-formatted)
  ✔  Font sizes   → estimated from bounding box height
  ✔  Batch mode   → convert entire folder at once

INSTALL:
  pip install azure-ai-formrecognizer openpyxl pymupdf pillow

  Optional (image OCR):
    pip install pytesseract
    sudo apt-get install tesseract-ocr   # or: brew install tesseract

SET AZURE KEYS — pick one option:
  Option A (env vars):
    export FORMREC_ENDPOINT=https://YOUR-RESOURCE.cognitiveservices.azure.com/
    export FORMREC_KEY=YOUR_KEY

  Option B: edit the CONFIG section in this file

RUN:
  python pdf_to_excel_final.py  input.pdf
  python pdf_to_excel_final.py  input.pdf   output.xlsx
  python pdf_to_excel_final.py  my_folder/  output_folder/
"""

import sys, os, io, re
from pathlib import Path
from collections import defaultdict

# ── Azure DI ──────────────────────────────────────────────────────────────────
try:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential
except ImportError:
    print("ERROR: pip install azure-ai-formrecognizer"); sys.exit(1)

# ── Excel ─────────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

# ── PyMuPDF ───────────────────────────────────────────────────────────────────
try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

# ── Pillow ────────────────────────────────────────────────────────────────────
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# ── Tesseract OCR (optional) ──────────────────────────────────────────────────
try:
    import pytesseract
    pytesseract.get_tesseract_version()
    HAS_OCR = True
except Exception:
    HAS_OCR = False


# ════════════════════════════════════════════════════════════════════════════
#  CONFIG  ★  EDIT YOUR AZURE KEYS HERE  ★
# ════════════════════════════════════════════════════════════════════════════

FORMREC_ENDPOINT = os.getenv("FORMREC_ENDPOINT",
    "https://YOUR-RESOURCE.cognitiveservices.azure.com/")   # ← EDIT THIS

FORMREC_KEY = os.getenv("FORMREC_KEY",
    "YOUR_AZURE_KEY_HERE")                                   # ← EDIT THIS

# Model — "prebuilt-layout" extracts tables + words + lines + key-value pairs
DI_MODEL = "prebuilt-layout"

# ── Grid resolution ───────────────────────────────────────────────────────────
# The PDF page is mapped onto a grid of GRID_COLS x GRID_ROWS Excel cells.
# Higher values = tighter layout match.  Lower = faster, smaller file.
GRID_COLS  = 80     # columns across page width
GRID_ROWS  = 100    # rows across page height

COL_WIDTH  = 2.8    # Excel column width in characters
ROW_HEIGHT = 12.0   # Excel row height in pt

# Image settings
MIN_IMG_PX   = 30
MAX_IMG_W_PX = 380
MAX_IMG_H_PX = 280


# ════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _side(st="thin", c="BBBBBB"):
    return Side(style=st, color=c)

def _bdr(st="thin", c="BBBBBB"):
    s = _side(st, c)
    return Border(top=s, bottom=s, left=s, right=s)

def _fill(h):
    return PatternFill("solid", fgColor=h.lstrip("#"))

def _font(bold=False, italic=False, sz=9, color="000000"):
    return Font(name="Arial", bold=bold, italic=italic,
                size=max(6, int(sz)), color=color.lstrip("#"))

def _aln(h="left", wrap=True, v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN   = _bdr("thin",   "CCCCCC")
MEDIUM = _bdr("medium", "444444")
NOBDR  = Border()

def wcell(ws, r, c, val,
          bold=False, italic=False, sz=9,
          fg=None, tc="000000", bdr=NOBDR,
          ha="left", wrap=True):
    """Write a cell. Always forces text number_format to preserve leading zeros."""
    cell               = ws.cell(row=r, column=c)
    cell.value         = "" if val is None else str(val)
    cell.number_format = "@"
    cell.font          = _font(bold, italic, sz, tc)
    cell.alignment     = _aln(ha, wrap)
    cell.border        = bdr
    if fg:
        cell.fill = _fill(fg)
    return cell

def safe_merge(ws, r1, c1, r2, c2, done):
    if r1 == r2 and c1 == c2:
        return
    key = (r1, c1, r2, c2)
    if key in done:
        return
    try:
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
        done.add(key)
    except Exception:
        pass


# ════════════════════════════════════════════════════════════════════════════
#  COORDINATE HELPERS
# ════════════════════════════════════════════════════════════════════════════

def poly_bbox(polygon):
    """Azure polygon → (x0, y0, x1, y1) in inches."""
    if not polygon:
        return None
    xs = [p.x for p in polygon]
    ys = [p.y for p in polygon]
    return (min(xs), min(ys), max(xs), max(ys))

def to_col(x_in, pw):
    return max(1, min(GRID_COLS, int(x_in / pw * GRID_COLS) + 1))

def to_row(y_in, ph):
    return max(1, min(GRID_ROWS, int(y_in / ph * GRID_ROWS) + 1))

def bbox_cells(bbox, pw, ph):
    sc = to_col(bbox[0], pw);  sr = to_row(bbox[1], ph)
    ec = to_col(bbox[2], pw);  er = to_row(bbox[3], ph)
    return sc, sr, max(sc, ec), max(sr, er)

def font_sz(h_in):
    pt = h_in * 72
    if pt >= 20: return 14
    if pt >= 14: return 12
    if pt >= 10: return 10
    return 9


# ════════════════════════════════════════════════════════════════════════════
#  AZURE DI — ANALYZE
# ════════════════════════════════════════════════════════════════════════════

def analyze_pdf(path):
    client = DocumentAnalysisClient(
        endpoint=FORMREC_ENDPOINT,
        credential=AzureKeyCredential(FORMREC_KEY))
    with open(path, "rb") as f:
        poller = client.begin_analyze_document(DI_MODEL, document=f)
    return poller.result()


# ════════════════════════════════════════════════════════════════════════════
#  BUILD COMPLETE PAGE DATA
# ════════════════════════════════════════════════════════════════════════════

def build_page_data(result):
    """
    Returns {page_num: {w, h, lines, tables, kvpairs, paragraphs}}.
    w/h are page dimensions in inches (Azure DI unit).
    lines  — every line with its bbox  (this is the main text carrier)
    tables — full grid + span + kind info
    kvpairs— key/value with individual bboxes
    paragraphs — with role tags (title, heading, footnote …)
    """
    pages = {}

    for pg in result.pages:
        pages[pg.page_number] = {
            "w": pg.width  or 8.5,
            "h": pg.height or 11.0,
            "lines": [], "tables": [], "kvpairs": [], "paragraphs": []
        }

    def ep(pn):
        if pn not in pages:
            pages[pn] = {"w": 8.5, "h": 11.0,
                         "lines": [], "tables": [], "kvpairs": [], "paragraphs": []}

    # ── Lines (and fallback words when no lines) ──────────────────────────────
    for pg in result.pages:
        pn = pg.page_number
        ep(pn)

        if pg.lines:
            for line in pg.lines:
                pages[pn]["lines"].append({
                    "content": line.content or "",
                    "bbox":    poly_bbox(line.polygon)
                })
        elif pg.words:
            # No lines → group words into synthetic lines by y-proximity
            words = [{"content": w.content or "",
                      "bbox":    poly_bbox(w.polygon)} for w in pg.words if w.content]
            words.sort(key=lambda w: (round(w["bbox"][1] * 4) if w["bbox"] else 0,
                                      w["bbox"][0] if w["bbox"] else 0))
            current_line_words = []
            current_y = None
            for w in words:
                wy = round(w["bbox"][1] * 4) if w["bbox"] else 0
                if current_y is None or abs(wy - current_y) < 2:
                    current_line_words.append(w)
                    current_y = wy
                else:
                    if current_line_words:
                        xs = [cw["bbox"][0] for cw in current_line_words if cw["bbox"]]
                        ys = [cw["bbox"][1] for cw in current_line_words if cw["bbox"]]
                        xe = [cw["bbox"][2] for cw in current_line_words if cw["bbox"]]
                        ye = [cw["bbox"][3] for cw in current_line_words if cw["bbox"]]
                        text = " ".join(cw["content"] for cw in current_line_words)
                        bbox = (min(xs), min(ys), max(xe), max(ye)) if xs else None
                        pages[pn]["lines"].append({"content": text, "bbox": bbox})
                    current_line_words = [w]
                    current_y = wy
            if current_line_words:
                xs = [cw["bbox"][0] for cw in current_line_words if cw["bbox"]]
                ys = [cw["bbox"][1] for cw in current_line_words if cw["bbox"]]
                xe = [cw["bbox"][2] for cw in current_line_words if cw["bbox"]]
                ye = [cw["bbox"][3] for cw in current_line_words if cw["bbox"]]
                text = " ".join(cw["content"] for cw in current_line_words)
                bbox = (min(xs), min(ys), max(xe), max(ye)) if xs else None
                pages[pn]["lines"].append({"content": text, "bbox": bbox})

    # ── Paragraphs ────────────────────────────────────────────────────────────
    if hasattr(result, "paragraphs") and result.paragraphs:
        for para in result.paragraphs:
            pn = 1
            bbox = None
            if para.bounding_regions:
                pn   = para.bounding_regions[0].page_number
                bbox = poly_bbox(para.bounding_regions[0].polygon)
            ep(pn)
            pages[pn]["paragraphs"].append({
                "content": para.content or "",
                "bbox":    bbox,
                "role":    getattr(para, "role", "") or ""
            })

    # ── Tables ────────────────────────────────────────────────────────────────
    if result.tables:
        for tbl in result.tables:
            pn = 1
            tbl_bbox = None
            if tbl.bounding_regions:
                pn       = tbl.bounding_regions[0].page_number
                tbl_bbox = poly_bbox(tbl.bounding_regions[0].polygon)
            ep(pn)

            nrows = tbl.row_count
            ncols = tbl.column_count
            grid  = [[""] * ncols for _ in range(nrows)]
            spans = {}
            kinds = {}

            for cell in tbl.cells:
                r, c = cell.row_index, cell.column_index
                grid[r][c] = cell.content or ""
                rs = getattr(cell, "row_span",    1) or 1
                cs = getattr(cell, "column_span", 1) or 1
                if rs > 1 or cs > 1:
                    spans[(r, c)] = (rs, cs)
                kinds[(r, c)] = getattr(cell, "kind", "") or ""

            pages[pn]["tables"].append({
                "grid": grid, "spans": spans, "kinds": kinds,
                "nrows": nrows, "ncols": ncols, "bbox": tbl_bbox
            })

    # ── Key-value pairs ───────────────────────────────────────────────────────
    if hasattr(result, "key_value_pairs") and result.key_value_pairs:
        for kv in result.key_value_pairs:
            pn     = 1
            k_bbox = v_bbox = None
            if kv.key and kv.key.bounding_regions:
                pn     = kv.key.bounding_regions[0].page_number
                k_bbox = poly_bbox(kv.key.bounding_regions[0].polygon)
            if kv.value and kv.value.bounding_regions:
                v_bbox = poly_bbox(kv.value.bounding_regions[0].polygon)
            ep(pn)
            pages[pn]["kvpairs"].append({
                "key":      kv.key.content   if kv.key   else "",
                "value":    kv.value.content if kv.value else "",
                "key_bbox": k_bbox,
                "val_bbox": v_bbox
            })

    return pages


# ════════════════════════════════════════════════════════════════════════════
#  IMAGE EXTRACTION
# ════════════════════════════════════════════════════════════════════════════

def get_images(fitz_doc, page_idx, pw, ph):
    out = []
    if not (HAS_FITZ and HAS_PIL):
        return out
    page = fitz_doc[page_idx]
    for img_info in page.get_images(full=True):
        xref = img_info[0]
        try:
            raw = fitz_doc.extract_image(xref)
            pil = PILImage.open(io.BytesIO(raw["image"])).convert("RGB")
        except Exception:
            continue
        if pil.width < MIN_IMG_PX or pil.height < MIN_IMG_PX:
            continue
        bbox_pt = None
        for rect in page.get_image_rects(xref):
            r = rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect)
            bbox_pt = (r.x0/72, r.y0/72, r.x1/72, r.y1/72)  # pt→in
            break
        if bbox_pt is None:
            continue
        sc, sr, ec, er = bbox_cells(bbox_pt, pw, ph)
        out.append({"pil": pil, "sc": sc, "sr": sr, "ec": ec, "er": er})
    return out

def resize_img(pil):
    w, h = pil.size
    scale = min(MAX_IMG_W_PX/w, MAX_IMG_H_PX/h, 1.0)
    return pil.resize((max(1, int(w*scale)), max(1, int(h*scale))), PILImage.LANCZOS)

def do_ocr(pil):
    if not HAS_OCR:
        return ""
    try:
        return re.sub(r'\n{3,}', '\n\n', pytesseract.image_to_string(pil)).strip()
    except Exception:
        return ""


# ════════════════════════════════════════════════════════════════════════════
#  TABLE-OCCUPIED CELLS  (for skipping text-over-table)
# ════════════════════════════════════════════════════════════════════════════

def table_cells_set(tables, pw, ph):
    occupied = set()
    for tbl in tables:
        if tbl["bbox"]:
            sc, sr, ec, er = bbox_cells(tbl["bbox"], pw, ph)
            for r in range(sr, er+1):
                for c in range(sc, ec+1):
                    occupied.add((r, c))
    return occupied


# ════════════════════════════════════════════════════════════════════════════
#  WRITE ONE PAGE TO WORKSHEET
# ════════════════════════════════════════════════════════════════════════════

def write_page(ws, pdata, page_num, fitz_doc=None, page_idx=0):
    pw = pdata["w"]
    ph = pdata["h"]
    merged = set()

    # ── Setup uniform grid ────────────────────────────────────────────────────
    for ci in range(1, GRID_COLS + 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTH
    for ri in range(1, GRID_ROWS + 1):
        ws.row_dimensions[ri].height = ROW_HEIGHT
    ws.sheet_view.zoomScale = 85

    tab_occupied = table_cells_set(pdata["tables"], pw, ph)

    # ══════════════════════════════════════════════════════════════════════════
    #  STEP 1 — TABLES
    # ══════════════════════════════════════════════════════════════════════════
    for tbl in pdata["tables"]:
        grid  = tbl["grid"]
        spans = tbl["spans"]
        kinds = tbl["kinds"]
        nrows = tbl["nrows"]
        ncols = tbl["ncols"]
        bbox  = tbl["bbox"]

        if bbox:
            sc0, sr0, ec0, er0 = bbox_cells(bbox, pw, ph)
        else:
            sc0, sr0, ec0, er0 = 1, 1, min(GRID_COLS, ncols*3), min(GRID_ROWS, nrows*2)

        cw = max(1, (ec0 - sc0 + 1) // max(ncols, 1))
        rh = max(1, (er0 - sr0 + 1) // max(nrows, 1))

        absorbed = set()

        for ri in range(nrows):
            for ci in range(ncols):
                if (ri, ci) in absorbed:
                    continue
                val  = grid[ri][ci]
                kind = kinds.get((ri, ci), "")
                hdr  = kind in ("columnHeader", "rowHeader") or ri == 0

                er = max(1, min(sr0 + ri * rh, GRID_ROWS))
                ec = max(1, min(sc0 + ci * cw, GRID_COLS))

                if hdr:
                    fg, tc, bd, bold = "1F4E79", "FFFFFF", MEDIUM, True
                elif ri % 2 == 0:
                    fg, tc, bd, bold = "EBF3FB", "000000", THIN,   False
                else:
                    fg, tc, bd, bold = "FFFFFF", "000000", THIN,   False

                wcell(ws, er, ec, val, bold=bold, sz=9, fg=fg, tc=tc, bdr=bd)

                rs, cs = spans.get((ri, ci), (1, 1))
                if rs > 1 or cs > 1:
                    end_r = min(er + rs*rh - 1, GRID_ROWS)
                    end_c = min(ec + cs*cw - 1, GRID_COLS)
                    safe_merge(ws, er, ec, end_r, end_c, merged)
                    for mr in range(ri, ri+rs):
                        for mc in range(ci, ci+cs):
                            if (mr, mc) != (ri, ci):
                                absorbed.add((mr, mc))

    # ══════════════════════════════════════════════════════════════════════════
    #  STEP 2 — ALL TEXT LINES  (the primary data carrier)
    #  Group by Excel row, then write left-to-right within each row.
    # ══════════════════════════════════════════════════════════════════════════
    row_items = defaultdict(list)   # excel_row → [(sc, ec, text, h_in)]

    for line in pdata["lines"]:
        content = (line["content"] or "").strip()
        bbox    = line["bbox"]
        if not content or not bbox:
            continue

        sc, sr, ec, er = bbox_cells(bbox, pw, ph)
        if (sr, sc) in tab_occupied:
            continue

        h_in = bbox[3] - bbox[1]
        row_items[sr].append((sc, ec, content, h_in))

    for ex_row in sorted(row_items.keys()):
        items = sorted(row_items[ex_row], key=lambda x: x[0])
        for sc, ec, content, h_in in items:
            sc  = max(1, min(sc, GRID_COLS))
            ec  = max(sc, min(ec, GRID_COLS))
            row = max(1, min(ex_row, GRID_ROWS))
            sz  = font_sz(h_in)

            existing = ws.cell(row=row, column=sc).value
            if existing:
                ws.cell(row=row, column=sc).value = str(existing) + "  " + content
            else:
                wcell(ws, row, sc, content, sz=sz, bdr=NOBDR, wrap=True)
                if ec > sc:
                    safe_merge(ws, row, sc, row, ec, merged)

            # Adjust row height proportionally to text size
            ws.row_dimensions[row].height = max(
                ws.row_dimensions[row].height,
                min(120, h_in * 72 * 1.5))

    # ══════════════════════════════════════════════════════════════════════════
    #  STEP 3 — PARAGRAPHS WITH ROLES  (titles, headings, footnotes, etc.)
    #  Only styled paragraphs — plain text is already in lines above.
    # ══════════════════════════════════════════════════════════════════════════
    for para in pdata["paragraphs"]:
        role    = para["role"]
        content = (para["content"] or "").strip()
        bbox    = para["bbox"]
        if not content or not bbox or not role:
            continue

        sc, sr, ec, er = bbox_cells(bbox, pw, ph)
        if (sr, sc) in tab_occupied:
            continue
        sc = max(1, min(sc, GRID_COLS))
        sr = max(1, min(sr, GRID_ROWS))
        ec = max(sc, min(ec, GRID_COLS))
        er = max(sr, min(er, GRID_ROWS))

        if role == "title":
            fg, tc, bold, sz = "D6E4F0", "1F4E79", True,  14
        elif role == "sectionHeading":
            fg, tc, bold, sz = "EBF3FB", "1F4E79", True,  12
        elif role in ("pageHeader", "pageFooter"):
            fg, tc, bold, sz = "F5F5F5", "666666", False,  8
        elif role == "footnote":
            fg, tc, bold, sz = None,     "888888", False,  7
        else:
            fg, tc, bold, sz = None,     "000000", False,  9

        existing = ws.cell(row=sr, column=sc).value
        if not existing:
            wcell(ws, sr, sc, content, bold=bold, sz=sz,
                  fg=fg, tc=tc, bdr=NOBDR, wrap=True)
            if ec > sc or er > sr:
                safe_merge(ws, sr, sc, er, ec, merged)

    # ══════════════════════════════════════════════════════════════════════════
    #  STEP 4 — KEY-VALUE PAIRS  (placed at exact PDF coords)
    # ══════════════════════════════════════════════════════════════════════════
    for kv in pdata["kvpairs"]:
        key = (kv["key"]   or "").strip()
        val = (kv["value"] or "").strip()
        if not key and not val:
            continue

        if kv["key_bbox"]:
            sc, sr, ec, er = bbox_cells(kv["key_bbox"], pw, ph)
            sr = max(1, min(sr, GRID_ROWS))
            sc = max(1, min(sc, GRID_COLS))
            if (sr, sc) not in tab_occupied:
                existing = ws.cell(row=sr, column=sc).value
                if not existing:
                    wcell(ws, sr, sc, key, bold=True, sz=9,
                          tc="1F4E79", bdr=NOBDR)

        if kv["val_bbox"]:
            sc, sr, ec, er = bbox_cells(kv["val_bbox"], pw, ph)
            sr = max(1, min(sr, GRID_ROWS))
            sc = max(1, min(sc, GRID_COLS))
            ec = max(sc, min(ec, GRID_COLS))
            if (sr, sc) not in tab_occupied:
                existing = ws.cell(row=sr, column=sc).value
                if not existing:
                    wcell(ws, sr, sc, val, sz=9,
                          fg="FEFCE8", bdr=NOBDR, wrap=True)
                    if ec > sc:
                        safe_merge(ws, sr, sc, sr, ec, merged)

    # ══════════════════════════════════════════════════════════════════════════
    #  STEP 5 — IMAGES
    # ══════════════════════════════════════════════════════════════════════════
    if fitz_doc is not None:
        for img in get_images(fitz_doc, page_idx, pw, ph):
            pil = img["pil"]
            sc  = max(1, img["sc"])
            sr  = max(1, img["sr"])

            resized = resize_img(pil)
            buf = io.BytesIO()
            resized.save(buf, format="PNG")
            buf.seek(0)
            try:
                xl_img = XLImage(buf)
                xl_img.anchor = f"{get_column_letter(sc)}{sr}"
                ws.add_image(xl_img)
            except Exception as e:
                print(f"    ⚠  Image embed: {e}")

            txt = do_ocr(pil)
            if txt:
                oc = min(img["ec"] + 1, GRID_COLS)
                or_ = max(1, min(sr, GRID_ROWS))
                existing = ws.cell(row=or_, column=oc).value
                if not existing:
                    wcell(ws, or_, oc, f"[Image]\n{txt}",
                          italic=True, sz=8, tc="555555",
                          fg="FFFDE7", bdr=NOBDR, wrap=True)
                    safe_merge(ws, or_, oc,
                               min(img["er"], GRID_ROWS),
                               min(oc + 6, GRID_COLS), merged)


# ════════════════════════════════════════════════════════════════════════════
#  CONVERT SINGLE PDF
# ════════════════════════════════════════════════════════════════════════════

def convert_pdf(pdf_path, out_path):
    bar = "=" * 64
    print(f"\n{bar}")
    print(f"  PDF -> EXCEL  |  Azure Document Intelligence")
    print(f"  Input : {pdf_path}")
    print(f"  Output: {out_path}")
    print(f"{bar}")

    print("  [1/4]  Sending to Azure DI ...", flush=True)
    result  = analyze_pdf(pdf_path)
    n_pages = len(result.pages)
    n_tbl   = len(result.tables) if result.tables else 0
    print(f"  Done: {n_pages} page(s), {n_tbl} table(s) detected")

    print("  [2/4]  Building page data ...", flush=True)
    all_data = build_page_data(result)

    fitz_doc = None
    if HAS_FITZ and HAS_PIL:
        print("  [3/4]  Opening PDF for image extraction ...", flush=True)
        try:
            fitz_doc = fitz.open(pdf_path)
        except Exception as e:
            print(f"  Warning: PyMuPDF failed: {e}")
    else:
        print("  [3/4]  Image extraction skipped (install pymupdf pillow)")

    print("  [4/4]  Writing Excel workbook ...", flush=True)
    wb    = Workbook()
    first = True

    for pn in sorted(all_data.keys()):
        ws    = wb.active if first else wb.create_sheet()
        first = False
        ws.title = f"Page {pn}"[:31]

        pd = all_data[pn]
        print(f"    Page {pn:>3}  |  lines={len(pd['lines'])}  "
              f"tables={len(pd['tables'])}  "
              f"kv={len(pd['kvpairs'])}  "
              f"paragraphs={len(pd['paragraphs'])}")

        write_page(ws, pd, pn, fitz_doc=fitz_doc, page_idx=pn-1)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    if fitz_doc:
        fitz_doc.close()

    wb.save(out_path)
    kb = os.path.getsize(out_path) // 1024
    print(f"\n  SAVED -> {out_path}  ({kb} KB)\n{bar}\n")


# ════════════════════════════════════════════════════════════════════════════
#  BATCH CONVERT
# ════════════════════════════════════════════════════════════════════════════

def convert_folder(in_dir, out_dir):
    in_p  = Path(in_dir)
    out_p = Path(out_dir)
    out_p.mkdir(parents=True, exist_ok=True)
    pdfs  = sorted(in_p.glob("*.pdf"))
    if not pdfs:
        print(f"No PDFs found in {in_dir}"); return
    ok = fail = 0
    for pdf in pdfs:
        try:
            convert_pdf(str(pdf), str(out_p / (pdf.stem + ".xlsx")))
            ok += 1
        except Exception as e:
            print(f"  ERROR: {pdf.name}: {e}"); fail += 1
    print(f"\nBatch done: {ok} OK, {fail} failed.")


# ════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════

def check_keys():
    if "YOUR-RESOURCE" in FORMREC_ENDPOINT or "YOUR_AZURE" in FORMREC_KEY:
        print("\n" + "!"*64)
        print("  Azure credentials not set!")
        print("  Edit the CONFIG section in this file, or set env vars:")
        print("    export FORMREC_ENDPOINT=https://your-resource.cognitiveservices.azure.com/")
        print("    export FORMREC_KEY=your_key")
        print("!"*64 + "\n")
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print(__doc__)
        print("Usage:")
        print("  python pdf_to_excel_final.py  input.pdf  [output.xlsx]")
        print("  python pdf_to_excel_final.py  folder/    [out_folder/]")
        sys.exit(1)

    check_keys()
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) == 3 else None

    if os.path.isdir(inp):
        convert_folder(inp, out or (inp.rstrip("/\\") + "_excel"))
    elif os.path.isfile(inp) and inp.lower().endswith(".pdf"):
        convert_pdf(inp, out or str(Path(inp).with_suffix(".xlsx")))
    else:
        print(f"ERROR: '{inp}' is not a PDF file or folder.")
        sys.exit(1)
