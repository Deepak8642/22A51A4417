"""
pdf_to_excel.py  —  PDF → Excel via Azure Document Intelligence
================================================================
Usage:
  python pdf_to_excel.py input.pdf
  python pdf_to_excel.py input.pdf output.xlsx
  python pdf_to_excel.py my_folder/ output_folder/

Set your Azure keys as environment variables (recommended):
  export FORMREC_ENDPOINT=https://your-resource.cognitiveservices.azure.com/
  export FORMREC_KEY=your_key_here

  -- OR --

Edit the CONFIG section below directly.
"""

import sys, os, io, re
from pathlib import Path
from collections import defaultdict

# ── Azure DI ──────────────────────────────────────────────────────────────
try:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential
except ImportError:
    print("ERROR: pip install azure-ai-formrecognizer"); sys.exit(1)

# ── Excel ─────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

# ── PyMuPDF ───────────────────────────────────────────────────────────────
try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False
    print("WARNING: pymupdf not installed — images will be skipped. pip install pymupdf")

# ── Pillow ────────────────────────────────────────────────────────────────
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("WARNING: pillow not installed — images will be skipped. pip install pillow")

# ── Tesseract OCR (optional) ──────────────────────────────────────────────
try:
    import pytesseract
    pytesseract.get_tesseract_version()
    HAS_OCR = True
except Exception:
    HAS_OCR = False


# ════════════════════════════════════════════════════════════════════════════
# ★  CONFIG — Edit your Azure keys here OR set environment variables  ★
# ════════════════════════════════════════════════════════════════════════════

# Read from environment first, fall back to hard-coded values below
FORMREC_ENDPOINT = os.environ.get(
    "FORMREC_ENDPOINT",
    "https://YOUR-RESOURCE.cognitiveservices.azure.com/"   # ← replace if not using env var
)
FORMREC_KEY = os.environ.get(
    "FORMREC_KEY",
    "YOUR_AZURE_KEY_HERE"                                  # ← replace if not using env var
)

# "prebuilt-layout" gives the best results: tables + text + checkboxes + KV pairs
DI_MODEL = "prebuilt-layout"

# Grid resolution: how many Excel columns/rows map across one full PDF page.
# Higher = tighter spatial match.  Lower = faster, smaller file.
GRID_COLS = 80
GRID_ROWS = 100

COL_WIDTH  = 2.8    # Excel column width (characters)
ROW_HEIGHT = 12.0   # Excel row height (points)

# Image sizing limits (pixels)
MIN_IMG_PX   = 30
MAX_IMG_W_PX = 380
MAX_IMG_H_PX = 280

# Checkbox characters — Azure DI emits these tokens for form checkboxes
CHECKBOX_SELECTED   = ":selected:"
CHECKBOX_UNSELECTED = ":unselected:"
CHECKBOX_CHECKED_CHAR   = "☑"   # rendered as checked box in Excel
CHECKBOX_UNCHECKED_CHAR = "☐"   # rendered as empty box in Excel


# ════════════════════════════════════════════════════════════════════════════
# STYLE HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _side(st="thin", c="BBBBBB"):
    return Side(style=st, color=c)

def _bdr(st="thin", c="BBBBBB"):
    s = _side(st, c)
    return Border(top=s, bottom=s, left=s, right=s)

def _fill(h):
    return PatternFill("solid", fgColor=h.lstrip("#"))

def _font(bold=False, italic=False, size=9, color="000000"):
    return Font(name="Arial", bold=bold, italic=italic,
                size=max(6, int(size)), color=color.lstrip("#"))

def _aln(h="left", wrap=True, v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN   = _bdr("thin",   "CCCCCC")
MEDIUM = _bdr("medium", "444444")
NOBDR  = Border()


def is_merged(ws, row, col):
    """Return True if the cell at (row, col) is part of a merged region (slave cell)."""
    return isinstance(ws.cell(row=row, column=col), MergedCell)


def normalize_checkboxes(text):
    """
    Replace Azure DI checkbox tokens with readable Unicode characters.
    Also handles inline occurrences such as ':selected: Yes'.
    """
    if not text:
        return text
    text = text.replace(CHECKBOX_SELECTED,   CHECKBOX_CHECKED_CHAR)
    text = text.replace(CHECKBOX_UNSELECTED, CHECKBOX_UNCHECKED_CHAR)
    return text


def wcell(ws, row, col, val,
          bold=False, italic=False, size=9,
          fg=None, tc="000000", bdr=NOBDR,
          ha="left", wrap=True, grid_rows=None, grid_cols=None):
    """
    Write value + formatting to a cell.
    - Normalises checkbox tokens in the value.
    - Skips silently if cell is a MergedCell (read-only slave cell).
    - Always forces text number_format (@) to preserve leading zeros.
    - grid_rows / grid_cols: clamp check uses these if provided.
    """
    gr = grid_rows or GRID_ROWS
    gc = grid_cols or GRID_COLS

    if row < 1 or row > gr or col < 1 or col > gc:
        return None
    if is_merged(ws, row, col):
        return None

    cell               = ws.cell(row=row, column=col)
    cell.value         = normalize_checkboxes("" if val is None else str(val))
    cell.number_format = "@"
    cell.font          = _font(bold, italic, size, tc)
    cell.alignment     = _aln(ha, wrap)
    cell.border        = bdr
    if fg:
        cell.fill = _fill(fg)
    return cell


def safe_merge(ws, r1, c1, r2, c2, done, grid_rows=None, grid_cols=None):
    """Merge cells safely, tracking already-merged regions to avoid duplicates."""
    gr = grid_rows or GRID_ROWS
    gc = grid_cols or GRID_COLS

    if r1 == r2 and c1 == c2:
        return
    # Clamp to actual grid for this page
    r1 = max(1, min(r1, gr));  r2 = max(1, min(r2, gr))
    c1 = max(1, min(c1, gc));  c2 = max(1, min(c2, gc))
    if r1 == r2 and c1 == c2:
        return
    # Ensure r1 ≤ r2, c1 ≤ c2
    if r1 > r2: r1, r2 = r2, r1
    if c1 > c2: c1, c2 = c2, c1

    key = (r1, c1, r2, c2)
    if key in done:
        return
    try:
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
        done.add(key)
    except Exception:
        pass


# ════════════════════════════════════════════════════════════════════════════
# COORDINATE HELPERS
# ════════════════════════════════════════════════════════════════════════════

def poly_bbox(polygon):
    """Azure DI polygon (list of Point) → (x0, y0, x1, y1) in inches."""
    if not polygon:
        return None
    xs = [p.x for p in polygon]
    ys = [p.y for p in polygon]
    return (min(xs), min(ys), max(xs), max(ys))


def to_col(x_in, pw, gc):
    """x in inches → 1-based Excel column clamped to gc."""
    if pw <= 0:
        return 1
    return max(1, min(gc, int(x_in / pw * gc) + 1))


def to_row(y_in, ph, gr):
    """y in inches → 1-based Excel row clamped to gr."""
    if ph <= 0:
        return 1
    return max(1, min(gr, int(y_in / ph * gr) + 1))


def bbox_to_cells(bbox, pw, ph, gc=None, gr=None):
    """(x0,y0,x1,y1) → (start_col, start_row, end_col, end_row) all 1-based."""
    gc = gc or GRID_COLS
    gr = gr or GRID_ROWS
    sc = to_col(bbox[0], pw, gc);  sr = to_row(bbox[1], ph, gr)
    ec = to_col(bbox[2], pw, gc);  er = to_row(bbox[3], ph, gr)
    return sc, sr, max(sc, ec), max(sr, er)


def font_size_from_height(h_in):
    """Estimate font pt size from bounding-box height in inches."""
    pt = h_in * 72
    if pt >= 20: return 14
    if pt >= 14: return 12
    if pt >= 10: return 10
    return 9


# ════════════════════════════════════════════════════════════════════════════
# AZURE DI — ANALYZE PDF
# ════════════════════════════════════════════════════════════════════════════

def analyze_pdf(path):
    client = DocumentAnalysisClient(
        endpoint=FORMREC_ENDPOINT,
        credential=AzureKeyCredential(FORMREC_KEY)
    )
    with open(path, "rb") as f:
        poller = client.begin_analyze_document(DI_MODEL, document=f)
    return poller.result()


# ════════════════════════════════════════════════════════════════════════════
# BUILD COMPLETE PAGE DATA from AnalyzeResult
# ════════════════════════════════════════════════════════════════════════════

def build_page_data(result):
    """
    Organises all Azure DI content by page number.
    Returns dict: { page_num: {w, h, lines, tables, kvpairs, paragraphs} }
      w/h         — page dimensions in inches
      lines       — every text line with bbox  (PRIMARY text carrier)
      tables      — structured table data
      kvpairs     — key-value pairs
      paragraphs  — paragraphs with role tags (title, heading, footnote …)
    """
    pages = {}

    def ensure(pn):
        if pn not in pages:
            pages[pn] = {"w": 8.5, "h": 11.0,
                         "lines": [], "tables": [], "kvpairs": [], "paragraphs": []}

    # ── Page dimensions ───────────────────────────────────────────────────
    for pg in result.pages:
        ensure(pg.page_number)
        pages[pg.page_number]["w"] = pg.width  or 8.5
        pages[pg.page_number]["h"] = pg.height or 11.0

    # ── Lines (every line of text on every page) ──────────────────────────
    for pg in result.pages:
        pn = pg.page_number
        ensure(pn)

        if pg.lines:
            for line in pg.lines:
                content = normalize_checkboxes((line.content or "").strip())
                if not content:
                    continue
                pages[pn]["lines"].append({
                    "content": content,
                    "bbox":    poly_bbox(line.polygon)
                })

        elif pg.words:
            # No lines returned — group words by y-position into synthetic lines
            words = []
            for w in pg.words:
                if w.content and w.content.strip():
                    bb = poly_bbox(w.polygon)
                    words.append({
                        "content": normalize_checkboxes(w.content.strip()),
                        "bbox":    bb
                    })
            # Sort by row (rounded to nearest 0.1 inch) then column
            words.sort(key=lambda w: (
                round(w["bbox"][1] * 10) if w["bbox"] else 0,
                w["bbox"][0] if w["bbox"] else 0
            ))
            # Group words sharing the same approximate y band (0.15 in tolerance)
            buckets = []
            cur_bucket = []
            cur_y = None
            for w in words:
                wy = w["bbox"][1] if w["bbox"] else 0
                if cur_y is None or abs(wy - cur_y) <= 0.15:
                    cur_bucket.append(w)
                    cur_y = wy if cur_y is None else (cur_y + wy) / 2
                else:
                    buckets.append(cur_bucket)
                    cur_bucket = [w]
                    cur_y = wy
            if cur_bucket:
                buckets.append(cur_bucket)

            for bucket in buckets:
                if not bucket:
                    continue
                valid = [b for b in bucket if b["bbox"]]
                if not valid:
                    continue
                xs = [b["bbox"][0] for b in valid]
                ys = [b["bbox"][1] for b in valid]
                xe = [b["bbox"][2] for b in valid]
                ye = [b["bbox"][3] for b in valid]
                pages[pn]["lines"].append({
                    "content": " ".join(b["content"] for b in bucket),
                    "bbox":    (min(xs), min(ys), max(xe), max(ye))
                })

    # ── Paragraphs (with role tags) ───────────────────────────────────────
    if hasattr(result, "paragraphs") and result.paragraphs:
        for para in result.paragraphs:
            pn   = 1
            bbox = None
            if para.bounding_regions:
                pn   = para.bounding_regions[0].page_number
                bbox = poly_bbox(para.bounding_regions[0].polygon)
            ensure(pn)
            content = normalize_checkboxes((para.content or "").strip())
            if content:
                pages[pn]["paragraphs"].append({
                    "content": content,
                    "bbox":    bbox,
                    "role":    getattr(para, "role", "") or ""
                })

    # ── Tables ────────────────────────────────────────────────────────────
    if result.tables:
        for tbl in result.tables:
            pn       = 1
            tbl_bbox = None
            if tbl.bounding_regions:
                pn       = tbl.bounding_regions[0].page_number
                tbl_bbox = poly_bbox(tbl.bounding_regions[0].polygon)
            ensure(pn)

            nrows = tbl.row_count
            ncols = tbl.column_count
            grid  = [[""] * ncols for _ in range(nrows)]
            spans = {}
            kinds = {}

            for cell in tbl.cells:
                r, c       = cell.row_index, cell.column_index
                grid[r][c] = normalize_checkboxes(cell.content or "")
                rs = getattr(cell, "row_span",    1) or 1
                cs = getattr(cell, "column_span", 1) or 1
                if rs > 1 or cs > 1:
                    spans[(r, c)] = (rs, cs)
                kinds[(r, c)] = getattr(cell, "kind", "") or ""

            pages[pn]["tables"].append({
                "grid": grid, "spans": spans, "kinds": kinds,
                "nrows": nrows, "ncols": ncols, "bbox": tbl_bbox
            })

    # ── Key-value pairs ───────────────────────────────────────────────────
    if hasattr(result, "key_value_pairs") and result.key_value_pairs:
        for kv in result.key_value_pairs:
            pn     = 1
            k_bbox = v_bbox = None
            if kv.key and kv.key.bounding_regions:
                pn     = kv.key.bounding_regions[0].page_number
                k_bbox = poly_bbox(kv.key.bounding_regions[0].polygon)
            if kv.value and kv.value.bounding_regions:
                v_bbox = poly_bbox(kv.value.bounding_regions[0].polygon)
            ensure(pn)
            pages[pn]["kvpairs"].append({
                "key":      normalize_checkboxes((kv.key.content   if kv.key   else "").strip()),
                "value":    normalize_checkboxes((kv.value.content if kv.value else "").strip()),
                "key_bbox": k_bbox,
                "val_bbox": v_bbox
            })

    return pages


# ════════════════════════════════════════════════════════════════════════════
# IMAGE EXTRACTION (PyMuPDF + Pillow)
# ════════════════════════════════════════════════════════════════════════════

def get_images(fitz_doc, page_idx, pw, ph, gc, gr):
    """Extract all images from a PDF page; returns list of positioned dicts."""
    out = []
    if not (HAS_FITZ and HAS_PIL):
        return out
    try:
        page = fitz_doc[page_idx]
    except Exception:
        return out

    for img_info in page.get_images(full=True):
        xref = img_info[0]
        try:
            raw = fitz_doc.extract_image(xref)
            pil = PILImage.open(io.BytesIO(raw["image"])).convert("RGB")
        except Exception:
            continue
        if pil.width < MIN_IMG_PX or pil.height < MIN_IMG_PX:
            continue

        bbox_in = None
        try:
            # PyMuPDF ≥ 1.18 — get_image_rects returns list of Rect
            rects = page.get_image_rects(xref)
            for r in rects:
                r = r if isinstance(r, fitz.Rect) else fitz.Rect(r)
                # fitz uses points (1 pt = 1/72 inch)
                bbox_in = (r.x0 / 72, r.y0 / 72, r.x1 / 72, r.y1 / 72)
                break
        except Exception:
            try:
                # Older PyMuPDF fallback — use clip rect from image list
                clip = page.get_image_bbox(img_info)
                if clip:
                    r = fitz.Rect(clip)
                    bbox_in = (r.x0 / 72, r.y0 / 72, r.x1 / 72, r.y1 / 72)
            except Exception:
                pass

        if bbox_in is None:
            continue

        sc, sr, ec, er = bbox_to_cells(bbox_in, pw, ph, gc, gr)
        out.append({"pil": pil, "sc": sc, "sr": sr, "ec": ec, "er": er})
    return out


def resize_img(pil):
    """Resize image to fit within MAX_IMG_W_PX × MAX_IMG_H_PX, preserving aspect ratio."""
    w, h = pil.size
    scale = min(MAX_IMG_W_PX / w, MAX_IMG_H_PX / h, 1.0)
    return pil.resize(
        (max(1, int(w * scale)), max(1, int(h * scale))),
        PILImage.LANCZOS
    )


def do_ocr(pil):
    if not HAS_OCR:
        return ""
    try:
        txt = pytesseract.image_to_string(pil, timeout=15).strip()
        return re.sub(r'\n{3,}', '\n\n', txt)
    except Exception:
        return ""


# ════════════════════════════════════════════════════════════════════════════
# TABLE-OCCUPIED CELL SET
# ════════════════════════════════════════════════════════════════════════════

def table_occupied_cells(tables, pw, ph, gc, gr):
    """Return set of (row, col) cells covered by any table's bounding box."""
    occ = set()
    for tbl in tables:
        if tbl["bbox"]:
            sc, sr, ec, er = bbox_to_cells(tbl["bbox"], pw, ph, gc, gr)
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    occ.add((r, c))
    return occ


# ════════════════════════════════════════════════════════════════════════════
# WRITE ONE PAGE → one Excel worksheet
# ════════════════════════════════════════════════════════════════════════════

def write_page(ws, pdata, page_num, fitz_doc=None, page_idx=0):
    pw = pdata["w"]
    ph = pdata["h"]

    # ── Per-page grid dimensions ──────────────────────────────────────────
    # Use global defaults; can be tuned per-page if needed
    gc = GRID_COLS
    gr = GRID_ROWS

    merged_done = set()   # tracks (r1,c1,r2,c2) already merged on this sheet

    # ── 1. Setup uniform grid ─────────────────────────────────────────────
    for ci in range(1, gc + 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTH
    for ri in range(1, gr + 1):
        ws.row_dimensions[ri].height = ROW_HEIGHT
    ws.sheet_view.zoomScale = 85

    tab_occ = table_occupied_cells(pdata["tables"], pw, ph, gc, gr)

    # ═════════════════════════════════════════════════════════════════════
    #  A. TABLES
    # ═════════════════════════════════════════════════════════════════════
    for tbl in pdata["tables"]:
        grid   = tbl["grid"]
        spans  = tbl["spans"]
        kinds  = tbl["kinds"]
        nrows  = tbl["nrows"]
        ncols  = tbl["ncols"]
        bbox   = tbl["bbox"]

        if bbox:
            sc0, sr0, ec0, er0 = bbox_to_cells(bbox, pw, ph, gc, gr)
        else:
            sc0, sr0 = 1, 1
            ec0 = min(gc, ncols * 3)
            er0 = min(gr, nrows * 2)

        # Cell-width / cell-height within the table area (at least 1)
        cw = max(1, (ec0 - sc0 + 1) // max(ncols, 1))
        rh = max(1, (er0 - sr0 + 1) // max(nrows, 1))

        absorbed = set()

        for ri in range(nrows):
            for ci in range(ncols):
                if (ri, ci) in absorbed:
                    continue

                val  = grid[ri][ci]
                kind = kinds.get((ri, ci), "")
                hdr  = kind in ("columnHeader", "rowHeader") or ri == 0

                ex_r = max(1, min(sr0 + ri * rh, gr))
                ex_c = max(1, min(sc0 + ci * cw, gc))

                if hdr:
                    fg, tc, bd, bold = "1F4E79", "FFFFFF", MEDIUM, True
                elif ri % 2 == 0:
                    fg, tc, bd, bold = "EBF3FB", "000000", THIN,   False
                else:
                    fg, tc, bd, bold = "FFFFFF", "000000", THIN,   False

                wcell(ws, ex_r, ex_c, val,
                      bold=bold, size=9, fg=fg, tc=tc, bdr=bd,
                      grid_rows=gr, grid_cols=gc)

                rs, cs = spans.get((ri, ci), (1, 1))
                if rs > 1 or cs > 1:
                    end_r = min(ex_r + rs * rh - 1, gr)
                    end_c = min(ex_c + cs * cw - 1, gc)
                    safe_merge(ws, ex_r, ex_c, end_r, end_c, merged_done, gr, gc)
                    for mr in range(ri, ri + rs):
                        for mc in range(ci, ci + cs):
                            if (mr, mc) != (ri, ci):
                                absorbed.add((mr, mc))

    # ═════════════════════════════════════════════════════════════════════
    #  B. ALL TEXT LINES — placed at their exact x/y position
    # ═════════════════════════════════════════════════════════════════════
    row_buckets = defaultdict(list)   # excel_row → [(sc, ec, text, h_in)]

    for line in pdata["lines"]:
        content = line["content"]
        bbox    = line["bbox"]
        if not content or not bbox:
            continue

        sc, sr, ec, er = bbox_to_cells(bbox, pw, ph, gc, gr)

        # Skip cells inside a table bounding box (table renderer handles them)
        if (sr, sc) in tab_occ:
            continue

        h_in = bbox[3] - bbox[1]
        row_buckets[sr].append((sc, ec, content, h_in))

    for ex_row in sorted(row_buckets.keys()):
        items = sorted(row_buckets[ex_row], key=lambda x: x[0])
        for sc, ec, content, h_in in items:
            sc  = max(1, min(sc, gc))
            ec  = max(sc, min(ec, gc))
            row = max(1, min(ex_row, gr))
            fsz = font_size_from_height(h_in)

            if not is_merged(ws, row, sc):
                existing = ws.cell(row=row, column=sc).value
                if existing:
                    # Append with separator instead of overwriting
                    ws.cell(row=row, column=sc).value = str(existing) + "  " + content
                else:
                    wcell(ws, row, sc, content,
                          size=fsz, bdr=NOBDR, wrap=True,
                          grid_rows=gr, grid_cols=gc)
                    if ec > sc:
                        safe_merge(ws, row, sc, row, ec, merged_done, gr, gc)

            # Adjust row height to accommodate text size
            new_h = min(120.0, max(ROW_HEIGHT, h_in * 72 * 1.5))
            if ws.row_dimensions[row].height < new_h:
                ws.row_dimensions[row].height = new_h

    # ═════════════════════════════════════════════════════════════════════
    #  C. PARAGRAPHS WITH ROLES  (titles, headings, footnotes)
    # ═════════════════════════════════════════════════════════════════════
    for para in pdata["paragraphs"]:
        role    = para["role"]
        content = para["content"]
        bbox    = para["bbox"]
        if not content or not bbox or not role:
            continue

        sc, sr, ec, er = bbox_to_cells(bbox, pw, ph, gc, gr)
        if (sr, sc) in tab_occ:
            continue
        sc = max(1, min(sc, gc));  sr = max(1, min(sr, gr))
        ec = max(sc, min(ec, gc)); er = max(sr, min(er, gr))

        if role == "title":
            fg, tc, bold, fsz = "D6E4F0", "1F4E79", True,  14
        elif role == "sectionHeading":
            fg, tc, bold, fsz = "EBF3FB", "1F4E79", True,  12
        elif role in ("pageHeader", "pageFooter"):
            fg, tc, bold, fsz = "F5F5F5", "666666", False,  8
        elif role == "footnote":
            fg, tc, bold, fsz = None,     "888888", False,  7
        else:
            fg, tc, bold, fsz = None,     "000000", False,  9

        if not is_merged(ws, sr, sc):
            existing = ws.cell(row=sr, column=sc).value
            if not existing:
                wcell(ws, sr, sc, content,
                      bold=bold, size=fsz, fg=fg, tc=tc, bdr=NOBDR, wrap=True,
                      grid_rows=gr, grid_cols=gc)
                if ec > sc or er > sr:
                    safe_merge(ws, sr, sc, er, ec, merged_done, gr, gc)

    # ═════════════════════════════════════════════════════════════════════
    #  D. KEY-VALUE PAIRS — key and value each at their own bbox
    # ═════════════════════════════════════════════════════════════════════
    for kv in pdata["kvpairs"]:
        key = kv["key"]
        val = kv["value"]

        if key and kv["key_bbox"]:
            sc, sr, ec, er = bbox_to_cells(kv["key_bbox"], pw, ph, gc, gr)
            sr = max(1, min(sr, gr));  sc = max(1, min(sc, gc))
            if (sr, sc) not in tab_occ and not is_merged(ws, sr, sc):
                if not ws.cell(row=sr, column=sc).value:
                    wcell(ws, sr, sc, key,
                          bold=True, size=9, tc="1F4E79", bdr=NOBDR,
                          grid_rows=gr, grid_cols=gc)

        if val and kv["val_bbox"]:
            sc, sr, ec, er = bbox_to_cells(kv["val_bbox"], pw, ph, gc, gr)
            sr = max(1, min(sr, gr));  sc = max(1, min(sc, gc))
            ec = max(sc, min(ec, gc))
            if (sr, sc) not in tab_occ and not is_merged(ws, sr, sc):
                if not ws.cell(row=sr, column=sc).value:
                    wcell(ws, sr, sc, val,
                          size=9, fg="FEFCE8", bdr=NOBDR, wrap=True,
                          grid_rows=gr, grid_cols=gc)
                    if ec > sc:
                        safe_merge(ws, sr, sc, sr, ec, merged_done, gr, gc)

    # ═════════════════════════════════════════════════════════════════════
    #  E. IMAGES
    # ═════════════════════════════════════════════════════════════════════
    if fitz_doc is not None:
        for img in get_images(fitz_doc, page_idx, pw, ph, gc, gr):
            pil = img["pil"]
            sc  = max(1, min(img["sc"], gc))
            sr  = max(1, min(img["sr"], gr))

            resized = resize_img(pil)
            buf = io.BytesIO()
            resized.save(buf, format="PNG")
            buf.seek(0)

            try:
                xl_img        = XLImage(buf)
                xl_img.anchor = f"{get_column_letter(sc)}{sr}"
                ws.add_image(xl_img)
            except Exception as e:
                print(f"    WARNING: Could not embed image — {e}")

            # OCR text beside the image
            txt = do_ocr(pil)
            if txt:
                oc  = min(img["ec"] + 1, gc)
                ore = max(1, min(sr, gr))
                if not is_merged(ws, ore, oc):
                    if not ws.cell(row=ore, column=oc).value:
                        wcell(ws, ore, oc, f"[Image text]\n{txt}",
                              italic=True, size=8, tc="555555",
                              fg="FFFDE7", bdr=NOBDR, wrap=True,
                              grid_rows=gr, grid_cols=gc)
                        safe_merge(ws, ore, oc,
                                   min(img["er"], gr),
                                   min(oc + 6, gc),
                                   merged_done, gr, gc)


# ════════════════════════════════════════════════════════════════════════════
# CONVERT SINGLE PDF → EXCEL
# ════════════════════════════════════════════════════════════════════════════

def convert_pdf(pdf_path, out_path):
    bar = "=" * 64
    print(f"\n{bar}")
    print(f"  PDF -> EXCEL | Azure Document Intelligence")
    print(f"  Input : {pdf_path}")
    print(f"  Output: {out_path}")
    print(f"{bar}")

    # Step 1 — Azure DI
    print("  [1/4]  Sending to Azure DI ...", flush=True)
    try:
        result = analyze_pdf(pdf_path)
    except Exception as e:
        print(f"\n  ERROR talking to Azure DI: {e}")
        print("  Check your FORMREC_ENDPOINT and FORMREC_KEY.")
        raise

    n_pages = len(result.pages)
    n_tbl   = len(result.tables) if result.tables else 0
    print(f"  Done: {n_pages} page(s), {n_tbl} table(s) detected")

    # Step 2 — Build page data
    print("  [2/4]  Building page data ...", flush=True)
    all_data = build_page_data(result)

    # Step 3 — Open with PyMuPDF for images
    fitz_doc = None
    if HAS_FITZ and HAS_PIL:
        print("  [3/4]  Opening PDF for image extraction ...", flush=True)
        try:
            fitz_doc = fitz.open(pdf_path)
        except Exception as e:
            print(f"  WARNING: PyMuPDF could not open file ({e}) — images skipped")
    else:
        print("  [3/4]  Image extraction skipped (install pymupdf + pillow)")

    # Step 4 — Write Excel
    print("  [4/4]  Writing Excel workbook ...", flush=True)
    wb    = Workbook()
    first = True

    for pn in sorted(all_data.keys()):
        ws    = wb.active if first else wb.create_sheet()
        first = False

        # Safe title: max 31 chars, strip illegal chars
        raw_title = f"Page {pn}"
        safe_title = re.sub(r'[\\/*?\[\]:]', '', raw_title)[:31]
        ws.title  = safe_title

        pd = all_data[pn]
        print(f"    Page {pn:>3}  |  "
              f"lines={len(pd['lines'])}  "
              f"tables={len(pd['tables'])}  "
              f"kv={len(pd['kvpairs'])}  "
              f"paragraphs={len(pd['paragraphs'])}")

        write_page(ws, pd, pn, fitz_doc=fitz_doc, page_idx=pn - 1)

    # Remove the blank default "Sheet" tab if named sheets were created
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    if fitz_doc:
        fitz_doc.close()

    wb.save(out_path)
    kb = os.path.getsize(out_path) // 1024
    print(f"\n  SAVED -> {out_path}  ({kb} KB)")
    print(f"{bar}\n")


# ════════════════════════════════════════════════════════════════════════════
# BATCH CONVERT (entire folder)
# ════════════════════════════════════════════════════════════════════════════

def convert_folder(in_dir, out_dir):
    in_p  = Path(in_dir)
    out_p = Path(out_dir)
    out_p.mkdir(parents=True, exist_ok=True)
    pdfs  = sorted(in_p.glob("*.pdf"))
    if not pdfs:
        print(f"No PDF files found in {in_dir}"); return

    print(f"\nBatch mode: {len(pdfs)} PDF(s) found in {in_dir}")
    ok = fail = 0
    for pdf in pdfs:
        try:
            convert_pdf(str(pdf), str(out_p / (pdf.stem + ".xlsx")))
            ok += 1
        except Exception as e:
            print(f"  ERROR: {pdf.name} -> {e}")
            fail += 1

    print(f"\nBatch complete: {ok} succeeded, {fail} failed.")


# ════════════════════════════════════════════════════════════════════════════
# VALIDATE KEYS
# ════════════════════════════════════════════════════════════════════════════

def check_keys():
    if "YOUR-RESOURCE" in FORMREC_ENDPOINT or "YOUR_AZURE" in FORMREC_KEY:
        print("\n" + "!" * 64)
        print("  Azure credentials are not set!")
        print("  Edit the CONFIG section at the top of this file, OR")
        print("  set these environment variables before running:")
        print()
        print("  Windows CMD:")
        print("    set FORMREC_ENDPOINT=https://your-resource.cognitiveservices.azure.com/")
        print("    set FORMREC_KEY=your_key_here")
        print()
        print("  Mac / Linux:")
        print("    export FORMREC_ENDPOINT=https://your-resource.cognitiveservices.azure.com/")
        print("    export FORMREC_KEY=your_key_here")
        print("!" * 64 + "\n")
        sys.exit(1)


# ════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════

USAGE = """
pdf_to_excel.py — Convert PDF to Excel using Azure Document Intelligence

Usage:
  python pdf_to_excel.py input.pdf
  python pdf_to_excel.py input.pdf output.xlsx
  python pdf_to_excel.py my_folder/ output_folder/
"""

if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print(USAGE)
        sys.exit(1)

    check_keys()

    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) == 3 else None

    if os.path.isdir(inp):
        convert_folder(inp, out or (inp.rstrip("/\\") + "_excel_output"))
    elif os.path.isfile(inp) and inp.lower().endswith(".pdf"):
        convert_pdf(inp, out or str(Path(inp).with_suffix(".xlsx")))
    else:
        print(f"\nERROR: '{inp}' is not a valid PDF file or folder.")
        print("Make sure the path is correct and the file ends in .pdf")
        sys.exit(1)
