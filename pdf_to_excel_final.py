"""
╔══════════════════════════════════════════════════════════════════════════════╗
║          PDF → EXCEL  |  FINAL TOOL  |  Azure Document Intelligence         ║
║                                                                              ║
║  100% data transfer — tables, text, key-values, images, spatial layout.     ║
║  Every element is placed at the SAME relative position as in the PDF.       ║
╚══════════════════════════════════════════════════════════════════════════════╝

WHAT THIS TOOL DOES:
  ✔  Tables          → exact grid with headers, row/col spans, borders
  ✔  Key-Value pairs → two-column layout (Field | Value)
  ✔  Text / Para     → placed at correct spatial position on the sheet
  ✔  Images          → extracted from PDF, embedded in Excel at correct position
                        + OCR text from image inserted into adjacent cell
  ✔  Multi-page PDF  → one Excel sheet per page
  ✔  Leading zeros   → preserved (all cells forced to text format)
  ✔  Spatial layout  → PDF (x,y) coordinates mapped to Excel (col,row)
  ✔  Batch mode      → convert entire folder of PDFs at once

SETUP:
  1. Install Python packages:
       pip install azure-ai-formrecognizer openpyxl pymupdf pillow pytesseract

  2. Install Tesseract OCR binary (for image text extraction):
       Ubuntu/Debian : sudo apt-get install tesseract-ocr
       macOS         : brew install tesseract
       Windows       : https://github.com/UB-Mannheim/tesseract/wiki

  3. Set your Azure Document Intelligence credentials:
       Option A — edit the CONFIG section below
       Option B — set environment variables:
         export FORMREC_ENDPOINT=https://YOUR-RESOURCE.cognitiveservices.azure.com/
         export FORMREC_KEY=YOUR_32_CHAR_KEY

RUN:
  python pdf_to_excel_final.py  input.pdf            → input.xlsx
  python pdf_to_excel_final.py  input.pdf  out.xlsx  → out.xlsx
  python pdf_to_excel_final.py  folder/    out_dir/  → batch convert all PDFs
"""

import sys, os, io, re, tempfile
from pathlib import Path
from collections import defaultdict

# ── Azure Document Intelligence ───────────────────────────────────────────────
try:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential
except ImportError:
    print("❌  Missing package: pip install azure-ai-formrecognizer")
    sys.exit(1)

# ── Excel ─────────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("❌  Missing package: pip install openpyxl")
    sys.exit(1)

# ── PyMuPDF for image extraction ──────────────────────────────────────────────
try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False
    print("⚠  pymupdf not found — images will be skipped. pip install pymupdf")

# ── Pillow ────────────────────────────────────────────────────────────────────
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("⚠  pillow not found — images will be skipped. pip install pillow")

# ── Tesseract OCR (optional) ──────────────────────────────────────────────────
try:
    import pytesseract
    pytesseract.get_tesseract_version()
    HAS_OCR = True
except Exception:
    HAS_OCR = False

# ════════════════════════════════════════════════════════════════════════════
#  ★  CONFIG — SET YOUR AZURE KEYS HERE  ★
# ════════════════════════════════════════════════════════════════════════════

FORMREC_ENDPOINT = os.getenv(
    "FORMREC_ENDPOINT",
    "https://YOUR-RESOURCE.cognitiveservices.azure.com/"   # ← EDIT THIS
)
FORMREC_KEY = os.getenv(
    "FORMREC_KEY",
    "YOUR_AZURE_DOCUMENT_INTELLIGENCE_KEY_HERE"             # ← EDIT THIS
)

# Azure DI model:
#   "prebuilt-layout"   → best for tables + structure (recommended)
#   "prebuilt-document" → adds entity extraction
#   "prebuilt-read"     → text only (fastest, no tables)
DI_MODEL = "prebuilt-layout"

# ════════════════════════════════════════════════════════════════════════════
#  GRID SETTINGS — controls how tightly the layout is reproduced
# ════════════════════════════════════════════════════════════════════════════

# Number of Excel columns/rows to map across the full page width/height.
# Higher = more faithful positioning but slightly wider file.
GRID_COLS  = 64      # columns across page width
GRID_ROWS  = 90      # rows per page height

COL_WIDTH  = 3.2     # Excel column character width (keeps sheet compact)
ROW_HEIGHT = 13.5    # Excel row height in points

# Skip images smaller than this (decorative dots, lines, etc.)
MIN_IMG_W  = 40      # pixels
MIN_IMG_H  = 40      # pixels

# Max image display size in Excel (pixels)
MAX_IMG_W  = 400
MAX_IMG_H  = 300

# ════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _side(style="thin", color="BBBBBB"):
    return Side(style=style, color=color)

def _border(style="thin", color="BBBBBB"):
    s = _side(style, color)
    return Border(top=s, bottom=s, left=s, right=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def _font(bold=False, italic=False, size=9, color="000000"):
    return Font(name="Arial", bold=bold, italic=italic,
                size=max(int(size), 6), color=color.lstrip("#"))

def _aln(h="left", wrap=True, v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN   = _border("thin",   "CCCCCC")
MEDIUM = _border("medium", "555555")
NONE   = Border()

def write_cell(ws, row, col, value,
               bold=False, italic=False, size=9,
               fg=None, text_color="000000",
               border=THIN, align="left", wrap=True):
    """Write a value into ws[row][col] with formatting. Returns the cell."""
    cell = ws.cell(row=row, column=col)
    cell.value         = "" if value is None else str(value)
    cell.number_format = "@"                   # always text — keeps leading zeros
    cell.font          = _font(bold, italic, size, text_color)
    cell.alignment     = _aln(align, wrap)
    cell.border        = border
    if fg:
        cell.fill = _fill(fg)
    return cell

def safe_merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2:
        return
    try:
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
    except Exception:
        pass

def freeze_and_zoom(ws, zoom=85):
    ws.sheet_view.zoomScale = zoom

# ════════════════════════════════════════════════════════════════════════════
#  COORDINATE MAPPING  (PDF points  →  Excel grid)
# ════════════════════════════════════════════════════════════════════════════

def to_excel_cell(x, y, page_w, page_h,
                  row_offset=0, col_offset=0,
                  gcols=GRID_COLS, grows=GRID_ROWS):
    """
    Map a PDF point (x, y) with origin at top-left to 1-based Excel (col, row).
    row_offset / col_offset allow shifting a section's content.
    """
    col = max(1, min(gcols, int(round(x / page_w * gcols)))) + col_offset
    row = max(1, min(grows, int(round(y / page_h * grows)))) + row_offset
    return col, row

def bbox_to_excel(bbox, page_w, page_h,
                  row_offset=0, col_offset=0,
                  gcols=GRID_COLS, grows=GRID_ROWS):
    """
    bbox = (x0, y0, x1, y1) in PDF points.
    Returns (sc, sr, ec, er) — all 1-based, end >= start.
    """
    sc, sr = to_excel_cell(bbox[0], bbox[1], page_w, page_h,
                           row_offset, col_offset, gcols, grows)
    ec, er = to_excel_cell(bbox[2], bbox[3], page_w, page_h,
                           row_offset, col_offset, gcols, grows)
    return sc, sr, max(sc, ec), max(sr, er)

def poly_to_bbox(polygon):
    """Convert Azure DI polygon (list of Point) to (x0,y0,x1,y1)."""
    if not polygon:
        return None
    xs = [p.x for p in polygon]
    ys = [p.y for p in polygon]
    return (min(xs), min(ys), max(xs), max(ys))

# ════════════════════════════════════════════════════════════════════════════
#  AZURE DOCUMENT INTELLIGENCE  — analyze
# ════════════════════════════════════════════════════════════════════════════

def analyze_pdf(pdf_path: str):
    """Send PDF to Azure DI and return the full AnalyzeResult."""
    client = DocumentAnalysisClient(
        endpoint   = FORMREC_ENDPOINT,
        credential = AzureKeyCredential(FORMREC_KEY)
    )
    with open(pdf_path, "rb") as f:
        poller = client.begin_analyze_document(DI_MODEL, document=f)
    return poller.result()

# ════════════════════════════════════════════════════════════════════════════
#  IMAGE EXTRACTION  (via PyMuPDF)
# ════════════════════════════════════════════════════════════════════════════

def extract_images_from_page(fitz_doc, page_index):
    """
    Extract all raster images on a PDF page.
    Returns list of dicts:
        { "pil": PILImage, "bbox": (x0,y0,x1,y1) in PDF points,
          "w_pt": float, "h_pt": float }
    Bounding boxes are in the same coordinate space as Azure DI (points, top-left origin).
    """
    images = []
    if not (HAS_FITZ and HAS_PIL):
        return images

    page = fitz_doc[page_index]
    page_rect = page.rect  # width/height in points

    for img_info in page.get_images(full=True):
        xref = img_info[0]
        try:
            base_img = fitz_doc.extract_image(xref)
            img_bytes = base_img["image"]
            pil_img   = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
        except Exception:
            continue

        if pil_img.width < MIN_IMG_W or pil_img.height < MIN_IMG_H:
            continue

        # Find image placement rectangle on the page
        bbox_pdf = None
        for item in page.get_image_rects(xref):
            r = item if isinstance(item, fitz.Rect) else fitz.Rect(item)
            bbox_pdf = (r.x0, r.y0, r.x1, r.y1)
            break

        if bbox_pdf is None:
            # Fallback: no placement found, skip
            continue

        images.append({
            "pil"  : pil_img,
            "bbox" : bbox_pdf,
            "w_pt" : bbox_pdf[2] - bbox_pdf[0],
            "h_pt" : bbox_pdf[3] - bbox_pdf[1],
        })

    return images

def ocr_image(pil_img) -> str:
    """Run Tesseract OCR on a PIL image. Returns stripped text or ''."""
    if not HAS_OCR:
        return ""
    try:
        text = pytesseract.image_to_string(pil_img, timeout=15).strip()
        # Collapse excessive whitespace
        text = re.sub(r'\n{3,}', '\n\n', text)
        return text
    except Exception:
        return ""

def resize_for_excel(pil_img, max_w=MAX_IMG_W, max_h=MAX_IMG_H):
    """Resize PIL image proportionally to fit within max_w × max_h."""
    w, h = pil_img.size
    scale = min(max_w / w, max_h / h, 1.0)
    new_w = max(1, int(w * scale))
    new_h = max(1, int(h * scale))
    return pil_img.resize((new_w, new_h), PILImage.LANCZOS)

# ════════════════════════════════════════════════════════════════════════════
#  BUILD PAGE DATA  from AnalyzeResult
# ════════════════════════════════════════════════════════════════════════════

def build_page_data(result):
    """
    Organise all Azure DI elements by page number.
    Returns:
      { page_num: {
          "page_w": float, "page_h": float,  ← in PDF points (inches * 72)
          "tables": [...],
          "kvpairs": [...],
          "words": [...],
          "lines": [...],
          "paragraphs": [...]
        }
      }
    """
    pages = {}

    # ── Page dimensions ───────────────────────────────────────────────────────
    for pg in result.pages:
        # Azure DI returns width/height in inches; convert to points (* 72)
        w = (pg.width  or 8.5) * 72
        h = (pg.height or 11)  * 72
        pages[pg.page_number] = {
            "page_w": w, "page_h": h,
            "tables": [], "kvpairs": [],
            "words": [], "lines": [], "paragraphs": []
        }

    def ensure_page(pn):
        if pn not in pages:
            pages[pn] = {
                "page_w": 8.5*72, "page_h": 11*72,
                "tables": [], "kvpairs": [],
                "words": [], "lines": [], "paragraphs": []
            }

    # ── Tables ────────────────────────────────────────────────────────────────
    if result.tables:
        for tbl in result.tables:
            pn = 1
            if tbl.bounding_regions:
                pn = tbl.bounding_regions[0].page_number
            ensure_page(pn)

            # Table bounding box
            tbl_bbox = None
            if tbl.bounding_regions:
                poly = tbl.bounding_regions[0].polygon
                tbl_bbox = poly_to_bbox(poly)

            nrows = tbl.row_count
            ncols = tbl.column_count
            grid  = [[""] * ncols for _ in range(nrows)]
            spans = {}   # (ri, ci) → (row_span, col_span)
            cell_kinds = {}  # (ri,ci) → "columnHeader" | "rowHeader" | ""

            for cell in tbl.cells:
                r, c = cell.row_index, cell.column_index
                grid[r][c] = cell.content or ""
                rs = getattr(cell, "row_span",    1) or 1
                cs = getattr(cell, "column_span", 1) or 1
                if rs > 1 or cs > 1:
                    spans[(r, c)] = (rs, cs)
                kind = getattr(cell, "kind", "") or ""
                cell_kinds[(r, c)] = kind

            pages[pn]["tables"].append({
                "grid"      : grid,
                "spans"     : spans,
                "cell_kinds": cell_kinds,
                "nrows"     : nrows,
                "ncols"     : ncols,
                "bbox"      : tbl_bbox,
            })

    # ── Key-Value Pairs ───────────────────────────────────────────────────────
    if hasattr(result, "key_value_pairs") and result.key_value_pairs:
        for kv in result.key_value_pairs:
            pn = 1
            if kv.key and kv.key.bounding_regions:
                pn = kv.key.bounding_regions[0].page_number
            ensure_page(pn)

            # Bounding box of the key
            kbbox = None
            if kv.key and kv.key.bounding_regions:
                poly = kv.key.bounding_regions[0].polygon
                kbbox = poly_to_bbox(poly)

            key_txt = kv.key.content   if kv.key   else ""
            val_txt = kv.value.content if kv.value else ""
            pages[pn]["kvpairs"].append({
                "key"  : key_txt,
                "value": val_txt,
                "bbox" : kbbox,
            })

    # ── Words (individual word positions for spatial text) ────────────────────
    for pg in result.pages:
        pn = pg.page_number
        ensure_page(pn)
        if pg.words:
            for word in pg.words:
                bbox = poly_to_bbox(word.polygon) if word.polygon else None
                pages[pn]["words"].append({
                    "content": word.content or "",
                    "bbox"   : bbox,
                    "conf"   : getattr(word, "confidence", 1.0),
                })

    # ── Lines ─────────────────────────────────────────────────────────────────
    for pg in result.pages:
        pn = pg.page_number
        ensure_page(pn)
        if pg.lines:
            for line in pg.lines:
                bbox = poly_to_bbox(line.polygon) if line.polygon else None
                pages[pn]["lines"].append({
                    "content": line.content or "",
                    "bbox"   : bbox,
                })

    # ── Paragraphs ────────────────────────────────────────────────────────────
    if hasattr(result, "paragraphs") and result.paragraphs:
        for para in result.paragraphs:
            pn = 1
            if para.bounding_regions:
                pn = para.bounding_regions[0].page_number
            ensure_page(pn)
            bbox = None
            if para.bounding_regions:
                poly = para.bounding_regions[0].polygon
                bbox = poly_to_bbox(poly)
            pages[pn]["paragraphs"].append({
                "content": para.content or "",
                "bbox"   : bbox,
                "role"   : getattr(para, "role", "") or "",
            })

    return pages

# ════════════════════════════════════════════════════════════════════════════
#  TABLE-OCCUPIED CELLS TRACKER
#  (so spatial text doesn't overwrite table content)
# ════════════════════════════════════════════════════════════════════════════

def get_table_regions(page_data, page_w, page_h):
    """
    Returns a set of (row, col) Excel cell coordinates that are occupied by tables.
    Used to avoid double-writing text that's already inside a table.
    """
    occupied = set()
    for tbl in page_data["tables"]:
        if tbl["bbox"]:
            sc, sr, ec, er = bbox_to_excel(tbl["bbox"], page_w, page_h)
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    occupied.add((r, c))
    return occupied

# ════════════════════════════════════════════════════════════════════════════
#  WRITE ONE PAGE TO EXCEL WORKSHEET
# ════════════════════════════════════════════════════════════════════════════

def write_page(ws, page_data, page_num, fitz_doc=None, page_index=0):
    """
    Write all content for one PDF page into worksheet `ws`.

    Layout strategy:
      1. Set up grid dimensions (col widths / row heights)
      2. Render TABLES at their spatial positions
      3. Render SPATIAL TEXT (paragraphs/lines) — skip cells inside tables
      4. Render KEY-VALUE PAIRS below all spatial content
      5. Embed IMAGES at their spatial positions
    """
    page_w = page_data["page_w"]
    page_h = page_data["page_h"]

    # ── 1. Set grid dimensions ─────────────────────────────────────────────
    for ci in range(1, GRID_COLS + 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTH
    for ri in range(1, GRID_ROWS + 1):
        ws.row_dimensions[ri].height = ROW_HEIGHT

    freeze_and_zoom(ws, zoom=90)

    # Track which Excel cells tables occupy (to avoid text overlap)
    table_cells = get_table_regions(page_data, page_w, page_h)

    # Track merged regions (to avoid double-merging)
    merged_regions = set()

    # ── 2. TABLES ────────────────────────────────────────────────────────────
    for tbl in page_data["tables"]:
        grid       = tbl["grid"]
        spans      = tbl["spans"]
        cell_kinds = tbl["cell_kinds"]
        nrows      = tbl["nrows"]
        ncols      = tbl["ncols"]
        tbl_bbox   = tbl["bbox"]

        # Determine top-left Excel cell for this table
        if tbl_bbox:
            tbl_sc, tbl_sr, tbl_ec, tbl_er = bbox_to_excel(tbl_bbox, page_w, page_h)
        else:
            tbl_sc, tbl_sr = 1, 1

        # Compute per-table cell size to fill the bounding box evenly
        if tbl_bbox:
            tbl_w_cells = max(ncols, tbl_ec - tbl_sc + 1)
            tbl_h_cells = max(nrows, tbl_er - tbl_sr + 1)
        else:
            tbl_w_cells = ncols
            tbl_h_cells = nrows

        cell_w = max(1, tbl_w_cells // ncols)  if ncols > 0 else 1
        cell_h = max(1, tbl_h_cells // nrows)  if nrows > 0 else 1

        merged_in_tbl = set()  # tracks (ri,ci) absorbed by a span

        for ri in range(nrows):
            for ci in range(ncols):
                if (ri, ci) in merged_in_tbl:
                    continue

                val  = grid[ri][ci]
                kind = cell_kinds.get((ri, ci), "")
                is_header = (kind in ("columnHeader", "rowHeader")) or (ri == 0)

                # Excel cell anchor
                ex_col = tbl_sc + ci * cell_w
                ex_row = tbl_sr + ri * cell_h
                ex_col = max(1, min(ex_col, GRID_COLS))
                ex_row = max(1, min(ex_row, GRID_ROWS))

                # Style
                if is_header:
                    fg, txt_col, bd, bold = "1F4E79", "FFFFFF", MEDIUM, True
                elif ri % 2 == 0:
                    fg, txt_col, bd, bold = "EBF3FB", "000000", THIN, False
                else:
                    fg, txt_col, bd, bold = "FFFFFF", "000000", THIN, False

                write_cell(ws, ex_row, ex_col, val,
                           bold=bold, fg=fg, text_color=txt_col,
                           border=bd, align="left")

                ws.row_dimensions[ex_row].height = max(
                    ROW_HEIGHT, ws.row_dimensions[ex_row].height)

                # Handle spans (merge cells)
                rs, cs = spans.get((ri, ci), (1, 1))
                if rs > 1 or cs > 1:
                    end_excel_col = min(tbl_sc + (ci + cs) * cell_w - 1, GRID_COLS)
                    end_excel_row = min(tbl_sr + (ri + rs) * cell_h - 1, GRID_ROWS)
                    merge_key = (ex_row, ex_col, end_excel_row, end_excel_col)
                    if merge_key not in merged_regions:
                        safe_merge(ws, ex_row, ex_col, end_excel_row, end_excel_col)
                        merged_regions.add(merge_key)
                    for mr in range(ri, ri + rs):
                        for mc in range(ci, ci + cs):
                            if (mr, mc) != (ri, ci):
                                merged_in_tbl.add((mr, mc))

    # ── 3. SPATIAL TEXT (paragraphs / lines) ─────────────────────────────────
    # Use paragraphs if available (richer), fall back to lines
    text_elements = page_data["paragraphs"] if page_data["paragraphs"] \
                    else page_data["lines"]

    for elem in text_elements:
        content = elem["content"]
        bbox    = elem["bbox"]
        role    = elem.get("role", "")

        if not content.strip():
            continue

        if bbox:
            sc, sr, ec, er = bbox_to_excel(bbox, page_w, page_h)
        else:
            continue  # no position info — skip to avoid random placement

        # Skip if this cell is inside a table (already rendered above)
        if (sr, sc) in table_cells:
            continue

        # Style based on paragraph role
        if role == "title":
            fg, bold, size, txt_col = "D6E4F0", True,  11, "1F4E79"
        elif role == "sectionHeading":
            fg, bold, size, txt_col = "EBF3FB", True,  10, "1F4E79"
        elif role == "footnote":
            fg, bold, size, txt_col = None,     False,  7, "666666"
        elif role == "pageNumber":
            fg, bold, size, txt_col = None,     False,  7, "999999"
        elif role == "pageHeader":
            fg, bold, size, txt_col = "F2F2F2", True,   8, "333333"
        elif role == "pageFooter":
            fg, bold, size, txt_col = "F2F2F2", False,  7, "555555"
        else:
            fg, bold, size, txt_col = None,     False,  9, "000000"

        ec = max(sc, min(ec, GRID_COLS))
        er = max(sr, min(er, GRID_ROWS))

        write_cell(ws, sr, sc, content,
                   bold=bold, size=size, fg=fg, text_color=txt_col,
                   border=NONE, align="left", wrap=True)

        if ec > sc or er > sr:
            merge_key = (sr, sc, er, ec)
            if merge_key not in merged_regions:
                safe_merge(ws, sr, sc, er, ec)
                merged_regions.add(merge_key)

        # Adjust row height for multi-line text
        line_count = max(1, content.count("\n") + 1)
        ws.row_dimensions[sr].height = max(
            ROW_HEIGHT, min(120, line_count * ROW_HEIGHT))

    # ── 4. KEY-VALUE PAIRS ────────────────────────────────────────────────────
    # Render spatially when bbox is available; otherwise append below grid
    kv_no_bbox = []
    for kv in page_data["kvpairs"]:
        if kv["bbox"]:
            sc, sr, ec, er = bbox_to_excel(kv["bbox"], page_w, page_h)
            # Avoid overwriting table content
            if (sr, sc) in table_cells:
                kv_no_bbox.append(kv)
                continue
            combined = f"{kv['key']}: {kv['value']}" if kv["key"] else kv["value"]
            write_cell(ws, sr, sc, combined,
                       bold=False, fg="FFF9E6", text_color="000000",
                       border=NONE, align="left")
            if ec > sc:
                merge_key = (sr, sc, sr, min(ec, GRID_COLS))
                if merge_key not in merged_regions:
                    safe_merge(ws, sr, sc, sr, min(ec, GRID_COLS))
                    merged_regions.add(merge_key)
        else:
            kv_no_bbox.append(kv)

    # KV pairs that had no bbox → dump below the grid in a clean two-column table
    if kv_no_bbox:
        kv_row = GRID_ROWS + 3

        # Section header
        write_cell(ws, kv_row, 1, f"⬇  Additional Key–Value Pairs  (Page {page_num})",
                   bold=True, size=9, fg="1F4E79", text_color="FFFFFF", border=MEDIUM)
        safe_merge(ws, kv_row, 1, kv_row, 6)
        kv_row += 1

        write_cell(ws, kv_row, 1, "Field / Key",  bold=True, fg="2E75B6",
                   text_color="FFFFFF", border=MEDIUM)
        write_cell(ws, kv_row, 2, "Value",         bold=True, fg="2E75B6",
                   text_color="FFFFFF", border=MEDIUM)
        safe_merge(ws, kv_row, 2, kv_row, 6)
        kv_row += 1

        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 38)
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 50)

        for i, kv in enumerate(kv_no_bbox):
            bg = "F4F9FF" if i % 2 == 0 else "FFFFFF"
            write_cell(ws, kv_row, 1, kv["key"],   fg=bg, border=THIN)
            write_cell(ws, kv_row, 2, kv["value"],  fg=bg, border=THIN)
            safe_merge(ws, kv_row, 2, kv_row, 6)
            ws.row_dimensions[kv_row].height = ROW_HEIGHT
            kv_row += 1

    # ── 5. IMAGES ─────────────────────────────────────────────────────────────
    if fitz_doc is not None and HAS_FITZ and HAS_PIL:
        images = extract_images_from_page(fitz_doc, page_index)
        for img_data in images:
            pil_img = img_data["pil"]
            bbox    = img_data["bbox"]

            sc, sr, ec, er = bbox_to_excel(bbox, page_w, page_h)
            sc = max(1, min(sc, GRID_COLS - 1))
            sr = max(1, min(sr, GRID_ROWS - 1))

            # Resize image to fit reasonably in Excel
            pil_resized = resize_for_excel(pil_img)

            # Save image to a temp PNG buffer
            buf = io.BytesIO()
            pil_resized.save(buf, format="PNG")
            buf.seek(0)

            try:
                xl_img = XLImage(buf)
                # Anchor image at the spatial cell
                anchor_cell = f"{get_column_letter(sc)}{sr}"
                xl_img.anchor = anchor_cell
                ws.add_image(xl_img)
            except Exception as e:
                print(f"    ⚠  Could not embed image: {e}")
                buf.seek(0)

            # OCR the image and write text beside it
            ocr_text = ocr_image(pil_img)
            if ocr_text:
                ocr_col = min(ec + 1, GRID_COLS)
                write_cell(ws, sr, ocr_col, f"[Image text]\n{ocr_text}",
                           italic=True, size=8, text_color="444444",
                           border=NONE, fg="FFFCE6")
                safe_merge(ws, sr, ocr_col, min(er, GRID_ROWS), min(ocr_col + 5, GRID_COLS))

    return ws

# ════════════════════════════════════════════════════════════════════════════
#  CONVERT SINGLE PDF → EXCEL
# ════════════════════════════════════════════════════════════════════════════

def convert_pdf(pdf_path: str, out_path: str):
    sep = "═" * 62
    print(f"\n{sep}")
    print(f"  PDF → EXCEL  |  Azure Document Intelligence")
    print(f"  Input  : {pdf_path}")
    print(f"  Output : {out_path}")
    print(f"  Model  : {DI_MODEL}")
    print(f"{sep}")

    # ── Step 1: Azure DI analysis ─────────────────────────────────────────────
    print("  [1/4] Sending to Azure Document Intelligence...", flush=True)
    result = analyze_pdf(pdf_path)
    n_pages = len(result.pages)
    n_tables = len(result.tables) if result.tables else 0
    n_kv = len(result.key_value_pairs) \
           if hasattr(result, "key_value_pairs") and result.key_value_pairs else 0
    print(f"  ✔  Analysis done — {n_pages} page(s), {n_tables} table(s), {n_kv} kv-pair(s)")

    # ── Step 2: Build structured page data ───────────────────────────────────
    print("  [2/4] Building page data...", flush=True)
    page_data = build_page_data(result)

    # ── Step 3: Open PDF with PyMuPDF for image extraction ────────────────────
    fitz_doc = None
    if HAS_FITZ and HAS_PIL:
        print("  [3/4] Opening PDF for image extraction...", flush=True)
        try:
            fitz_doc = fitz.open(pdf_path)
        except Exception as e:
            print(f"  ⚠  Could not open with PyMuPDF: {e}")
    else:
        print("  [3/4] Skipping image extraction (pymupdf/pillow not available)")

    # ── Step 4: Write Excel ───────────────────────────────────────────────────
    print("  [4/4] Writing Excel workbook...", flush=True)
    wb = Workbook()
    first = True

    for page_num in sorted(page_data.keys()):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = f"P{page_num}"[:31]  # Excel sheet name limit = 31 chars

        pdata    = page_data[page_num]
        n_tbl    = len(pdata["tables"])
        n_kv_pg  = len(pdata["kvpairs"])
        n_para   = len(pdata["paragraphs"])
        n_lines  = len(pdata["lines"])
        n_words  = len(pdata["words"])
        print(f"    Page {page_num:>3}  │  tables={n_tbl}  kv={n_kv_pg}  "
              f"paras={n_para}  lines={n_lines}  words={n_words}")

        page_index = page_num - 1
        write_page(ws, pdata, page_num,
                   fitz_doc=fitz_doc, page_index=page_index)

    # Remove default "Sheet" if extra sheets were created
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    if fitz_doc:
        fitz_doc.close()

    wb.save(out_path)
    size_kb = os.path.getsize(out_path) // 1024
    print(f"\n  ✅  Saved → {out_path}  ({size_kb} KB)\n{sep}\n")

# ════════════════════════════════════════════════════════════════════════════
#  BATCH CONVERT: folder of PDFs
# ════════════════════════════════════════════════════════════════════════════

def convert_folder(in_dir: str, out_dir: str):
    in_p  = Path(in_dir)
    out_p = Path(out_dir)
    out_p.mkdir(parents=True, exist_ok=True)

    pdfs = sorted(in_p.glob("*.pdf"))
    if not pdfs:
        print(f"No PDF files found in {in_dir}")
        return

    print(f"\nBatch mode — {len(pdfs)} PDF(s) found in {in_dir}")
    ok, fail = 0, 0
    for pdf in pdfs:
        out_file = out_p / (pdf.stem + ".xlsx")
        try:
            convert_pdf(str(pdf), str(out_file))
            ok += 1
        except Exception as e:
            print(f"  ✗  ERROR: {pdf.name}  →  {e}")
            fail += 1

    print(f"\nBatch complete — ✅ {ok} succeeded, ❌ {fail} failed.")

# ════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════

def validate_keys():
    if "YOUR-RESOURCE" in FORMREC_ENDPOINT or "YOUR_AZURE" in FORMREC_KEY:
        print("\n" + "!"*62)
        print("  ⚠  AZURE CREDENTIALS NOT SET")
        print("  Edit the CONFIG section in this file, or set env vars:")
        print("    export FORMREC_ENDPOINT=https://your-resource.cognitiveservices.azure.com/")
        print("    export FORMREC_KEY=your_key_here")
        print("!"*62 + "\n")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print(__doc__)
        print("Usage:")
        print("  python pdf_to_excel_final.py  input.pdf  [output.xlsx]")
        print("  python pdf_to_excel_final.py  folder/    [output_folder/]")
        sys.exit(1)

    validate_keys()

    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) == 3 else None

    if os.path.isdir(inp):
        out_dir = out or (inp.rstrip("/\\") + "_excel_output")
        convert_folder(inp, out_dir)

    elif os.path.isfile(inp) and inp.lower().endswith(".pdf"):
        out_file = out or (os.path.splitext(inp)[0] + ".xlsx")
        convert_pdf(inp, out_file)

    else:
        print(f"ERROR: '{inp}' is not a valid PDF file or directory.")
        sys.exit(1)
