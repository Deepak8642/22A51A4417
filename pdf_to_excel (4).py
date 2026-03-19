"""
PDF -> Excel Converter  —  FINAL ACCURATE VERSION
==================================================
Key insight that makes this accurate:
  - ROW boundaries = unique word 'top' positions (NOT horizontal lines)
    Because h_lines in PDFs often span multiple visual rows
  - COLUMN boundaries = actual vertical lines from the PDF
  - Text within a zone (row x column) is joined into one cell
  - Section headers (black rects) → full-width merged, white bold text
  - Column headers (grey rects)   → grey background, bold
  - Text alignment detected per-zone from x-position math
  - Images placed using exact PDF coordinates

INSTALL (pip only, no admin needed):
    pip install pdfplumber pymupdf pillow openpyxl

USAGE:
    python pdf_to_excel.py  input.pdf  output.xlsx
"""

import sys, os, io, re
from collections import defaultdict

try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import pdfplumber
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# STYLING
# ══════════════════════════════════════════════════════════════════════════════
def solid(h): return PatternFill("solid", fgColor=h)

def bdr(t="thin",b="thin",l="thin",r="thin",c="AAAAAA"):
    s = lambda st: Side(style=st,color=c) if st else Side(style=None)
    return Border(top=s(t),bottom=s(b),left=s(l),right=s(r))

THIN  = bdr()
THICK = bdr("medium","medium","medium","medium",c="444444")
NONE  = bdr(None,None,None,None)

# ══════════════════════════════════════════════════════════════════════════════
# IMAGES
# ══════════════════════════════════════════════════════════════════════════════
def get_images(pdf_path):
    if not HAS_FITZ:
        print("  [INFO] pip install pymupdf  to keep images")
        return []
    out, doc = [], fitz.open(pdf_path)
    for pn in range(len(doc)):
        pg = doc[pn]
        for info in pg.get_images(full=True):
            xref = info[0]
            try:
                b   = doc.extract_image(xref)
                pil = PILImage.open(io.BytesIO(b["image"])).convert("RGBA")
                rs  = pg.get_image_rects(xref)
                r   = rs[0] if rs else fitz.Rect(0,0,pil.width,pil.height)
                out.append(dict(page=pn,x0=r.x0,y0=r.y0,x1=r.x1,y1=r.y1,pil=pil))
            except: pass
    doc.close(); return out

def pil_to_xl(pil, wp, hp):
    wp,hp = max(wp,20), max(hp,20)
    bg = PILImage.new("RGB",pil.size,(255,255,255))
    bg.paste(pil.convert("RGB"), mask=pil.split()[3] if pil.mode=="RGBA" else None)
    bg = bg.resize((wp,hp), PILImage.LANCZOS)
    buf=io.BytesIO(); bg.save(buf,"PNG"); buf.seek(0)
    xl=XLImage(buf); xl.width=wp; xl.height=hp; return xl

# ══════════════════════════════════════════════════════════════════════════════
# CORE: snap helper
# ══════════════════════════════════════════════════════════════════════════════
def snap(vals, tol=3):
    out = []
    for v in sorted(set(round(float(x),1) for x in vals)):
        if out and v - out[-1] <= tol:
            out[-1] = (out[-1]+v)/2
        else:
            out.append(v)
    return sorted(out)

# ══════════════════════════════════════════════════════════════════════════════
# GET COLUMN BOUNDARIES from vertical lines
# ══════════════════════════════════════════════════════════════════════════════
def get_col_boundaries(page):
    v_raw = [l["x0"] for l in page.lines if abs(l["x0"]-l["x1"]) < 2]
    v_raw += [0, float(page.width)]
    return snap(v_raw, tol=4)

# ══════════════════════════════════════════════════════════════════════════════
# GET ROW BOUNDARIES from word tops (NOT h_lines)
# ══════════════════════════════════════════════════════════════════════════════
def get_row_boundaries(words, page_height):
    tops = [float(w["top"]) for w in words]
    tops += [0, float(page_height)]
    return snap(tops, tol=2)   # 2pt tolerance — keeps rows separate

# ══════════════════════════════════════════════════════════════════════════════
# GET FILLED RECTANGLES
# ══════════════════════════════════════════════════════════════════════════════
def get_fills(page):
    out = []
    for r in page.rects:
        c = r.get("non_stroking_color")
        if c is None: continue
        if isinstance(c,(int,float)): c=(float(c),)*3
        if not (isinstance(c,(list,tuple)) and len(c)>=3): continue
        r0,g0,b0 = float(c[0]),float(c[1]),float(c[2])
        if   r0<0.15 and g0<0.15 and b0<0.15: kind="black"
        elif r0>0.55 and g0>0.55 and b0>0.55: kind="grey"
        else: kind="other"
        out.append(dict(top=float(r["top"]),bottom=float(r["bottom"]),
                        x0=float(r["x0"]),x1=float(r["x1"]),kind=kind))
    return out

def fill_at(y, fills):
    for f in fills:
        if f["top"]-3 <= y <= f["bottom"]+3:
            return f["kind"]
    return None

# ══════════════════════════════════════════════════════════════════════════════
# ASSIGN WORDS → (row_idx, col_idx)
# ══════════════════════════════════════════════════════════════════════════════
def find_idx(val, boundaries):
    for i in range(len(boundaries)-1):
        if boundaries[i]-2 <= val < boundaries[i+1]+2:
            return i
    return len(boundaries)-2

def assign(words, row_bounds, col_bounds):
    grid = defaultdict(list)
    for w in words:
        mid_y = (float(w["top"]) + float(w["bottom"])) / 2
        ri = find_idx(mid_y, row_bounds)
        ci = find_idx(float(w["x0"])+1, col_bounds)
        grid[(ri,ci)].append(w)
    return grid

# ══════════════════════════════════════════════════════════════════════════════
# ALIGNMENT DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def get_align(wds, cx0, cx1):
    if not wds: return "left"
    xmin = min(float(w["x0"]) for w in wds)
    xmax = max(float(w["x1"]) for w in wds)
    mid_t = (xmin+xmax)/2
    mid_c = (cx0+cx1)/2
    cw    = cx1-cx0
    if cw < 5: return "left"
    if abs(mid_t-mid_c) <= cw*0.15: return "center"
    if xmin > cx0+cw*0.6: return "right"
    return "left"

# ══════════════════════════════════════════════════════════════════════════════
# WRITE ONE PAGE
# ══════════════════════════════════════════════════════════════════════════════
def write_page(ws, page, page_images):
    pw = float(page.width)
    ph = float(page.height)

    words = page.extract_words(x_tolerance=2, y_tolerance=2,
                               keep_blank_chars=False,
                               extra_attrs=["fontname","size"])

    col_bounds = get_col_boundaries(page)
    row_bounds = get_row_boundaries(words, ph)
    fills      = get_fills(page)
    grid       = assign(words, row_bounds, col_bounds)

    n_rows = len(row_bounds)-1
    n_cols = len(col_bounds)-1

    # Column widths — proportional to PDF width
    for ci in range(n_cols):
        cw = col_bounds[ci+1]-col_bounds[ci]
        ws.column_dimensions[get_column_letter(ci+1)].width = max(3, round(cw/5.5))

    for ri in range(n_rows):
        y_top = row_bounds[ri]
        y_bot = row_bounds[min(ri+1, len(row_bounds)-1)]
        mid_y = (y_top+y_bot)/2
        xl_row = ri+1

        # Row height proportional to PDF
        ws.row_dimensions[xl_row].height = max(10, round((y_bot-y_top)*0.88))

        kind = fill_at(mid_y, fills)

        # Collect all words in this row
        all_row_words = []
        for ci in range(n_cols):
            all_row_words.extend(grid.get((ri,ci),[]))

        if not all_row_words and kind is None:
            continue

        # ── BLACK SECTION HEADER ─────────────────────────────────────────────
        if kind == "black":
            text = " ".join(w["text"] for w in
                            sorted(all_row_words, key=lambda w:float(w["x0"])))
            cell = ws.cell(row=xl_row, column=1, value=text.strip())
            cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill      = solid("111111")
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True)
            cell.border    = THICK
            ws.row_dimensions[xl_row].height = max(15, round((y_bot-y_top)*0.88))
            if n_cols > 1:
                try:
                    ws.merge_cells(start_row=xl_row, start_column=1,
                                   end_row=xl_row, end_column=n_cols)
                except: pass
            continue

        # ── GREY COLUMN HEADER ───────────────────────────────────────────────
        if kind == "grey":
            for ci in range(n_cols):
                wds = sorted(grid.get((ri,ci),[]), key=lambda w:float(w["x0"]))
                txt = " ".join(w["text"] for w in wds).strip()
                if not txt: continue
                cx0 = col_bounds[ci]; cx1 = col_bounds[min(ci+1,len(col_bounds)-1)]
                al  = get_align(wds, cx0, cx1)
                cell = ws.cell(row=xl_row, column=ci+1, value=txt)
                cell.font      = Font(name="Arial", bold=True, size=8, color="111111")
                cell.fill      = solid("D0D0D0")
                cell.alignment = Alignment(horizontal=al, vertical="center",
                                           wrap_text=True)
                cell.border    = THIN
            continue

        # ── NORMAL ROW ───────────────────────────────────────────────────────
        all_bold = (all_row_words and
                    all("Bold" in str(w.get("fontname","")) for w in all_row_words))
        sz_list  = [float(w.get("size") or 9) for w in all_row_words if w.get("size")]
        avg_sz   = round(sum(sz_list)/len(sz_list)) if sz_list else 9
        bg       = "F6F6F6" if (xl_row % 2 == 0) else "FFFFFF"

        # Find contiguous column spans that have text and should be merged
        # Rule: a span starts at each non-empty zone; it extends rightward
        # only if the words' x1 extends past the next column boundary
        ci = 0
        while ci < n_cols:
            wds = sorted(grid.get((ri,ci),[]), key=lambda w:float(w["x0"]))
            txt = " ".join(w["text"] for w in wds).strip()

            if not txt:
                ci += 1
                continue

            # Try to extend span: if words reach into adjacent column zones
            xmax = max(float(w["x1"]) for w in wds)
            span_end = ci
            while span_end+1 < n_cols:
                next_bound = col_bounds[span_end+1]
                # Only extend if text physically crosses into next zone
                if xmax > next_bound - 5:
                    extra = sorted(grid.get((ri,span_end+1),[]),
                                   key=lambda w:float(w["x0"]))
                    if extra:
                        wds  += extra
                        txt   = " ".join(w["text"] for w in
                                         sorted(wds, key=lambda w:float(w["x0"]))).strip()
                        xmax  = max(float(w["x1"]) for w in wds)
                        span_end += 1
                    else:
                        break
                else:
                    break

            cx0 = col_bounds[ci]
            cx1 = col_bounds[min(span_end+1, len(col_bounds)-1)]
            al  = get_align(wds, cx0, cx1)

            cell = ws.cell(row=xl_row, column=ci+1, value=txt)
            cell.font      = Font(name="Arial", bold=all_bold,
                                  size=max(avg_sz,8), color="111111")
            cell.fill      = solid(bg)
            cell.alignment = Alignment(horizontal=al, vertical="top",
                                       wrap_text=True)
            cell.border    = THIN

            if span_end > ci:
                try:
                    ws.merge_cells(start_row=xl_row, start_column=ci+1,
                                   end_row=xl_row,   end_column=span_end+1)
                except: pass

            ci = span_end + 1

    # ── IMAGES ────────────────────────────────────────────────────────────────
    for img in page_images:
        PX = 96/72
        wp = max(int((img["x1"]-img["x0"])*PX), 30)
        hp = max(int((img["y1"]-img["y0"])*PX), 20)
        try:
            xl_img = pil_to_xl(img["pil"], wp, hp)
            ar = find_idx(img["y0"], row_bounds) + 1
            ac = find_idx(img["x0"], col_bounds) + 1
            ws.add_image(xl_img, f"{get_column_letter(ac)}{ar}")
            print(f"    img -> {get_column_letter(ac)}{ar} ({wp}x{hp}px)")
        except Exception as e:
            print(f"    img skipped: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def convert(pdf_path, out_path):
    print(f"\nConverting: {pdf_path}")
    imgs = get_images(pdf_path)
    print(f"  {len(imgs)} image(s) found")
    wb = Workbook()
    with pdfplumber.open(pdf_path) as pdf:
        for pn, page in enumerate(pdf.pages):
            ws = wb.active if pn==0 else wb.create_sheet()
            ws.title = f"Page {pn+1}"
            print(f"  [Page {pn+1}]  {page.width:.0f}x{page.height:.0f} pts")
            write_page(ws, page, [im for im in imgs if im["page"]==pn])
    wb.save(out_path)
    print(f"\n  Saved -> {out_path}\n")

if __name__ == "__main__":
    if len(sys.argv)!=3:
        print("Usage: python pdf_to_excel.py  input.pdf  output.xlsx")
        sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"Not found: {sys.argv[1]}"); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
