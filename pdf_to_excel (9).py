"""
PDF → Excel  ·  FINAL ACCURATE VERSION
========================================
Uses pdfplumber's line-based table extraction.
Row fill detection uses actual char y-positions (not h-line midpoints).

INSTALL:  pip install pdfplumber pymupdf pillow openpyxl
RUN:      python pdf_to_excel.py  input.pdf  output.xlsx
"""

import sys, os, io
from collections import defaultdict

try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import pdfplumber
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── styles ─────────────────────────────────────────────────────────────────────
def solid(h): return PatternFill("solid", fgColor=h)
def bdr(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)
THIN  = bdr("thin","CCCCCC")
THICK = bdr("medium","444444")
def fnt(bold=False, size=9, color="111111"):
    return Font(name="Arial", bold=bold, size=int(max(size,8)), color=color)
def aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

# ── images ─────────────────────────────────────────────────────────────────────
def get_images(pdf_path):
    if not HAS_FITZ:
        print("  [INFO] pip install pymupdf  to preserve images"); return []
    out, doc = [], fitz.open(pdf_path)
    for pn in range(len(doc)):
        pg = doc[pn]
        for info in pg.get_images(full=True):
            xref = info[0]
            try:
                b = doc.extract_image(xref)
                pil = PILImage.open(io.BytesIO(b["image"])).convert("RGBA")
                rs = pg.get_image_rects(xref)
                r = rs[0] if rs else fitz.Rect(0,0,pil.width,pil.height)
                out.append(dict(page=pn,x0=r.x0,y0=r.y0,x1=r.x1,y1=r.y1,pil=pil))
            except: pass
    doc.close(); return out

def pil_to_xl(pil, wp, hp):
    wp,hp = max(wp,20),max(hp,20)
    bg = PILImage.new("RGB",pil.size,(255,255,255))
    bg.paste(pil.convert("RGB"), mask=pil.split()[3] if pil.mode=="RGBA" else None)
    bg = bg.resize((wp,hp),PILImage.LANCZOS)
    buf=io.BytesIO(); bg.save(buf,"PNG"); buf.seek(0)
    xl=XLImage(buf); xl.width=wp; xl.height=hp; return xl

# ── fills ──────────────────────────────────────────────────────────────────────
def get_fills(page):
    out = []
    for r in page.rects:
        c = r.get("non_stroking_color")
        if c is None: continue
        if isinstance(c,(int,float)): c=(float(c),)*3
        if not (isinstance(c,(list,tuple)) and len(c)>=3): continue
        r0,g0,b0 = float(c[0]),float(c[1]),float(c[2])
        if   r0<0.15 and g0<0.15 and b0<0.15: kind="black"
        elif r0>0.50 and g0>0.50 and b0>0.50: kind="grey"
        else: continue
        out.append((float(r["top"]),float(r["bottom"]),kind))
    return out

def fill_at(y, fills, tol=3):
    for top,bot,kind in fills:
        if top-tol <= y <= bot+tol: return kind
    return None

# ── alignment ──────────────────────────────────────────────────────────────────
def detect_align(chs, cx0, cx1):
    if not chs: return "left"
    xs = [float(c["x0"]) for c in chs]; xe = [float(c["x1"]) for c in chs]
    mid_t=(min(xs)+max(xe))/2; mid_c=(cx0+cx1)/2; cw=cx1-cx0
    if cw<5: return "left"
    if abs(mid_t-mid_c)<=cw*0.20: return "center"
    if min(xs)>cx0+cw*0.55: return "right"
    return "left"

# ── write page ─────────────────────────────────────────────────────────────────
def write_page(ws, page, page_images):
    pw,ph = float(page.width),float(page.height)
    fills = get_fills(page)
    all_chars = page.chars

    # Build char-y lookup: for each table row, find the actual y of chars in it
    # Group chars by y-bucket
    char_by_y = defaultdict(list)
    for ch in all_chars:
        char_by_y[round(float(ch["top"]))].append(ch)

    # Extract table using PDF's actual drawn lines
    tables = page.find_tables({
        "vertical_strategy":   "lines",
        "horizontal_strategy": "lines",
        "snap_tolerance": 5, "join_tolerance": 3,
    })
    if not tables:
        print("    [WARN] No table found"); return

    table = max(tables, key=lambda t:(t.bbox[2]-t.bbox[0])*(t.bbox[3]-t.bbox[1]))
    data  = table.extract()
    if not data: return

    n_rows = len(data)
    n_cols = max(len(r) for r in data)
    bbox   = table.bbox

    # Column x-boundaries from vertical lines
    v_xs = sorted(set(
        round(l["x0"],1) for l in page.lines
        if abs(l["x0"]-l["x1"])<1 and bbox[1]-5<=l["y0"]<=bbox[3]+5
    ))
    v_xs = [x for x in v_xs if bbox[0]-5<=x<=bbox[2]+5]

    # Row y-boundaries from horizontal lines  
    h_ys = sorted(set(
        round(l["y0"],1) for l in page.lines
        if abs(l["y0"]-l["y1"])<1 and bbox[1]-5<=l["y0"]<=bbox[3]+5
    ))

    # Set column widths
    for ci in range(min(len(v_xs)-1, n_cols)):
        cw = v_xs[ci+1]-v_xs[ci]
        ws.column_dimensions[get_column_letter(ci+1)].width = max(2,round(cw/5.2))

    # Process each row
    for ri, row in enumerate(data):
        xl_row = ri+1

        # Get the y-band for this row from h_lines
        if ri < len(h_ys)-1:
            row_top = h_ys[ri]; row_bot = h_ys[ri+1]
        else:
            row_top = bbox[1]+(ri/n_rows)*(bbox[3]-bbox[1])
            row_bot = bbox[1]+((ri+1)/n_rows)*(bbox[3]-bbox[1])

        # Row height
        ws.row_dimensions[xl_row].height = max(10, round((row_bot-row_top)*0.88))

        # Find ACTUAL chars in this row band (for fill detection + alignment)
        row_chars = [ch for ch in all_chars
                     if row_top-2 <= float(ch["top"]) <= row_bot+2]

        # Fill detection: use ACTUAL char y positions
        row_char_ys = [float(ch["top"]) for ch in row_chars]
        mid_y = sum(row_char_ys)/len(row_char_ys) if row_char_ys else (row_top+row_bot)/2
        kind  = fill_at(mid_y, fills)

        # Build spans: (col_start, col_end, text)
        spans = []
        ci = 0
        while ci < len(row):
            val = row[ci]
            if val is not None:
                text = str(val).strip()
                # Extend span over trailing None cells
                end = ci
                for k in range(ci+1, len(row)):
                    if row[k] is None: end = k
                    else: break
                spans.append((ci, end, text))
                ci = end+1
            else:
                ci += 1

        # ── BLACK SECTION HEADER ─────────────────────────────────────────────
        if kind == "black":
            all_text = "  " + "   ".join(s[2] for s in spans if s[2])
            cell = ws.cell(row=xl_row, column=1, value=all_text)
            cell.font = fnt(bold=True, size=10, color="FFFFFF")
            cell.fill = solid("111111")
            cell.alignment = aln("left")
            cell.border = THICK
            ws.row_dimensions[xl_row].height = 15
            if n_cols > 1:
                try: ws.merge_cells(start_row=xl_row,start_column=1,end_row=xl_row,end_column=n_cols)
                except: pass
            continue

        # ── GREY HEADER ───────────────────────────────────────────────────────
        if kind == "grey":
            ws.row_dimensions[xl_row].height = 14
            used = set()
            for (cs,ce,text) in spans:
                if not text or cs+1 in used: continue
                cx0 = v_xs[cs]   if cs   < len(v_xs) else bbox[0]
                cx1 = v_xs[ce+1] if ce+1 < len(v_xs) else bbox[2]
                chs = [ch for ch in row_chars if cx0-2<=float(ch["x0"])<=cx1+2]
                al  = detect_align(chs, cx0, cx1)
                cell = ws.cell(row=xl_row, column=cs+1, value=text)
                cell.font = fnt(bold=True, size=9, color="111111")
                cell.fill = solid("D2D2D2")
                cell.alignment = aln(al)
                cell.border = THIN
                if ce>cs:
                    try:
                        ws.merge_cells(start_row=xl_row,start_column=cs+1,end_row=xl_row,end_column=ce+1)
                        for c in range(cs+1,ce+2): used.add(c)
                    except: pass
                else: used.add(cs+1)
            continue

        # ── NORMAL ROW ────────────────────────────────────────────────────────
        bg   = "F5F5F5" if (xl_row%2==0) else "FFFFFF"
        used = set()
        for (cs,ce,text) in spans:
            if not text or cs+1 in used: continue
            cx0 = v_xs[cs]   if cs   < len(v_xs) else bbox[0]
            cx1 = v_xs[ce+1] if ce+1 < len(v_xs) else bbox[2]
            chs = [ch for ch in row_chars if cx0-2<=float(ch["x0"])<=cx1+2]
            is_bold = any("Bold" in str(c.get("fontname","")) for c in chs)
            sz  = float(chs[0].get("size") or 9) if chs else 9
            al  = detect_align(chs, cx0, cx1)
            cell = ws.cell(row=xl_row, column=cs+1, value=text)
            cell.font = fnt(bold=is_bold, size=max(int(sz),8))
            cell.fill = solid(bg)
            cell.alignment = aln(al)
            cell.border = THIN
            if ce>cs:
                try:
                    ws.merge_cells(start_row=xl_row,start_column=cs+1,end_row=xl_row,end_column=ce+1)
                    for c in range(cs+1,ce+2): used.add(c)
                except: pass
            else: used.add(cs+1)

    # ── images ────────────────────────────────────────────────────────────────
    PX=96/72
    for img in page_images:
        wp=max(int((img["x1"]-img["x0"])*PX),30)
        hp=max(int((img["y1"]-img["y0"])*PX),20)
        try:
            xl_img=pil_to_xl(img["pil"],wp,hp)
            ri_img=max(0,int((img["y0"]-bbox[1])/(bbox[3]-bbox[1])*n_rows))
            ci_img=max(0,int((img["x0"]-bbox[0])/(bbox[2]-bbox[0])*n_cols))
            ws.add_image(xl_img,f"{get_column_letter(ci_img+1)}{ri_img+1}")
        except Exception as e: print(f"    img: {e}")

# ── main ───────────────────────────────────────────────────────────────────────
def convert(pdf_path, out_path):
    print(f"\nConverting: {pdf_path}")
    images = get_images(pdf_path)
    print(f"  {len(images)} image(s)")
    wb = Workbook()
    with pdfplumber.open(pdf_path) as pdf:
        for pn,page in enumerate(pdf.pages):
            ws = wb.active if pn==0 else wb.create_sheet()
            ws.title = f"Page {pn+1}"
            pg_imgs = [im for im in images if im["page"]==pn]
            print(f"  [Page {pn+1}] {page.width:.0f}×{page.height:.0f}")
            write_page(ws, page, pg_imgs)
    wb.save(out_path)
    print(f"\n  ✓  Saved → {out_path}\n")

if __name__=="__main__":
    if len(sys.argv)!=3:
        print("Usage: python pdf_to_excel.py  input.pdf  output.xlsx"); sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"Not found: {sys.argv[1]}"); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
