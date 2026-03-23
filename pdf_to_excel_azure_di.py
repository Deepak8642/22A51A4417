"""
PDF → Excel  |  Azure Document Intelligence Edition
====================================================
Uses Azure Document Intelligence (Form Recognizer) to extract
tables, key-value pairs, and text from PDFs and writes them
to a clean, formatted Excel file — one sheet per PDF page.

KEYS  (set these):
    FORMREC_ENDPOINT  =  https://<your-resource>.cognitiveservices.azure.com/
    FORMREC_KEY       =  <your-32-char-key>

INSTALL:
    pip install azure-ai-formrecognizer openpyxl pdfplumber

RUN:
    python pdf_to_excel_azure_di.py  input.pdf  [output.xlsx]

    # batch — all PDFs in a folder:
    python pdf_to_excel_azure_di.py  folder/  [output_folder/]

OUTPUT FORMAT (matches document_pdf.xlsx style):
    • Each PDF page → one Excel sheet
    • Tables  → rows/columns with header styling
    • Key-value pairs → two-column layout (Key | Value)
    • Raw paragraphs → merged full-width rows
    • Leading zeros and special numbers preserved as text
"""

import sys, os, json, re
from pathlib import Path

# ── Azure Document Intelligence SDK ─────────────────────────────────────────
try:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential
    HAS_DI = True
except ImportError:
    HAS_DI = False
    print("ERROR: pip install azure-ai-formrecognizer")
    sys.exit(1)

# ── Excel writer ─────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════════════════
#  ★  CONFIG — FILL IN YOUR KEYS HERE  ★
# ════════════════════════════════════════════════════════════════════════════
FORMREC_ENDPOINT = os.getenv("FORMREC_ENDPOINT",
    "https://YOUR-RESOURCE.cognitiveservices.azure.com/")

FORMREC_KEY      = os.getenv("FORMREC_KEY",
    "YOUR_FORM_RECOGNIZER_KEY_HERE")

# Model: "prebuilt-layout" gives tables + key-values + paragraphs (best for tax forms)
# Other options: "prebuilt-document", "prebuilt-read" (text only)
DI_MODEL = "prebuilt-layout"

# ── Style helpers ─────────────────────────────────────────────────────────────
def _side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def _bdr(style="thin", color="CCCCCC"):
    s = _side(style, color)
    return Border(top=s, bottom=s, left=s, right=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=9, color="000000"):
    return Font(name="Arial", bold=bold, size=max(int(size), 7), color=color)

def _aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

THIN  = _bdr("thin",   "CCCCCC")
THICK = _bdr("medium", "444444")

# ── Safe merge (skip already-merged cells) ────────────────────────────────────
def safe_merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2:
        return
    try:
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
    except Exception:
        pass

# ── Force text format (preserves leading zeros) ──────────────────────────────
TEXT_FMT = "@"

def set_cell(ws, row, col, value, bold=False, size=9,
             fill_hex=None, align="left", border=THIN, color="000000"):
    cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    cell.number_format = TEXT_FMT
    cell.font          = _font(bold=bold, size=size, color=color)
    cell.alignment     = _aln(align)
    cell.border        = border
    if fill_hex:
        cell.fill = _fill(fill_hex)
    return cell

# ════════════════════════════════════════════════════════════════════════════
#  Azure Document Intelligence — analyze PDF
# ════════════════════════════════════════════════════════════════════════════
def analyze_pdf(pdf_path: str):
    """
    Returns the full AnalyzeResult from Azure DI.
    Sends the file as binary stream (works for local files up to 500 MB).
    """
    client = DocumentAnalysisClient(
        endpoint   = FORMREC_ENDPOINT,
        credential = AzureKeyCredential(FORMREC_KEY)
    )
    with open(pdf_path, "rb") as f:
        poller = client.begin_analyze_document(DI_MODEL, document=f)
    return poller.result()

# ════════════════════════════════════════════════════════════════════════════
#  Build page-level data structures from AnalyzeResult
# ════════════════════════════════════════════════════════════════════════════
def page_number_of(element, result):
    """Return 1-based page number for a content element."""
    if hasattr(element, "bounding_regions") and element.bounding_regions:
        return element.bounding_regions[0].page_number
    return 1

def build_page_data(result):
    """
    Returns dict: { page_num (int): {"tables": [...], "kvpairs": [...], "paragraphs": [...]} }

    tables    — list of 2D lists (rows × cols) of cell content strings
    kvpairs   — list of (key_str, value_str)
    paragraphs— list of paragraph content strings (excluding table/kv content)
    """
    pages = {}
    for pg in result.pages:
        pages[pg.page_number] = {"tables": [], "kvpairs": [], "paragraphs": []}

    # ── Tables ───────────────────────────────────────────────────────────────
    if result.tables:
        for tbl in result.tables:
            pg = page_number_of(tbl, result)
            if pg not in pages:
                pages[pg] = {"tables": [], "kvpairs": [], "paragraphs": []}

            nrows = tbl.row_count
            ncols = tbl.column_count
            grid  = [[""] * ncols for _ in range(nrows)]
            spans = {}  # (row, col) → (row_span, col_span)

            for cell in tbl.cells:
                r, c = cell.row_index, cell.column_index
                grid[r][c] = cell.content or ""
                rs = getattr(cell, "row_span",   1) or 1
                cs = getattr(cell, "column_span", 1) or 1
                if rs > 1 or cs > 1:
                    spans[(r, c)] = (rs, cs)

            pages[pg]["tables"].append({"grid": grid, "spans": spans,
                                        "nrows": nrows, "ncols": ncols})

    # ── Key-Value Pairs ──────────────────────────────────────────────────────
    if hasattr(result, "key_value_pairs") and result.key_value_pairs:
        for kv in result.key_value_pairs:
            pg = 1
            if kv.key and kv.key.bounding_regions:
                pg = kv.key.bounding_regions[0].page_number
            if pg not in pages:
                pages[pg] = {"tables": [], "kvpairs": [], "paragraphs": []}
            key_text = kv.key.content if kv.key else ""
            val_text = kv.value.content if kv.value else ""
            pages[pg]["kvpairs"].append((key_text, val_text))

    # ── Paragraphs ───────────────────────────────────────────────────────────
    if hasattr(result, "paragraphs") and result.paragraphs:
        for para in result.paragraphs:
            pg = page_number_of(para, result)
            if pg not in pages:
                pages[pg] = {"tables": [], "kvpairs": [], "paragraphs": []}
            pages[pg]["paragraphs"].append(para.content or "")

    return pages

# ════════════════════════════════════════════════════════════════════════════
#  Write one page to an Excel worksheet
# ════════════════════════════════════════════════════════════════════════════
def write_page(ws, page_data, page_num):
    """
    Layout per sheet:
      1. Tables — rendered as proper grids with header row styled
      2. Key-Value pairs — two-column layout
      3. Paragraphs — full-width merged rows
    """
    row = 1  # current Excel row

    # ── Section label helper ─────────────────────────────────────────────────
    def section_label(label, n_cols=8):
        cell = set_cell(ws, row, 1, label,
                        bold=True, size=8, fill_hex="1F4E79", color="FFFFFF",
                        border=THICK)
        if n_cols > 1:
            safe_merge(ws, row, 1, row, n_cols)
        ws.row_dimensions[row].height = 14
        return row + 1

    # ────────────────────────────────────────────────────────────────────────
    #  1. TABLES
    # ────────────────────────────────────────────────────────────────────────
    tables = page_data.get("tables", [])
    if tables:
        row = section_label(f"Page {page_num} — Tables ({len(tables)} found)")

        for t_idx, tbl in enumerate(tables):
            grid  = tbl["grid"]
            spans = tbl["spans"]
            nrows = tbl["nrows"]
            ncols = tbl["ncols"]

            # set column widths (shared across all tables on this sheet)
            for ci in range(ncols):
                col_letter = get_column_letter(ci + 1)
                if ws.column_dimensions[col_letter].width < 18:
                    ws.column_dimensions[col_letter].width = 18

            # track merged regions from DI spans
            merged = set()

            for ri in range(nrows):
                ws.row_dimensions[row].height = 15

                for ci in range(ncols):
                    if (ri, ci) in merged:
                        continue

                    val = grid[ri][ci]
                    is_header = (ri == 0)

                    # style
                    fill_hex = "2E75B6" if is_header else ("F2F2F2" if ri % 2 == 0 else "FFFFFF")
                    txt_col  = "FFFFFF" if is_header else "000000"
                    bdr      = THICK if is_header else THIN

                    set_cell(ws, row, ci + 1, val,
                             bold=is_header, size=9,
                             fill_hex=fill_hex, color=txt_col, border=bdr)

                    # apply span merges
                    rs, cs = spans.get((ri, ci), (1, 1))
                    if rs > 1 or cs > 1:
                        end_r = row + rs - 1
                        end_c = ci + cs
                        safe_merge(ws, row, ci + 1, end_r, end_c)
                        for mr in range(ri, ri + rs):
                            for mc in range(ci, ci + cs):
                                if (mr, mc) != (ri, ci):
                                    merged.add((mr, mc))

                row += 1

            row += 1  # blank row between tables

    # ────────────────────────────────────────────────────────────────────────
    #  2. KEY-VALUE PAIRS
    # ────────────────────────────────────────────────────────────────────────
    kvpairs = page_data.get("kvpairs", [])
    if kvpairs:
        row = section_label("Key–Value Pairs", n_cols=8)

        # Header row
        set_cell(ws, row, 1, "Field / Key",   bold=True, fill_hex="D6E4F0", border=THICK)
        set_cell(ws, row, 2, "Value",          bold=True, fill_hex="D6E4F0", border=THICK)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 40)
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 40)
        ws.row_dimensions[row].height = 14
        row += 1

        for ki, (key, val) in enumerate(kvpairs):
            fill_hex = "F9F9F9" if ki % 2 == 0 else "FFFFFF"
            set_cell(ws, row, 1, key, fill_hex=fill_hex, border=THIN)
            set_cell(ws, row, 2, val, fill_hex=fill_hex, border=THIN)
            ws.row_dimensions[row].height = 13
            row += 1

        row += 1

    # ────────────────────────────────────────────────────────────────────────
    #  3. PARAGRAPHS / RAW TEXT
    # ────────────────────────────────────────────────────────────────────────
    paragraphs = page_data.get("paragraphs", [])
    if paragraphs:
        row = section_label("Text / Paragraphs", n_cols=8)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 80)

        for para in paragraphs:
            if not para.strip():
                continue
            cell = set_cell(ws, row, 1, para, size=9, border=THIN,
                            fill_hex="FFFFFF" if row % 2 else "F7F7F7")
            safe_merge(ws, row, 1, row, 8)
            ws.row_dimensions[row].height = max(13, min(60, 13 * para.count("\n") + 13))
            row += 1

    return row

# ════════════════════════════════════════════════════════════════════════════
#  Convert a single PDF → Excel
# ════════════════════════════════════════════════════════════════════════════
def convert_pdf(pdf_path: str, out_path: str):
    print(f"\n{'='*60}")
    print(f"PDF → Excel  |  Azure Document Intelligence")
    print(f"Input : {pdf_path}")
    print(f"Output: {out_path}")
    print(f"Model : {DI_MODEL}")
    print(f"{'='*60}")

    print("  Sending to Azure Document Intelligence...", flush=True)
    result = analyze_pdf(pdf_path)
    print(f"  ✓ Analysis complete  ({len(result.pages)} pages)")

    page_data = build_page_data(result)
    wb = Workbook()

    for page_num in sorted(page_data.keys()):
        ws = wb.active if page_num == 1 else wb.create_sheet()
        ws.title = f"Page {page_num}"

        pdata = page_data[page_num]
        n_tbl = len(pdata["tables"])
        n_kv  = len(pdata["kvpairs"])
        n_par = len(pdata["paragraphs"])
        print(f"  [Page {page_num}]  tables={n_tbl}  kv-pairs={n_kv}  paragraphs={n_par}")

        write_page(ws, pdata, page_num)

    # Remove default empty sheet if we created pages
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(out_path)
    print(f"\n✅  Saved → {out_path}\n")

# ════════════════════════════════════════════════════════════════════════════
#  Batch convert: folder of PDFs
# ════════════════════════════════════════════════════════════════════════════
def convert_folder(in_dir: str, out_dir: str):
    in_p  = Path(in_dir)
    out_p = Path(out_dir)
    out_p.mkdir(parents=True, exist_ok=True)

    pdfs = sorted(in_p.glob("*.pdf"))
    if not pdfs:
        print(f"No PDF files found in {in_dir}")
        return

    print(f"Found {len(pdfs)} PDF(s) in {in_dir}")
    for pdf in pdfs:
        out_file = out_p / (pdf.stem + ".xlsx")
        try:
            convert_pdf(str(pdf), str(out_file))
        except Exception as e:
            print(f"  ✗ ERROR processing {pdf.name}: {e}")

# ════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print(__doc__)
        print("\nUsage:")
        print("  python pdf_to_excel_azure_di.py  input.pdf  [output.xlsx]")
        print("  python pdf_to_excel_azure_di.py  folder/    [output_folder/]")
        sys.exit(1)

    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) == 3 else None

    # Validate keys
    if "YOUR-RESOURCE" in FORMREC_ENDPOINT or "YOUR_FORM" in FORMREC_KEY:
        print("\n⚠  ERROR: Please set FORMREC_ENDPOINT and FORMREC_KEY in the script")
        print("   or export them as environment variables:")
        print("     export FORMREC_ENDPOINT=https://your-resource.cognitiveservices.azure.com/")
        print("     export FORMREC_KEY=your_key_here")
        sys.exit(1)

    if os.path.isdir(inp):
        out_dir = out or (inp.rstrip("/\\") + "_excel_output")
        convert_folder(inp, out_dir)
    elif os.path.isfile(inp) and inp.lower().endswith(".pdf"):
        out_file = out or (os.path.splitext(inp)[0] + ".xlsx")
        convert_pdf(inp, out_file)
    else:
        print(f"ERROR: '{inp}' is not a valid PDF file or directory.")
        sys.exit(1)
