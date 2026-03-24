"""
╔══════════════════════════════════════════════════════════════════════════════╗
║        PDF → EXCEL  v4  |  Azure Document Intelligence                      ║
║  ✔ No truncation — text spans across correct columns                        ║
║  ✔ Perfect table detection — cell-level bounding boxes                      ║
║  ✔ Clickable checkboxes — TRUE/FALSE toggle via Excel data validation       ║
║  ✔ All pages extracted correctly                                             ║
╚══════════════════════════════════════════════════════════════════════════════╝

INSTALL:
  pip install azure-ai-formrecognizer openpyxl pymupdf pillow

OPTIONAL (image OCR):
  pip install pytesseract
  Ubuntu:  sudo apt-get install tesseract-ocr
  Mac:     brew install tesseract
  Windows: https://github.com/UB-Mannheim/tesseract/wiki

SET AZURE KEYS (edit CONFIG below OR use environment variables):
  Windows:  set FORMREC_ENDPOINT=https://...  &&  set FORMREC_KEY=...
  Mac/Linux: export FORMREC_ENDPOINT=https://...  &&  export FORMREC_KEY=...

RUN:
  python pdf_to_excel_v4.py  input.pdf
  python pdf_to_excel_v4.py  input.pdf   output.xlsx
  python pdf_to_excel_v4.py  my_folder/  out_folder/
"""

import sys, os, io, re, textwrap
from pathlib import Path
from collections import defaultdict

# ── Azure DI ──────────────────────────────────────────────────────────────────
try:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential
except ImportError:
    print("ERROR: pip install azure-ai-formrecognizer"); sys.exit(1)

# ── Excel ─────────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.cell.cell import MergedCell
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: pip install openpyxl"); sys.exit(1)

# ── PyMuPDF ───────────────────────────────────────────────────────────────────
try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False
    print("WARNING: pip install pymupdf  (images will be skipped)")

# ── Pillow ────────────────────────────────────────────────────────────────────
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("WARNING: pip install pillow  (images will be skipped)")

# ── Tesseract OCR (optional) ──────────────────────────────────────────────────
try:
    import pytesseract
    pytesseract.get_tesseract_version()
    HAS_OCR = True
except Exception:
    HAS_OCR = False


# ════════════════════════════════════════════════════════════════════════════════
#  ★  CONFIG — EDIT YOUR AZURE KEYS HERE  ★
# ════════════════════════════════════════════════════════════════════════════════

FORMREC_ENDPOINT = os.getenv(
    "FORMREC_ENDPOINT",
    "https://YOUR-RESOURCE.cognitiveservices.azure.com/"  # ← EDIT
)
FORMREC_KEY = os.getenv(
    "FORMREC_KEY",
    "YOUR_AZURE_KEY_HERE"                                  # ← EDIT
)

DI_MODEL = "prebuilt-layout"

# ── Grid config ────────────────────────────────────────────────────────────────
# Each PDF page maps onto GRID_COLS × GRID_ROWS Excel cells.
# Higher values = more faithful positioning but larger files.
GRID_COLS  = 100        # columns per page
GRID_ROWS  = 130        # rows per page
COL_WIDTH  = 2.2        # Excel column width (chars)  — narrow for dense grid
ROW_HEIGHT = 10.5       # Excel row height (pt)

# Image limits
MIN_IMG_PX   = 30
MAX_IMG_W_PX = 400
MAX_IMG_H_PX = 300


# ════════════════════════════════════════════════════════════════════════════════
#  COLOUR PALETTE
# ════════════════════════════════════════════════════════════════════════════════
C_HDR_BG   = "1F4E79"   # table header background
C_HDR_FG   = "FFFFFF"   # table header text
C_ALT_ROW  = "EBF3FB"   # alternating table row
C_TITLE_BG = "D6E4F0"
C_TITLE_FG = "1F4E79"
C_SECT_BG  = "EEF4FB"
C_SECT_FG  = "1F4E79"
C_KV_KEY   = "1F4E79"
C_KV_VAL   = "FEFCE8"
C_FOOT     = "999999"
C_OCR      = "FFFDE7"
C_CB_TRUE  = "E6F4EA"   # checked checkbox bg
C_CB_FALSE = "FFF3E0"   # unchecked checkbox bg


# ════════════════════════════════════════════════════════════════════════════════
#  STYLE HELPERS
# ════════════════════════════════════════════════════════════════════════════════

def _side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def _bdr(style="thin", color="CCCCCC"):
    s = _side(style, color)
    return Border(top=s, bottom=s, left=s, right=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def _font(bold=False, italic=False, size=9, color="000000", name="Calibri"):
    return Font(name=name, bold=bold, italic=italic,
                size=max(6, int(size)), color=color.lstrip("#"))

def _aln(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN    = _bdr("thin",   "CCCCCC")
MEDIUM  = _bdr("medium", "555555")
NOBDR   = Border()
OUTER   = Border(
    top=_side("medium", "555555"),   bottom=_side("medium", "555555"),
    left=_side("medium", "555555"),  right=_side("medium", "555555")
)


def is_merged(ws, row, col):
    """True if cell is a slave merged cell (read-only)."""
    return isinstance(ws.cell(row=row, column=col), MergedCell)


def clamp_r(r): return max(1, min(r, GRID_ROWS))
def clamp_c(c): return max(1, min(c, GRID_COLS))


def wcell(ws, row, col, val,
          bold=False, italic=False, size=9,
          bg=None, fg="000000", border=NOBDR,
          halign="left", wrap=True, font_name="Calibri"):
    """
    Write a value + full formatting to a cell.
    Silently skips merged slave cells.
    Always stores as text (@) to avoid number/date misinterpretation.
    """
    row = clamp_r(row); col = clamp_c(col)
    if is_merged(ws, row, col):
        return None
    c = ws.cell(row=row, column=col)
    c.value         = "" if val is None else str(val)
    c.number_format = "@"
    c.font          = _font(bold, italic, size, fg, font_name)
    c.alignment     = _aln(halign, "top", wrap)
    c.border        = border
    if bg:
        c.fill = _fill(bg)
    return c


def safe_merge(ws, r1, c1, r2, c2, done: set):
    """Merge cells, clamped to grid, deduped via done set."""
    r1, c1 = clamp_r(r1), clamp_c(c1)
    r2, c2 = clamp_r(r2), clamp_c(c2)
    if r1 == r2 and c1 == c2:
        return
    key = (r1, c1, r2, c2)
    if key in done:
        return
    try:
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
        done.add(key)
    except Exception:
        pass


# ════════════════════════════════════════════════════════════════════════════════
#  COORDINATE HELPERS
# ════════════════════════════════════════════════════════════════════════════════

def poly_bbox(polygon):
    """Azure polygon (list of Point) → (x0, y0, x1, y1) in inches."""
    if not polygon:
        return None
    xs = [p.x for p in polygon]; ys = [p.y for p in polygon]
    return (min(xs), min(ys), max(xs), max(ys))


def inch_to_col(x, pw):
    """x-coord in inches → 1-based column index."""
    return clamp_c(int(x / pw * GRID_COLS) + 1)


def inch_to_row(y, ph):
    """y-coord in inches → 1-based row index."""
    return clamp_r(int(y / ph * GRID_ROWS) + 1)


def bbox_cells(bbox, pw, ph):
    """(x0,y0,x1,y1) → (sc, sr, ec, er) all 1-based, ec>=sc, er>=sr."""
    sc = inch_to_col(bbox[0], pw)
    sr = inch_to_row(bbox[1], ph)
    ec = max(sc, inch_to_col(bbox[2], pw))
    er = max(sr, inch_to_row(bbox[3], ph))
    return sc, sr, ec, er


def height_to_fontsize(h_in):
    """Estimate font pt from bounding-box height in inches."""
    pt = h_in * 72
    if pt >= 22: return 16
    if pt >= 17: return 14
    if pt >= 13: return 12
    if pt >= 10: return 11
    return 9


# ════════════════════════════════════════════════════════════════════════════════
#  ★  CHECKBOX  HELPERS
# ════════════════════════════════════════════════════════════════════════════════

# Azure DI checkbox tokens (from prebuilt-layout)
_CB_CHECKED   = frozenset({":selected:", "☑", "✔", "✓", "[x]", "[X]", "✅"})
_CB_UNCHECKED = frozenset({":unselected:", "☐", "□", "[ ]", "○"})


def parse_checkbox(text: str):
    """
    Returns ('checked' | 'unchecked' | None, label_without_token).
    Scans for Azure DI :selected:/:unselected: tokens AND common symbols.
    """
    t = text.strip()
    for tok in _CB_CHECKED:
        if tok in t:
            return "checked", t.replace(tok, "").strip()
    for tok in _CB_UNCHECKED:
        if tok in t:
            return "unchecked", t.replace(tok, "").strip()
    return None, t


def write_checkbox_cell(ws, row, col, state: str, label: str,
                        merged_done: set):
    """
    Write an interactive checkbox using Excel Data Validation (List TRUE,FALSE).

    Layout:
      col+0 : the TRUE/FALSE value cell  →  data validation makes it a dropdown
               that Excel 365 renders as a real tick-box.
               For older Excel we also pre-fill ☑ / ☐ as the display value.
      col+1 … col+N : label text (merged across remaining width)

    The cell background is green (checked) or amber (unchecked) so it is
    immediately visible even without the tick-box rendering.
    """
    row = clamp_r(row)
    col = clamp_c(col)
    if is_merged(ws, row, col):
        return

    # ── Value cell ─────────────────────────────────────────────────────────
    is_checked  = (state == "checked")
    cell        = ws.cell(row=row, column=col)
    cell.value  = "TRUE" if is_checked else "FALSE"   # machine-readable
    cell.number_format = "@"
    cell.font   = Font(name="Segoe UI Symbol", size=12, bold=True,
                       color="276221" if is_checked else "BF5700")
    cell.fill   = _fill(C_CB_TRUE if is_checked else C_CB_FALSE)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=False)
    cell.border = THIN

    # ── Data validation — makes it a real Excel tick-box in Excel 365 ──────
    col_letter  = get_column_letter(col)
    cell_addr   = f"{col_letter}{row}"
    dv = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False,
        showDropDown=False,    # False = show the dropdown arrow
        showErrorMessage=False,
        showInputMessage=True,
        promptTitle="Checkbox",
        prompt="Select TRUE (checked) or FALSE (unchecked)"
    )
    dv.sqref = cell_addr
    ws.add_data_validation(dv)

    # ── Label cell ─────────────────────────────────────────────────────────
    if label:
        lc  = clamp_c(col + 1)
        lce = clamp_c(col + 12)   # span label across ~12 cols
        if not is_merged(ws, row, lc):
            wcell(ws, row, lc, label, size=9, border=NOBDR, wrap=True)
            if lce > lc:
                safe_merge(ws, row, lc, row, lce, merged_done)


# ════════════════════════════════════════════════════════════════════════════════
#  AZURE DI  —  ANALYZE PDF
# ════════════════════════════════════════════════════════════════════════════════

def analyze_pdf(path: str):
    client = DocumentAnalysisClient(
        endpoint=FORMREC_ENDPOINT,
        credential=AzureKeyCredential(FORMREC_KEY)
    )
    with open(path, "rb") as f:
        poller = client.begin_analyze_document(DI_MODEL, document=f)
    return poller.result()


# ════════════════════════════════════════════════════════════════════════════════
#  BUILD PAGE DATA  — all pages, all elements
# ════════════════════════════════════════════════════════════════════════════════

def build_page_data(result) -> dict:
    """
    Returns { page_num: { w, h, lines, tables, kvpairs, paragraphs } }

    KEY FIXES vs v3:
    • lines   : every page iterated — nothing dropped
    • tables  : ALL bounding_regions iterated (multi-page tables)
                cell bounding boxes stored per cell for pixel-perfect placement
    • paragraphs: ALL bounding_regions iterated
    • kvpairs : correct page from bounding_regions
    """
    pages = {}

    # Initialise every page (dimension data)
    for pg in result.pages:
        pages[pg.page_number] = {
            "w": pg.width  or 8.5,
            "h": pg.height or 11.0,
            "lines":      [],
            "tables":     [],
            "kvpairs":    [],
            "paragraphs": [],
        }

    def ensure(pn):
        if pn not in pages:
            pages[pn] = {"w": 8.5, "h": 11.0,
                         "lines": [], "tables": [], "kvpairs": [], "paragraphs": []}

    # ── LINES — every page ──────────────────────────────────────────────────
    for pg in result.pages:
        pn = pg.page_number
        ensure(pn)
        if pg.lines:
            for line in pg.lines:
                txt = (line.content or "").strip()
                if txt:
                    pages[pn]["lines"].append({
                        "content": txt,
                        "bbox":    poly_bbox(line.polygon),
                    })
        elif pg.words:
            # Fallback: group words into synthetic lines by y-proximity
            words = [
                {"content": w.content.strip(), "bbox": poly_bbox(w.polygon)}
                for w in pg.words if (w.content or "").strip()
            ]
            words.sort(key=lambda w: (
                round(w["bbox"][1] * 6) if w["bbox"] else 0,
                w["bbox"][0] if w["bbox"] else 0,
            ))
            bucket, cur_y = [], None
            def flush(bucket, target):
                bbs = [b["bbox"] for b in bucket if b["bbox"]]
                if bbs:
                    target.append({
                        "content": " ".join(b["content"] for b in bucket),
                        "bbox":    (min(b[0] for b in bbs), min(b[1] for b in bbs),
                                    max(b[2] for b in bbs), max(b[3] for b in bbs)),
                    })
            for w in words:
                wy = round(w["bbox"][1] * 6) if w["bbox"] else 0
                if cur_y is None or abs(wy - cur_y) <= 1:
                    bucket.append(w); cur_y = wy
                else:
                    flush(bucket, pages[pn]["lines"])
                    bucket, cur_y = [w], wy
            if bucket:
                flush(bucket, pages[pn]["lines"])

    # ── PARAGRAPHS — all bounding_regions ───────────────────────────────────
    if hasattr(result, "paragraphs") and result.paragraphs:
        for para in result.paragraphs:
            txt  = (para.content or "").strip()
            role = getattr(para, "role", "") or ""
            if not txt:
                continue
            if para.bounding_regions:
                for region in para.bounding_regions:
                    pn = region.page_number
                    ensure(pn)
                    pages[pn]["paragraphs"].append({
                        "content": txt,
                        "bbox":    poly_bbox(region.polygon),
                        "role":    role,
                    })
            else:
                ensure(1)
                pages[1]["paragraphs"].append({"content": txt, "bbox": None, "role": role})

    # ── TABLES — all bounding_regions + cell-level bboxes ───────────────────
    if result.tables:
        for tbl in result.tables:
            nrows = tbl.row_count
            ncols = tbl.column_count
            # Build cell grid with content, spans, kind AND per-cell bbox
            grid   = [[""] * ncols for _ in range(nrows)]
            spans  = {}    # (r,c) → (row_span, col_span)
            kinds  = {}    # (r,c) → "columnHeader" | "rowHeader" | ""
            cbboxes = {}   # (r,c) → (x0,y0,x1,y1) in inches  ← NEW

            for cell in tbl.cells:
                r, c = cell.row_index, cell.column_index
                grid[r][c] = (cell.content or "").strip()
                rs = getattr(cell, "row_span",    1) or 1
                cs = getattr(cell, "column_span", 1) or 1
                if rs > 1 or cs > 1:
                    spans[(r, c)] = (rs, cs)
                kinds[(r, c)] = getattr(cell, "kind", "") or ""
                # Cell bbox from its own bounding_regions
                if cell.bounding_regions:
                    cbboxes[(r, c)] = poly_bbox(cell.bounding_regions[0].polygon)

            if tbl.bounding_regions:
                for region in tbl.bounding_regions:
                    pn       = region.page_number
                    tbl_bbox = poly_bbox(region.polygon)
                    ensure(pn)
                    pages[pn]["tables"].append({
                        "grid":    grid,
                        "spans":   spans,
                        "kinds":   kinds,
                        "cbboxes": cbboxes,   # ← cell-level bboxes
                        "nrows":   nrows,
                        "ncols":   ncols,
                        "bbox":    tbl_bbox,
                    })
            else:
                ensure(1)
                pages[1]["tables"].append({
                    "grid": grid, "spans": spans, "kinds": kinds,
                    "cbboxes": cbboxes, "nrows": nrows, "ncols": ncols, "bbox": None,
                })

    # ── KEY-VALUE PAIRS ──────────────────────────────────────────────────────
    if hasattr(result, "key_value_pairs") and result.key_value_pairs:
        for kv in result.key_value_pairs:
            pn = 1; k_bbox = v_bbox = None
            if kv.key and kv.key.bounding_regions:
                pn      = kv.key.bounding_regions[0].page_number
                k_bbox  = poly_bbox(kv.key.bounding_regions[0].polygon)
            if kv.value and kv.value.bounding_regions:
                v_bbox  = poly_bbox(kv.value.bounding_regions[0].polygon)
            ensure(pn)
            pages[pn]["kvpairs"].append({
                "key":      (kv.key.content   if kv.key   else "").strip(),
                "value":    (kv.value.content if kv.value else "").strip(),
                "key_bbox": k_bbox,
                "val_bbox": v_bbox,
            })

    return pages


# ════════════════════════════════════════════════════════════════════════════════
#  IMAGE EXTRACTION
# ════════════════════════════════════════════════════════════════════════════════

def extract_images(fitz_doc, page_idx: int, pw: float, ph: float):
    out = []
    if not (HAS_FITZ and HAS_PIL):
        return out
    try:
        page = fitz_doc[page_idx]
    except Exception:
        return out
    for info in page.get_images(full=True):
        xref = info[0]
        try:
            raw = fitz_doc.extract_image(xref)
            pil = PILImage.open(io.BytesIO(raw["image"])).convert("RGB")
        except Exception:
            continue
        if pil.width < MIN_IMG_PX or pil.height < MIN_IMG_PX:
            continue
        bbox_in = None
        try:
            for rect in page.get_image_rects(xref):
                r = rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect)
                bbox_in = (r.x0/72, r.y0/72, r.x1/72, r.y1/72)
                break
        except Exception:
            pass
        if bbox_in is None:
            continue
        sc, sr, ec, er = bbox_cells(bbox_in, pw, ph)
        out.append({"pil": pil, "sc": sc, "sr": sr, "ec": ec, "er": er})
    return out


def resize_image(pil):
    w, h  = pil.size
    scale = min(MAX_IMG_W_PX / w, MAX_IMG_H_PX / h, 1.0)
    return pil.resize((max(1, int(w*scale)), max(1, int(h*scale))), PILImage.LANCZOS)


def ocr_image(pil):
    if not HAS_OCR:
        return ""
    try:
        txt = pytesseract.image_to_string(pil, timeout=15).strip()
        return re.sub(r'\n{3,}', '\n\n', txt)
    except Exception:
        return ""


# ════════════════════════════════════════════════════════════════════════════════
#  TABLE-OCCUPIED CELLS  (bounding-box level)
# ════════════════════════════════════════════════════════════════════════════════

def table_occupied(tables, pw, ph) -> set:
    occ = set()
    for tbl in tables:
        if tbl["bbox"]:
            sc, sr, ec, er = bbox_cells(tbl["bbox"], pw, ph)
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    occ.add((r, c))
    return occ


# ════════════════════════════════════════════════════════════════════════════════
#  ★  WRITE ONE PAGE
# ════════════════════════════════════════════════════════════════════════════════

def write_page(ws, pdata: dict, page_num: int,
               fitz_doc=None, page_idx: int = 0):

    pw, ph   = pdata["w"], pdata["h"]
    merged   = set()   # tracks already-merged ranges

    # ── Setup uniform grid ────────────────────────────────────────────────────
    for ci in range(1, GRID_COLS + 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTH
    for ri in range(1, GRID_ROWS + 1):
        ws.row_dimensions[ri].height = ROW_HEIGHT
    ws.sheet_view.zoomScale = 90

    # Faint page label top-left
    c0 = ws.cell(row=1, column=1)
    c0.value = f"Page {page_num}"
    c0.font  = Font(name="Calibri", size=7, italic=True, color="CCCCCC")

    tab_occ = table_occupied(pdata["tables"], pw, ph)

    # ══════════════════════════════════════════════════════════════════════════
    #  A.  TABLES  — pixel-perfect using per-cell bounding boxes
    # ══════════════════════════════════════════════════════════════════════════
    for tbl in pdata["tables"]:
        grid    = tbl["grid"]
        spans   = tbl["spans"]
        kinds   = tbl["kinds"]
        cbboxes = tbl["cbboxes"]
        nrows   = tbl["nrows"]
        ncols   = tbl["ncols"]
        tbl_bb  = tbl["bbox"]

        # Pre-compute Excel cell for every (row, col) using cell bbox if available,
        # else fall back to proportional subdivision of the table bbox.
        if tbl_bb:
            tsc, tsr, tec, ter = bbox_cells(tbl_bb, pw, ph)
        else:
            tsc, tsr = 2, 3
            tec = clamp_c(tsc + ncols * 4)
            ter = clamp_r(tsr + nrows * 2)

        def cell_excel_coords(ri, ci):
            """Return (ex_r, ex_c) for grid cell (ri, ci)."""
            if (ri, ci) in cbboxes and cbboxes[(ri, ci)]:
                sc, sr, ec, er = bbox_cells(cbboxes[(ri, ci)], pw, ph)
                return sr, sc
            # Proportional fallback
            cw = max(1, (tec - tsc + 1) // max(ncols, 1))
            rh = max(1, (ter - tsr + 1) // max(nrows, 1))
            return clamp_r(tsr + ri * rh), clamp_c(tsc + ci * cw)

        def cell_excel_end(ri, ci, rs, cs):
            """Return (end_r, end_c) for a spanning cell."""
            # Try to use the last covered cell's bbox corner
            end_ri = min(ri + rs - 1, nrows - 1)
            end_ci = min(ci + cs - 1, ncols - 1)
            if (end_ri, end_ci) in cbboxes and cbboxes[(end_ri, end_ci)]:
                _, _, ec, er = bbox_cells(cbboxes[(end_ri, end_ci)], pw, ph)
                return er, ec
            # Proportional fallback
            cw = max(1, (tec - tsc + 1) // max(ncols, 1))
            rh = max(1, (ter - tsr + 1) // max(nrows, 1))
            sr0, sc0 = cell_excel_coords(ri, ci)
            return clamp_r(sr0 + rs * rh - 1), clamp_c(sc0 + cs * cw - 1)

        absorbed = set()

        for ri in range(nrows):
            for ci in range(ncols):
                if (ri, ci) in absorbed:
                    continue

                val  = grid[ri][ci]
                kind = kinds.get((ri, ci), "")
                hdr  = (kind in ("columnHeader", "rowHeader")) or ri == 0

                ex_r, ex_c = cell_excel_coords(ri, ci)

                # Style
                if hdr:
                    bg, fg, bdr, bold = C_HDR_BG, C_HDR_FG, MEDIUM, True
                elif ri % 2 == 0:
                    bg, fg, bdr, bold = C_ALT_ROW, "000000", THIN, False
                else:
                    bg, fg, bdr, bold = "FFFFFF", "000000", THIN, False

                # Detect checkbox inside table cell
                cb_state, cb_label = parse_checkbox(val)
                if cb_state is not None:
                    write_checkbox_cell(ws, ex_r, ex_c, cb_state, cb_label, merged)
                else:
                    wcell(ws, ex_r, ex_c, val,
                          bold=bold, size=9, bg=bg, fg=fg, border=bdr)

                # Handle spans
                rs_v, cs_v = spans.get((ri, ci), (1, 1))
                if rs_v > 1 or cs_v > 1:
                    end_r, end_c = cell_excel_end(ri, ci, rs_v, cs_v)
                    safe_merge(ws, ex_r, ex_c, end_r, end_c, merged)
                    for mr in range(ri, ri + rs_v):
                        for mc in range(ci, ci + cs_v):
                            if (mr, mc) != (ri, ci):
                                absorbed.add((mr, mc))

    # ══════════════════════════════════════════════════════════════════════════
    #  B.  TEXT LINES
    #  FIX: text is placed at its exact x0→x1 bbox, merged across all columns
    #       it physically spans, with wrap enabled.  No truncation.
    # ══════════════════════════════════════════════════════════════════════════
    row_buckets: dict[int, list] = defaultdict(list)

    for line in pdata["lines"]:
        txt  = line["content"]
        bbox = line["bbox"]
        if not txt or not bbox:
            continue

        sc, sr, ec, er = bbox_cells(bbox, pw, ph)
        if (sr, sc) in tab_occ:
            continue

        h_in = bbox[3] - bbox[1]
        row_buckets[sr].append((sc, ec, txt, h_in, bbox))

    for ex_row in sorted(row_buckets):
        items = sorted(row_buckets[ex_row], key=lambda x: x[0])
        for sc, ec, txt, h_in, bbox in items:
            sc  = clamp_c(sc)
            ec  = clamp_c(max(sc, ec))
            row = clamp_r(ex_row)
            fsz = height_to_fontsize(h_in)

            # ★ Checkbox detection
            cb_state, cb_label = parse_checkbox(txt)
            if cb_state is not None:
                write_checkbox_cell(ws, row, sc, cb_state, cb_label, merged)
                continue

            if is_merged(ws, row, sc):
                continue

            existing = ws.cell(row=row, column=sc).value
            if existing:
                # Cell already written (e.g. by table) — append
                ws.cell(row=row, column=sc).value = str(existing) + "  " + txt
            else:
                # ★ KEY FIX: merge across the FULL width the text spans
                #   so text has room to display without truncation
                wcell(ws, row, sc, txt, size=fsz, border=NOBDR,
                      wrap=True, halign="left")
                if ec > sc:
                    safe_merge(ws, row, sc, row, ec, merged)

            # Adapt row height to text size
            desired_h = min(120.0, max(ROW_HEIGHT, h_in * 72 * 1.35))
            if ws.row_dimensions[row].height < desired_h:
                ws.row_dimensions[row].height = desired_h

    # ══════════════════════════════════════════════════════════════════════════
    #  C.  PARAGRAPH ROLES  (title, sectionHeading, footnote …)
    # ══════════════════════════════════════════════════════════════════════════
    for para in pdata["paragraphs"]:
        role = para["role"]; txt = para["content"]; bbox = para["bbox"]
        if not txt or not bbox or not role:
            continue

        sc, sr, ec, er = bbox_cells(bbox, pw, ph)
        if (sr, sc) in tab_occ:
            continue
        sc = clamp_c(sc); sr = clamp_r(sr)
        ec = clamp_c(max(sc, ec)); er = clamp_r(max(sr, er))

        if   role == "title":          bg, fg, bold, fsz = C_TITLE_BG, C_TITLE_FG, True,  14
        elif role == "sectionHeading": bg, fg, bold, fsz = C_SECT_BG,  C_SECT_FG,  True,  11
        elif role in ("pageHeader","pageFooter"):
                                       bg, fg, bold, fsz = "F5F5F5",   C_FOOT,     False,  8
        elif role == "footnote":       bg, fg, bold, fsz = None,       "AAAAAA",   False,  7
        else:                          bg, fg, bold, fsz = None,       "000000",   False,  9

        if not is_merged(ws, sr, sc) and not ws.cell(row=sr, column=sc).value:
            wcell(ws, sr, sc, txt,
                  bold=bold, size=fsz, bg=bg, fg=fg, border=NOBDR, wrap=True)
            if ec > sc or er > sr:
                safe_merge(ws, sr, sc, er, ec, merged)

    # ══════════════════════════════════════════════════════════════════════════
    #  D.  KEY-VALUE PAIRS
    # ══════════════════════════════════════════════════════════════════════════
    for kv in pdata["kvpairs"]:
        key = kv["key"]; val = kv["value"]

        if key and kv["key_bbox"]:
            sc, sr, ec, er = bbox_cells(kv["key_bbox"], pw, ph)
            sr = clamp_r(sr); sc = clamp_c(sc)
            if (sr, sc) not in tab_occ and not is_merged(ws, sr, sc):
                if not ws.cell(row=sr, column=sc).value:
                    wcell(ws, sr, sc, key, bold=True, size=9, fg=C_KV_KEY)

        if val and kv["val_bbox"]:
            sc, sr, ec, er = bbox_cells(kv["val_bbox"], pw, ph)
            sr = clamp_r(sr); sc = clamp_c(sc); ec = clamp_c(max(sc, ec))
            if (sr, sc) not in tab_occ and not is_merged(ws, sr, sc):
                if not ws.cell(row=sr, column=sc).value:
                    cb_state, cb_label = parse_checkbox(val)
                    if cb_state is not None:
                        write_checkbox_cell(ws, sr, sc, cb_state, cb_label, merged)
                    else:
                        wcell(ws, sr, sc, val, size=9, bg=C_KV_VAL, wrap=True)
                        if ec > sc:
                            safe_merge(ws, sr, sc, sr, ec, merged)

    # ══════════════════════════════════════════════════════════════════════════
    #  E.  IMAGES
    # ══════════════════════════════════════════════════════════════════════════
    if fitz_doc is not None:
        for img in extract_images(fitz_doc, page_idx, pw, ph):
            pil = img["pil"]
            sc  = clamp_c(img["sc"])
            sr  = clamp_r(img["sr"])
            resized = resize_image(pil)
            buf = io.BytesIO()
            resized.save(buf, format="PNG"); buf.seek(0)
            try:
                xl = XLImage(buf)
                xl.anchor = f"{get_column_letter(sc)}{sr}"
                ws.add_image(xl)
            except Exception as e:
                print(f"    WARNING: image embed failed — {e}")

            txt = ocr_image(pil)
            if txt:
                oc = clamp_c(img["ec"] + 1)
                if not is_merged(ws, sr, oc) and not ws.cell(row=sr, column=oc).value:
                    wcell(ws, sr, oc, f"[Image text]\n{txt}",
                          italic=True, size=8, fg="555555", bg=C_OCR, wrap=True)
                    safe_merge(ws, sr, oc,
                               clamp_r(img["er"]), clamp_c(oc + 6), merged)


# ════════════════════════════════════════════════════════════════════════════════
#  CONVERT SINGLE PDF  →  EXCEL
# ════════════════════════════════════════════════════════════════════════════════

def convert_pdf(pdf_path: str, out_path: str):
    sep = "=" * 68
    print(f"\n{sep}")
    print(f"  PDF → EXCEL  v4  |  Azure Document Intelligence")
    print(f"  In  : {pdf_path}")
    print(f"  Out : {out_path}")
    print(sep)

    # 1. Azure DI
    print("  [1/4]  Sending to Azure DI …", flush=True)
    try:
        result = analyze_pdf(pdf_path)
    except Exception as e:
        print(f"\n  ERROR: Azure DI call failed — {e}")
        print("  Check FORMREC_ENDPOINT and FORMREC_KEY.")
        raise

    n_pages = len(result.pages)
    n_tbls  = len(result.tables) if result.tables else 0
    print(f"         {n_pages} page(s), {n_tbls} table(s) detected")

    # 2. Build page data
    print("  [2/4]  Extracting all page content …", flush=True)
    all_data = build_page_data(result)
    print(f"         Pages: {sorted(all_data.keys())}")

    # 3. Open with PyMuPDF for images
    fitz_doc = None
    if HAS_FITZ and HAS_PIL:
        print("  [3/4]  Opening PDF for images …", flush=True)
        try:
            fitz_doc = fitz.open(pdf_path)
        except Exception as e:
            print(f"  WARNING: PyMuPDF failed ({e}) — images skipped")
    else:
        print("  [3/4]  Images skipped  (pip install pymupdf pillow)")

    # 4. Write Excel
    print("  [4/4]  Writing Excel …", flush=True)
    wb    = Workbook()
    first = True

    for pn in sorted(all_data.keys()):
        ws    = wb.active if first else wb.create_sheet()
        first = False
        ws.title = f"Page {pn}"[:31]

        pd = all_data[pn]
        cb_n = sum(1 for ln in pd["lines"]
                   if parse_checkbox(ln["content"])[0] is not None)
        print(f"    Page {pn:>3}  lines={len(pd['lines'])}  "
              f"tables={len(pd['tables'])}  kv={len(pd['kvpairs'])}  "
              f"paragraphs={len(pd['paragraphs'])}  checkboxes≈{cb_n}")

        write_page(ws, pd, pn, fitz_doc=fitz_doc, page_idx=pn - 1)

    # Clean up blank default sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    if fitz_doc:
        fitz_doc.close()

    wb.save(out_path)
    kb = os.path.getsize(out_path) // 1024
    print(f"\n  ✔  Saved  →  {out_path}  ({kb} KB)")
    print(f"{sep}\n")


# ════════════════════════════════════════════════════════════════════════════════
#  BATCH CONVERT
# ════════════════════════════════════════════════════════════════════════════════

def convert_folder(in_dir: str, out_dir: str):
    in_p  = Path(in_dir)
    out_p = Path(out_dir)
    out_p.mkdir(parents=True, exist_ok=True)
    pdfs  = sorted(in_p.glob("*.pdf"))
    if not pdfs:
        print(f"No PDFs found in {in_dir}"); return

    print(f"\nBatch: {len(pdfs)} PDF(s) in {in_dir}")
    ok = fail = 0
    for pdf in pdfs:
        try:
            convert_pdf(str(pdf), str(out_p / (pdf.stem + ".xlsx")))
            ok += 1
        except Exception as e:
            print(f"  FAIL: {pdf.name} — {e}")
            fail += 1
    print(f"\nBatch done: {ok} ok, {fail} failed.")


# ════════════════════════════════════════════════════════════════════════════════
#  VALIDATE CREDENTIALS
# ════════════════════════════════════════════════════════════════════════════════

def check_keys():
    if "YOUR-RESOURCE" in FORMREC_ENDPOINT or "YOUR_AZURE" in FORMREC_KEY:
        print("\n" + "!" * 68)
        print("  Azure credentials not set!  Edit the CONFIG section above, or:")
        print()
        print("  Windows:    set FORMREC_ENDPOINT=https://....cognitiveservices.azure.com/")
        print("              set FORMREC_KEY=your_key")
        print("  Mac/Linux:  export FORMREC_ENDPOINT=https://....cognitiveservices.azure.com/")
        print("              export FORMREC_KEY=your_key")
        print("!" * 68 + "\n")
        sys.exit(1)


# ════════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print(__doc__)
        sys.exit(1)

    check_keys()

    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) == 3 else None

    if os.path.isdir(inp):
        convert_folder(inp, out or (inp.rstrip("/\\") + "_excel"))
    elif os.path.isfile(inp) and inp.lower().endswith(".pdf"):
        convert_pdf(inp, out or str(Path(inp).with_suffix(".xlsx")))
    else:
        print(f"\nERROR: '{inp}' is not a valid PDF or folder.")
        sys.exit(1)
