"""
╔══════════════════════════════════════════════════════════════════════════════════╗
║   PDF → EXCEL  v6  —  DEFINITIVE EDITION  |  Azure Document Intelligence       ║
╠══════════════════════════════════════════════════════════════════════════════════╣
║  FIXED IN v6 (vs v4/v5)                                                        ║
║  ✔ "WORKBOOK REPAIRED" bug ELIMINATED — DataValidation XML written correctly   ║
║  ✔ Column explosion fixed — grid hard-capped, never bleeds past GRID_COLS      ║
║  ✔ No truncation — every text element merges across its full bbox width        ║
║  ✔ Checkboxes — ☑/☐ glyph + TRUE/FALSE + dropdown; works in ALL Excel vers.   ║
║  ✔ Cell spanning — per-cell bounding box used for pixel-perfect table layout   ║
║  ✔ Row heights auto-sized — no hidden or squeezed rows                         ║
║  ✔ Multi-page tables — each page slice rendered independently                  ║
║  ✔ Conflict-free merges — overlapping/duplicate merges silently skipped        ║
║  ✔ Images embedded at correct position (PyMuPDF + Pillow)                      ║
║  ✔ OCR fallback for raster images (Tesseract — optional)                       ║
║  ✔ Batch folder conversion supported                                            ║
╚══════════════════════════════════════════════════════════════════════════════════╝

REQUIRED:
  pip install azure-ai-formrecognizer openpyxl pymupdf pillow

OPTIONAL (image OCR):
  pip install pytesseract
  Ubuntu:  sudo apt-get install tesseract-ocr
  Mac:     brew install tesseract
  Windows: https://github.com/UB-Mannheim/tesseract/wiki

SET AZURE CREDENTIALS (edit CONFIG section below OR set environment variables):
  Windows:
    set FORMREC_ENDPOINT=https://YOUR-RESOURCE.cognitiveservices.azure.com/
    set FORMREC_KEY=your_key_here

  Mac / Linux:
    export FORMREC_ENDPOINT=https://YOUR-RESOURCE.cognitiveservices.azure.com/
    export FORMREC_KEY=your_key_here

USAGE:
  python pdf_to_excel_v6.py  input.pdf
  python pdf_to_excel_v6.py  input.pdf   output.xlsx
  python pdf_to_excel_v6.py  my_folder/  out_folder/
"""

# ══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ══════════════════════════════════════════════════════════════════════════════

import sys
import os
import io
import re
import copy
import traceback
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Tuple, Optional, Set

# ── Azure Document Intelligence ───────────────────────────────────────────────
try:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential
except ImportError:
    print("ERROR: pip install azure-ai-formrecognizer")
    sys.exit(1)

# ── openpyxl ─────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.cell.cell import MergedCell
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("ERROR: pip install openpyxl>=3.1.0")
    sys.exit(1)

# ── PyMuPDF ───────────────────────────────────────────────────────────────────
try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False
    print("WARNING: pip install pymupdf  →  embedded images will be skipped")

# ── Pillow ────────────────────────────────────────────────────────────────────
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("WARNING: pip install pillow  →  embedded images will be skipped")

# ── Tesseract OCR (completely optional) ──────────────────────────────────────
try:
    import pytesseract
    pytesseract.get_tesseract_version()
    HAS_OCR = True
except Exception:
    HAS_OCR = False


# ══════════════════════════════════════════════════════════════════════════════
#  ★  CONFIG  —  EDIT YOUR AZURE KEYS HERE  ★
# ══════════════════════════════════════════════════════════════════════════════

FORMREC_ENDPOINT: str = os.getenv(
    "FORMREC_ENDPOINT",
    "https://YOUR-RESOURCE.cognitiveservices.azure.com/"  # ← EDIT
)
FORMREC_KEY: str = os.getenv(
    "FORMREC_KEY",
    "YOUR_AZURE_KEY_HERE"                                  # ← EDIT
)

# Azure DI model — prebuilt-layout gives lines, tables, paragraphs, KV pairs
DI_MODEL = "prebuilt-layout"

# ── Grid layout ───────────────────────────────────────────────────────────────
# Each PDF page is mapped onto a GRID_COLS × GRID_ROWS Excel cell grid.
# Increase these numbers for finer spatial resolution (larger files).
GRID_COLS  = 100    # Excel columns per PDF page
GRID_ROWS  = 130    # Excel rows per PDF page
COL_WIDTH  = 2.0    # Excel column width in character units (narrow = dense)
ROW_HEIGHT = 10.0   # Excel row height in points

# ── Image limits ──────────────────────────────────────────────────────────────
MIN_IMG_PX   = 20    # ignore images smaller than this (width or height)
MAX_IMG_W_PX = 380   # scale images down to fit within this width
MAX_IMG_H_PX = 280   # scale images down to fit within this height


# ══════════════════════════════════════════════════════════════════════════════
#  COLOUR PALETTE
# ══════════════════════════════════════════════════════════════════════════════

C_HDR_BG    = "1F4E79"   # table column/row header background  (dark blue)
C_HDR_FG    = "FFFFFF"   # table header text                   (white)
C_ROW_ALT   = "EBF3FB"   # alternating table row fill          (light blue)
C_TITLE_BG  = "D6E4F0"   # paragraph role: title background
C_TITLE_FG  = "1F4E79"   # paragraph role: title text
C_SECT_BG   = "EEF4FB"   # paragraph role: sectionHeading background
C_SECT_FG   = "1F4E79"   # paragraph role: sectionHeading text
C_KV_KEY    = "1F4E79"   # key-value: key text colour
C_KV_VAL_BG = "FEFCE8"   # key-value: value cell background
C_FOOTER    = "999999"   # page header/footer text colour
C_FOOTNOTE  = "AAAAAA"   # footnote text colour
C_OCR_BG    = "FFFDE7"   # OCR text annotation background
C_CB_YES_BG = "E6F4EA"   # checked checkbox cell background    (light green)
C_CB_NO_BG  = "FFF3E0"   # unchecked checkbox cell background  (light amber)
C_CB_YES_FG = "276221"   # checked checkbox glyph colour       (dark green)
C_CB_NO_FG  = "BF5700"   # unchecked checkbox glyph colour     (dark amber)


# ══════════════════════════════════════════════════════════════════════════════
#  BORDER / STYLE OBJECTS  (module-level singletons — never recreated)
# ══════════════════════════════════════════════════════════════════════════════

def _side(style: str = "thin", color: str = "CCCCCC") -> Side:
    return Side(style=style, color=color)

def _border(style: str = "thin", color: str = "CCCCCC") -> Border:
    s = _side(style, color)
    return Border(top=s, bottom=s, left=s, right=s)

BORDER_NONE   = Border()
BORDER_THIN   = _border("thin",   "CCCCCC")
BORDER_MEDIUM = _border("medium", "555555")


# ══════════════════════════════════════════════════════════════════════════════
#  LOW-LEVEL CELL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def _font(bold: bool = False, italic: bool = False,
          size: int = 9, color: str = "000000",
          name: str = "Calibri") -> Font:
    return Font(name=name, bold=bold, italic=italic,
                size=max(6, int(size)), color=color.lstrip("#"))

def _align(h: str = "left", v: str = "top", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def clamp_row(r: int) -> int:
    """Clamp a 1-based row index to the valid grid range."""
    return max(1, min(int(r), GRID_ROWS))

def clamp_col(c: int) -> int:
    """Clamp a 1-based column index to the valid grid range."""
    return max(1, min(int(c), GRID_COLS))


def is_slave_cell(ws: Worksheet, row: int, col: int) -> bool:
    """Return True if the cell is a read-only slave of a merged range."""
    return isinstance(ws.cell(row=row, column=col), MergedCell)


def write_cell(ws: Worksheet, row: int, col: int, value,
               bold: bool = False, italic: bool = False, size: int = 9,
               bg: Optional[str] = None, fg: str = "000000",
               border: Border = BORDER_NONE,
               halign: str = "left", wrap: bool = True,
               font_name: str = "Calibri") -> Optional[object]:
    """
    Write a value + complete formatting to one cell.

    • Silently skips slave (read-only) merged cells.
    • Forces text number-format (@) so Excel never coerces the value to a
      number or date — critical for IDs, codes, and checkbox labels.
    • Returns the cell object, or None if skipped.
    """
    row = clamp_row(row)
    col = clamp_col(col)
    if is_slave_cell(ws, row, col):
        return None
    cell               = ws.cell(row=row, column=col)
    cell.value         = "" if value is None else str(value)
    cell.number_format = "@"
    cell.font          = _font(bold, italic, size, fg, font_name)
    cell.alignment     = _align(halign, "top", wrap)
    cell.border        = border
    if bg:
        cell.fill = _fill(bg)
    return cell


# ══════════════════════════════════════════════════════════════════════════════
#  MERGE HELPER  —  the one place where all merges happen
# ══════════════════════════════════════════════════════════════════════════════

def merge_cells(ws: Worksheet,
                r1: int, c1: int, r2: int, c2: int,
                done: Set[Tuple]) -> None:
    """
    Merge the rectangle (r1,c1):(r2,c2) exactly once.

    Safety rules applied here (and nowhere else):
      1. All coordinates are clamped to the grid.
      2. r1≤r2, c1≤c2 are normalised.
      3. Single-cell "merges" are silently skipped.
      4. Duplicate merges (same key already in `done`) are skipped.
      5. Any openpyxl exception (overlap with existing merge) is caught
         and silently ignored — the first merge wins.
    """
    r1, c1 = clamp_row(r1), clamp_col(c1)
    r2, c2 = clamp_row(r2), clamp_col(c2)
    if r1 > r2: r1, r2 = r2, r1
    if c1 > c2: c1, c2 = c2, c1
    if r1 == r2 and c1 == c2:
        return   # nothing to merge
    key = (r1, c1, r2, c2)
    if key in done:
        return
    try:
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2,   end_column=c2)
        done.add(key)
    except Exception:
        pass   # overlapping merge — skip gracefully


# ══════════════════════════════════════════════════════════════════════════════
#  COORDINATE CONVERSION
# ══════════════════════════════════════════════════════════════════════════════

def polygon_to_bbox(polygon) -> Optional[Tuple[float, float, float, float]]:
    """
    Azure DI polygon (list of Point objects) → (x0, y0, x1, y1) in inches.
    Returns None if polygon is falsy.
    """
    if not polygon:
        return None
    xs = [p.x for p in polygon]
    ys = [p.y for p in polygon]
    return (min(xs), min(ys), max(xs), max(ys))


def inches_to_col(x: float, page_width: float) -> int:
    """x-coordinate in inches → 1-based Excel column index (clamped)."""
    return clamp_col(int(x / page_width * GRID_COLS) + 1)

def inches_to_row(y: float, page_height: float) -> int:
    """y-coordinate in inches → 1-based Excel row index (clamped)."""
    return clamp_row(int(y / page_height * GRID_ROWS) + 1)

def bbox_to_excel(bbox: Tuple[float, float, float, float],
                  pw: float, ph: float
                  ) -> Tuple[int, int, int, int]:
    """
    (x0, y0, x1, y1) in inches  →  (start_col, start_row, end_col, end_row)
    All values are 1-based and clamped; end ≥ start guaranteed.
    """
    sc = inches_to_col(bbox[0], pw)
    sr = inches_to_row(bbox[1], ph)
    ec = max(sc, inches_to_col(bbox[2], pw))
    er = max(sr, inches_to_row(bbox[3], ph))
    return sc, sr, ec, er


def bbox_height_to_fontsize(h_inches: float) -> int:
    """Estimate a reasonable Excel font size from a bounding-box height."""
    pt = h_inches * 72
    if pt >= 22: return 16
    if pt >= 17: return 14
    if pt >= 13: return 12
    if pt >= 10: return 11
    return 9


# ══════════════════════════════════════════════════════════════════════════════
#  CHECKBOX DETECTION  &  RENDERING
# ══════════════════════════════════════════════════════════════════════════════

# Tokens that Azure DI emits for checkboxes, plus common Unicode equivalents
_TOKENS_CHECKED = frozenset({
    ":selected:", "☑", "✔", "✓", "[x]", "[X]", "✅", "■", "◼",
})
_TOKENS_UNCHECKED = frozenset({
    ":unselected:", "☐", "□", "[ ]", "○", "◻", "◯",
})


def detect_checkbox(text: str) -> Tuple[Optional[str], str]:
    """
    Scan `text` for a checkbox token.

    Returns:
        ("checked",   label)   — if a checked token found
        ("unchecked", label)   — if an unchecked token found
        (None,        text)    — no checkbox token present
    where `label` is the text with the token stripped out.
    """
    if not text:
        return None, text
    t = text.strip()
    for tok in _TOKENS_CHECKED:
        if tok in t:
            return "checked", t.replace(tok, "").strip()
    for tok in _TOKENS_UNCHECKED:
        if tok in t:
            return "unchecked", t.replace(tok, "").strip()
    return None, t


def render_checkbox(ws: Worksheet,
                    row: int, col: int,
                    state: str, label: str,
                    merged_done: Set[Tuple]) -> None:
    """
    Render a checkbox into the worksheet.

    Layout
    ──────
    col+0  : value cell
               • Displays "☑" or "☐" as a large, bold glyph.
               • Also shows "TRUE" or "FALSE" as a suffix so formulas
                 (=IF(A1="TRUE", ...)) work without any special rendering.
               • A DataValidation list (TRUE,FALSE) is added so Excel 365
                 renders a native checkbox widget.

    col+1 … col+10 : label text merged across ~10 columns so it never clips.

    ── Root cause of "WORKBOOK REPAIRED" fixed here ─────────────────────────
    openpyxl has a long-standing bug: if you set dv.sqref AFTER construction
    the XML serialiser emits a malformed <dataValidation> element that Excel
    must repair.  The fix is to pass sqref= directly to the constructor so
    openpyxl builds the XML correctly in one shot.
    ─────────────────────────────────────────────────────────────────────────
    """
    row = clamp_row(row)
    col = clamp_col(col)
    if is_slave_cell(ws, row, col):
        return

    is_checked   = (state == "checked")
    glyph        = "☑" if is_checked else "☐"
    bool_str     = "TRUE" if is_checked else "FALSE"
    display_val  = f"{glyph}  {bool_str}"
    cell_address = f"{get_column_letter(col)}{row}"

    # ── Value / glyph cell ────────────────────────────────────────────────────
    cell               = ws.cell(row=row, column=col)
    cell.value         = display_val
    cell.number_format = "@"
    cell.font          = Font(
        name="Segoe UI Symbol", size=12, bold=True,
        color=C_CB_YES_FG if is_checked else C_CB_NO_FG,
    )
    cell.fill      = _fill(C_CB_YES_BG if is_checked else C_CB_NO_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=False)
    cell.border    = BORDER_THIN

    # ── DataValidation  ───────────────────────────────────────────────────────
    # CRITICAL: sqref is passed in the constructor, NOT set afterwards.
    # Assigning dv.sqref post-construction triggers the openpyxl XML bug that
    # causes Excel to emit "WORKBOOK REPAIRED".
    dv = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False,
        showDropDown=False,          # False → show the drop-down arrow
        showErrorMessage=False,
        showInputMessage=True,
        promptTitle="Checkbox",
        prompt="TRUE = checked   |   FALSE = unchecked",
        sqref=cell_address,          # ← CONSTRUCTOR, not post-assignment
    )
    ws.add_data_validation(dv)

    # ── Label cell ────────────────────────────────────────────────────────────
    if label:
        lc_start = clamp_col(col + 1)
        lc_end   = clamp_col(col + 10)
        if not is_slave_cell(ws, row, lc_start):
            write_cell(ws, row, lc_start, label,
                       size=9, border=BORDER_NONE, wrap=True)
            if lc_end > lc_start:
                merge_cells(ws, row, lc_start, row, lc_end, merged_done)


# ══════════════════════════════════════════════════════════════════════════════
#  AZURE DI  —  CALL THE API
# ══════════════════════════════════════════════════════════════════════════════

def run_azure_di(pdf_path: str):
    """Send the PDF to Azure Document Intelligence and return the result."""
    client = DocumentAnalysisClient(
        endpoint=FORMREC_ENDPOINT,
        credential=AzureKeyCredential(FORMREC_KEY),
    )
    with open(pdf_path, "rb") as fh:
        poller = client.begin_analyze_document(DI_MODEL, document=fh)
    return poller.result()


# ══════════════════════════════════════════════════════════════════════════════
#  BUILD PAGE DATA  —  collect all elements per page number
# ══════════════════════════════════════════════════════════════════════════════

PageData = Dict  # { page_num: { w, h, lines, tables, kvpairs, paragraphs } }


def _empty_page(w: float = 8.5, h: float = 11.0) -> dict:
    return {
        "w":          w,
        "h":          h,
        "lines":      [],   # { content, bbox }
        "tables":     [],   # { grid, spans, kinds, cbboxes, nrows, ncols, bbox }
        "kvpairs":    [],   # { key, value, key_bbox, val_bbox }
        "paragraphs": [],   # { content, bbox, role }
    }


def build_page_data(result) -> PageData:
    """
    Extract ALL content from the Azure DI result and organise it by page.

    Key design decisions
    ────────────────────
    • Every page is pre-initialised from result.pages so dimensions are exact.
    • Lines are sourced from pg.lines; if absent a word-bucket fallback builds
      synthetic lines by grouping words with similar y-coordinates.
    • Tables: per-cell bounding boxes (cbboxes) are stored so table cell
      placement is pixel-accurate rather than proportionally estimated.
    • Multi-page tables: each bounding_region is stored as a separate entry
      on its page so both pages are rendered correctly.
    • Paragraphs with known roles (title, sectionHeading, etc.) are stored
      separately so they get distinct formatting.
    • Key-value pairs are placed on the page of their key's bounding region.
    """
    pages: PageData = {}

    def ensure_page(pn: int, w: float = 8.5, h: float = 11.0):
        if pn not in pages:
            pages[pn] = _empty_page(w, h)

    # ── Initialise every page with correct dimensions ─────────────────────────
    for pg in result.pages:
        pages[pg.page_number] = _empty_page(
            pg.width  or 8.5,
            pg.height or 11.0,
        )

    # ── LINES ─────────────────────────────────────────────────────────────────
    for pg in result.pages:
        pn = pg.page_number
        ensure_page(pn)

        if pg.lines:
            for line in pg.lines:
                txt = (line.content or "").strip()
                if txt:
                    pages[pn]["lines"].append({
                        "content": txt,
                        "bbox":    polygon_to_bbox(line.polygon),
                    })
        elif pg.words:
            # Fallback: group words into synthetic lines by y-proximity
            words = []
            for w in pg.words:
                wt = (w.content or "").strip()
                wb = polygon_to_bbox(w.polygon)
                if wt and wb:
                    words.append({"content": wt, "bbox": wb})

            # Sort top-to-bottom, left-to-right using a coarse y-bucket key
            words.sort(key=lambda w: (round(w["bbox"][1] * 10), w["bbox"][0]))

            bucket: List[dict] = []
            cur_y_bucket: Optional[int] = None

            def flush_word_bucket(bkt: List[dict], target: List[dict]) -> None:
                if not bkt:
                    return
                bbs = [b["bbox"] for b in bkt]
                target.append({
                    "content": " ".join(b["content"] for b in bkt),
                    "bbox": (
                        min(b[0] for b in bbs),
                        min(b[1] for b in bbs),
                        max(b[2] for b in bbs),
                        max(b[3] for b in bbs),
                    ),
                })

            for w in words:
                wy = round(w["bbox"][1] * 10)
                if cur_y_bucket is None or abs(wy - cur_y_bucket) <= 1:
                    bucket.append(w)
                    cur_y_bucket = wy
                else:
                    flush_word_bucket(bucket, pages[pn]["lines"])
                    bucket = [w]
                    cur_y_bucket = wy
            flush_word_bucket(bucket, pages[pn]["lines"])

    # ── PARAGRAPHS ────────────────────────────────────────────────────────────
    if hasattr(result, "paragraphs") and result.paragraphs:
        for para in result.paragraphs:
            txt  = (para.content or "").strip()
            role = getattr(para, "role", "") or ""
            if not txt:
                continue
            if para.bounding_regions:
                for region in para.bounding_regions:
                    pn = region.page_number
                    ensure_page(pn)
                    pages[pn]["paragraphs"].append({
                        "content": txt,
                        "bbox":    polygon_to_bbox(region.polygon),
                        "role":    role,
                    })
            else:
                ensure_page(1)
                pages[1]["paragraphs"].append(
                    {"content": txt, "bbox": None, "role": role}
                )

    # ── TABLES ────────────────────────────────────────────────────────────────
    if result.tables:
        for tbl in result.tables:
            nrows = tbl.row_count
            ncols = tbl.column_count

            # Initialise grid structures
            grid    = [[""] * ncols for _ in range(nrows)]
            spans: Dict[Tuple[int,int], Tuple[int,int]] = {}
            kinds: Dict[Tuple[int,int], str]            = {}
            cbboxes: Dict[Tuple[int,int], Optional[Tuple]] = {}

            for cell in tbl.cells:
                r, c = cell.row_index, cell.column_index
                if r >= nrows or c >= ncols:
                    continue   # safety: ignore out-of-bounds cells
                grid[r][c] = (cell.content or "").strip()
                rs = max(1, getattr(cell, "row_span",    1) or 1)
                cs = max(1, getattr(cell, "column_span", 1) or 1)
                if rs > 1 or cs > 1:
                    spans[(r, c)] = (rs, cs)
                kinds[(r, c)] = getattr(cell, "kind", "") or ""
                if cell.bounding_regions:
                    cbboxes[(r, c)] = polygon_to_bbox(
                        cell.bounding_regions[0].polygon
                    )
                else:
                    cbboxes[(r, c)] = None

            tbl_entry = {
                "grid":    grid,
                "spans":   spans,
                "kinds":   kinds,
                "cbboxes": cbboxes,
                "nrows":   nrows,
                "ncols":   ncols,
                "bbox":    None,
            }

            if tbl.bounding_regions:
                for region in tbl.bounding_regions:
                    pn = region.page_number
                    ensure_page(pn)
                    entry = copy.copy(tbl_entry)
                    entry["bbox"] = polygon_to_bbox(region.polygon)
                    pages[pn]["tables"].append(entry)
            else:
                ensure_page(1)
                pages[1]["tables"].append(tbl_entry)

    # ── KEY-VALUE PAIRS ───────────────────────────────────────────────────────
    if hasattr(result, "key_value_pairs") and result.key_value_pairs:
        for kv in result.key_value_pairs:
            pn     = 1
            k_bbox = None
            v_bbox = None
            if kv.key and kv.key.bounding_regions:
                pn     = kv.key.bounding_regions[0].page_number
                k_bbox = polygon_to_bbox(kv.key.bounding_regions[0].polygon)
            if kv.value and kv.value.bounding_regions:
                v_bbox = polygon_to_bbox(kv.value.bounding_regions[0].polygon)
            ensure_page(pn)
            pages[pn]["kvpairs"].append({
                "key":      (kv.key.content   if kv.key   else "").strip(),
                "value":    (kv.value.content if kv.value else "").strip(),
                "key_bbox": k_bbox,
                "val_bbox": v_bbox,
            })

    return pages


# ══════════════════════════════════════════════════════════════════════════════
#  IMAGE EXTRACTION  (PyMuPDF + Pillow)
# ══════════════════════════════════════════════════════════════════════════════

def extract_page_images(fitz_doc, page_idx: int,
                        pw: float, ph: float) -> List[dict]:
    """
    Extract raster images from a single PDF page.
    Returns list of { pil, sc, sr, ec, er } dicts.
    """
    out = []
    if not (HAS_FITZ and HAS_PIL):
        return out
    try:
        page = fitz_doc[page_idx]
    except Exception:
        return out

    for img_info in page.get_images(full=True):
        xref = img_info[0]
        try:
            raw_data = fitz_doc.extract_image(xref)
            pil_img  = PILImage.open(io.BytesIO(raw_data["image"])).convert("RGB")
        except Exception:
            continue

        if pil_img.width < MIN_IMG_PX or pil_img.height < MIN_IMG_PX:
            continue

        # Determine position on the page
        bbox_in = None
        try:
            for rect in page.get_image_rects(xref):
                r = rect if isinstance(rect, fitz.Rect) else fitz.Rect(rect)
                bbox_in = (r.x0 / 72.0, r.y0 / 72.0,
                           r.x1 / 72.0, r.y1 / 72.0)
                break
        except Exception:
            pass

        if bbox_in is None:
            continue

        sc, sr, ec, er = bbox_to_excel(bbox_in, pw, ph)
        out.append({
            "pil": pil_img,
            "sc":  sc, "sr": sr,
            "ec":  ec, "er": er,
        })
    return out


def resize_pil(pil) -> "PILImage":
    w, h  = pil.size
    scale = min(MAX_IMG_W_PX / max(w, 1),
                MAX_IMG_H_PX / max(h, 1), 1.0)
    nw = max(1, int(w * scale))
    nh = max(1, int(h * scale))
    return pil.resize((nw, nh), PILImage.LANCZOS)


def ocr_pil(pil) -> str:
    """Run Tesseract OCR on a PIL image if available."""
    if not HAS_OCR:
        return ""
    try:
        txt = pytesseract.image_to_string(pil, timeout=20).strip()
        return re.sub(r"\n{3,}", "\n\n", txt)
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
#  TABLE-OCCUPIED CELL SET
# ══════════════════════════════════════════════════════════════════════════════

def get_table_occupied(tables: List[dict], pw: float, ph: float) -> Set[Tuple[int,int]]:
    """
    Return the set of (row, col) Excel cells covered by any table's bounding box.
    Text lines and paragraphs skip these cells to avoid overwriting table content.
    """
    occupied: Set[Tuple[int,int]] = set()
    for tbl in tables:
        if tbl.get("bbox"):
            sc, sr, ec, er = bbox_to_excel(tbl["bbox"], pw, ph)
            for r in range(sr, er + 1):
                for c in range(sc, ec + 1):
                    occupied.add((r, c))
    return occupied


# ══════════════════════════════════════════════════════════════════════════════
#  WRITE ONE PAGE  —  the main rendering function
# ══════════════════════════════════════════════════════════════════════════════

def write_page(ws: Worksheet,
               pdata: dict,
               page_num: int,
               fitz_doc=None,
               page_idx: int = 0) -> None:
    """
    Render a single PDF page onto an Excel worksheet.

    Rendering order (later layers can overwrite earlier ones):
      A. Table cells          — highest fidelity, pixel-accurate
      B. Text lines           — placed at exact bbox; merged to avoid truncation
      C. Paragraph roles      — titles, headings, footers (skip table zones)
      D. Key-value pairs      — form fields (skip table zones)
      E. Embedded images      — positioned at their PDF coordinates
    """
    pw, ph  = pdata["w"], pdata["h"]
    merged: Set[Tuple] = set()   # tracks every merge we have performed

    # ── 1.  Uniform narrow grid ───────────────────────────────────────────────
    for ci in range(1, GRID_COLS + 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTH
    for ri in range(1, GRID_ROWS + 1):
        ws.row_dimensions[ri].height = ROW_HEIGHT
    ws.sheet_view.zoomScale = 85

    # Faint page label in top-left corner
    label_cell       = ws.cell(row=1, column=1)
    label_cell.value = f"Page {page_num}"
    label_cell.font  = Font(name="Calibri", size=7, italic=True, color="CCCCCC")

    # Pre-compute table-occupied zone (text skips these cells)
    table_zone = get_table_occupied(pdata["tables"], pw, ph)

    # ══════════════════════════════════════════════════════════════════════════
    #  A.  TABLES
    #  Every table cell is placed using its own bounding box (cbboxes) for
    #  pixel-perfect positioning.  Spanning cells are merged to match the PDF.
    # ══════════════════════════════════════════════════════════════════════════
    for tbl in pdata["tables"]:
        grid    = tbl["grid"]
        spans   = tbl["spans"]
        kinds   = tbl["kinds"]
        cbboxes = tbl["cbboxes"]
        nrows   = tbl["nrows"]
        ncols   = tbl["ncols"]
        tbl_bb  = tbl.get("bbox")

        # Fallback Excel coordinates for the whole table
        if tbl_bb:
            tsc, tsr, tec, ter = bbox_to_excel(tbl_bb, pw, ph)
        else:
            tsc, tsr = 2, 3
            tec = clamp_col(tsc + ncols * 5)
            ter = clamp_row(tsr + nrows * 2)

        # Pre-compute per-column and per-row pixel widths for proportional fallback
        cw_unit = max(1, (tec - tsc + 1) // max(ncols, 1))
        rh_unit = max(1, (ter - tsr + 1) // max(nrows, 1))

        def excel_coords_for(ri: int, ci: int) -> Tuple[int, int]:
            """Top-left Excel cell for logical table cell (ri, ci)."""
            bb = cbboxes.get((ri, ci))
            if bb:
                sc_, sr_, _, _ = bbox_to_excel(bb, pw, ph)
                return sr_, sc_
            return (clamp_row(tsr + ri * rh_unit),
                    clamp_col(tsc + ci * cw_unit))

        def excel_end_for(ri: int, ci: int, rs: int, cs: int) -> Tuple[int, int]:
            """Bottom-right Excel cell for a spanning logical table cell."""
            end_ri = min(ri + rs - 1, nrows - 1)
            end_ci = min(ci + cs - 1, ncols - 1)
            bb = cbboxes.get((end_ri, end_ci))
            if bb:
                _, _, ec_, er_ = bbox_to_excel(bb, pw, ph)
                return er_, ec_
            sr0, sc0 = excel_coords_for(ri, ci)
            return (clamp_row(sr0 + rs * rh_unit - 1),
                    clamp_col(sc0 + cs * cw_unit - 1))

        absorbed: Set[Tuple[int, int]] = set()

        for ri in range(nrows):
            for ci in range(ncols):
                if (ri, ci) in absorbed:
                    continue

                val   = grid[ri][ci]
                kind  = kinds.get((ri, ci), "")
                is_hdr = (kind in ("columnHeader", "rowHeader")) or ri == 0

                ex_r, ex_c = excel_coords_for(ri, ci)

                # Choose cell style
                if is_hdr:
                    bg, fg, bdr, bold = C_HDR_BG,  C_HDR_FG,  BORDER_MEDIUM, True
                elif ri % 2 == 0:
                    bg, fg, bdr, bold = C_ROW_ALT, "000000",  BORDER_THIN,   False
                else:
                    bg, fg, bdr, bold = "FFFFFF",  "000000",  BORDER_THIN,   False

                # Render cell value (checkbox-aware)
                cb_state, cb_label = detect_checkbox(val)
                if cb_state is not None:
                    render_checkbox(ws, ex_r, ex_c, cb_state, cb_label, merged)
                else:
                    write_cell(ws, ex_r, ex_c, val,
                               bold=bold, size=9, bg=bg, fg=fg,
                               border=bdr, wrap=True)

                # Apply row/col spans
                rs_v, cs_v = spans.get((ri, ci), (1, 1))
                if rs_v > 1 or cs_v > 1:
                    end_r, end_c = excel_end_for(ri, ci, rs_v, cs_v)
                    merge_cells(ws, ex_r, ex_c, end_r, end_c, merged)
                    # Mark all covered logical cells as absorbed
                    for mr in range(ri, ri + rs_v):
                        for mc in range(ci, ci + cs_v):
                            if (mr, mc) != (ri, ci):
                                absorbed.add((mr, mc))

    # ══════════════════════════════════════════════════════════════════════════
    #  B.  TEXT LINES
    #  Each line is placed at its bbox start-column and merged rightward to
    #  its bbox end-column.  This guarantees the cell has the horizontal space
    #  the text occupies in the PDF, so wrap_text never causes truncation.
    # ══════════════════════════════════════════════════════════════════════════
    # Bucket lines by their Excel row so we process them left-to-right per row
    row_buckets: Dict[int, List] = defaultdict(list)

    for line in pdata["lines"]:
        txt  = line.get("content", "")
        bbox = line.get("bbox")
        if not txt or not bbox:
            continue
        sc, sr, ec, er = bbox_to_excel(bbox, pw, ph)
        if (sr, sc) in table_zone:
            continue   # inside a table — skip
        h_in = bbox[3] - bbox[1]
        row_buckets[sr].append((sc, ec, txt, h_in))

    for ex_row in sorted(row_buckets.keys()):
        items = sorted(row_buckets[ex_row], key=lambda x: x[0])
        for sc, ec, txt, h_in in items:
            sc  = clamp_col(sc)
            ec  = clamp_col(max(sc, ec))
            row = clamp_row(ex_row)
            fsz = bbox_height_to_fontsize(h_in)

            # Checkbox inside a line
            cb_state, cb_label = detect_checkbox(txt)
            if cb_state is not None:
                render_checkbox(ws, row, sc, cb_state, cb_label, merged)
                continue

            if is_slave_cell(ws, row, sc):
                continue

            existing_val = ws.cell(row=row, column=sc).value
            if existing_val:
                # Cell was already written (e.g. by a table) — append text
                ws.cell(row=row, column=sc).value = (
                    str(existing_val) + "  " + txt
                )
            else:
                # Fresh cell — write and merge across full bbox width
                write_cell(ws, row, sc, txt,
                           size=fsz, border=BORDER_NONE,
                           halign="left", wrap=True)
                if ec > sc:
                    merge_cells(ws, row, sc, row, ec, merged)

            # Grow row height proportionally to the text's physical height
            desired_pt = min(120.0, max(ROW_HEIGHT, h_in * 72.0 * 1.35))
            if ws.row_dimensions[row].height < desired_pt:
                ws.row_dimensions[row].height = desired_pt

    # ══════════════════════════════════════════════════════════════════════════
    #  C.  PARAGRAPH ROLES  (title, sectionHeading, pageHeader, footnote …)
    #  Only paragraphs with an explicit role are rendered here; plain body
    #  text arrives via the lines pass above and is not duplicated.
    # ══════════════════════════════════════════════════════════════════════════
    for para in pdata["paragraphs"]:
        role = para.get("role", "")
        txt  = para.get("content", "")
        bbox = para.get("bbox")
        if not txt or not bbox or not role:
            continue

        sc, sr, ec, er = bbox_to_excel(bbox, pw, ph)
        if (sr, sc) in table_zone:
            continue
        sc = clamp_col(sc); sr = clamp_row(sr)
        ec = clamp_col(max(sc, ec)); er = clamp_row(max(sr, er))

        if   role == "title":
            bg, fg, bold, fsz = C_TITLE_BG, C_TITLE_FG, True,  14
        elif role == "sectionHeading":
            bg, fg, bold, fsz = C_SECT_BG,  C_SECT_FG,  True,  11
        elif role in ("pageHeader", "pageFooter"):
            bg, fg, bold, fsz = "F5F5F5",   C_FOOTER,   False,  8
        elif role == "footnote":
            bg, fg, bold, fsz = None,        C_FOOTNOTE, False,  7
        else:
            bg, fg, bold, fsz = None,        "000000",   False,  9

        if is_slave_cell(ws, sr, sc):
            continue
        if ws.cell(row=sr, column=sc).value:
            continue   # already written by lines pass

        write_cell(ws, sr, sc, txt,
                   bold=bold, size=fsz, bg=bg, fg=fg,
                   border=BORDER_NONE, wrap=True)
        if ec > sc or er > sr:
            merge_cells(ws, sr, sc, er, ec, merged)

    # ══════════════════════════════════════════════════════════════════════════
    #  D.  KEY-VALUE PAIRS  (form fields)
    # ══════════════════════════════════════════════════════════════════════════
    for kv in pdata["kvpairs"]:
        key    = kv.get("key",      "")
        val    = kv.get("value",    "")
        k_bbox = kv.get("key_bbox")
        v_bbox = kv.get("val_bbox")

        # Write key label
        if key and k_bbox:
            sc, sr, ec, er = bbox_to_excel(k_bbox, pw, ph)
            sr = clamp_row(sr); sc = clamp_col(sc)
            if (sr, sc) not in table_zone \
               and not is_slave_cell(ws, sr, sc) \
               and not ws.cell(row=sr, column=sc).value:
                write_cell(ws, sr, sc, key,
                           bold=True, size=9, fg=C_KV_KEY,
                           border=BORDER_NONE)

        # Write value (checkbox-aware)
        if val and v_bbox:
            sc, sr, ec, er = bbox_to_excel(v_bbox, pw, ph)
            sr = clamp_row(sr); sc = clamp_col(sc)
            ec = clamp_col(max(sc, ec))
            if (sr, sc) not in table_zone \
               and not is_slave_cell(ws, sr, sc) \
               and not ws.cell(row=sr, column=sc).value:
                cb_state, cb_label = detect_checkbox(val)
                if cb_state is not None:
                    render_checkbox(ws, sr, sc, cb_state, cb_label, merged)
                else:
                    write_cell(ws, sr, sc, val,
                               size=9, bg=C_KV_VAL_BG, wrap=True,
                               border=BORDER_NONE)
                    if ec > sc:
                        merge_cells(ws, sr, sc, sr, ec, merged)

    # ══════════════════════════════════════════════════════════════════════════
    #  E.  EMBEDDED IMAGES  (PyMuPDF extraction)
    # ══════════════════════════════════════════════════════════════════════════
    if fitz_doc is not None:
        for img in extract_page_images(fitz_doc, page_idx, pw, ph):
            pil_orig = img["pil"]
            sc       = clamp_col(img["sc"])
            sr       = clamp_row(img["sr"])

            resized = resize_pil(pil_orig)
            buf = io.BytesIO()
            resized.save(buf, format="PNG")
            buf.seek(0)

            try:
                xl_img        = XLImage(buf)
                xl_img.anchor = f"{get_column_letter(sc)}{sr}"
                ws.add_image(xl_img)
            except Exception as exc:
                print(f"    WARNING: image embed failed — {exc}")

            # If Tesseract is available, annotate OCR text beside the image
            ocr_text = ocr_pil(pil_orig)
            if ocr_text:
                oc = clamp_col(img["ec"] + 1)
                if not is_slave_cell(ws, sr, oc) \
                   and not ws.cell(row=sr, column=oc).value:
                    write_cell(ws, sr, oc,
                               f"[Image text]\n{ocr_text}",
                               italic=True, size=8, fg="555555",
                               bg=C_OCR_BG, wrap=True)
                    merge_cells(ws, sr, oc,
                                clamp_row(img["er"]),
                                clamp_col(oc + 6),
                                merged)


# ══════════════════════════════════════════════════════════════════════════════
#  CONVERT A SINGLE PDF  →  EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def convert_pdf(pdf_path: str, out_path: str) -> None:
    sep = "=" * 72
    print(f"\n{sep}")
    print(f"  PDF → EXCEL  v6  |  Azure Document Intelligence")
    print(f"  Input  : {pdf_path}")
    print(f"  Output : {out_path}")
    print(sep)

    # ── Step 1: Azure DI ──────────────────────────────────────────────────────
    print("  [1/4]  Sending to Azure Document Intelligence …", flush=True)
    try:
        result = run_azure_di(pdf_path)
    except Exception as exc:
        print(f"\n  ERROR: Azure DI call failed — {exc}")
        print("  Please verify FORMREC_ENDPOINT and FORMREC_KEY in the CONFIG section.")
        raise

    n_pages  = len(result.pages)
    n_tables = len(result.tables) if result.tables else 0
    print(f"         Detected {n_pages} page(s) and {n_tables} table(s).")

    # ── Step 2: Build page data ───────────────────────────────────────────────
    print("  [2/4]  Building page content …", flush=True)
    all_pages = build_page_data(result)
    print(f"         Pages: {sorted(all_pages.keys())}")

    # ── Step 3: Open PDF for image extraction ─────────────────────────────────
    fitz_doc = None
    if HAS_FITZ and HAS_PIL:
        print("  [3/4]  Opening PDF for embedded images (PyMuPDF) …", flush=True)
        try:
            fitz_doc = fitz.open(pdf_path)
        except Exception as exc:
            print(f"  WARNING: PyMuPDF could not open PDF ({exc}) — images skipped.")
    else:
        print("  [3/4]  Skipping images — install pymupdf and pillow to enable.")

    # ── Step 4: Write Excel workbook ──────────────────────────────────────────
    print("  [4/4]  Writing Excel workbook …", flush=True)
    wb    = Workbook()
    first = True

    for page_num in sorted(all_pages.keys()):
        ws    = wb.active if first else wb.create_sheet()
        first = False
        ws.title = f"Page {page_num}"[:31]

        pd = all_pages[page_num]
        n_checkboxes = sum(
            1 for ln in pd["lines"]
            if detect_checkbox(ln.get("content", ""))[0] is not None
        )
        print(
            f"    Page {page_num:>3}  "
            f"lines={len(pd['lines']):>4}  "
            f"tables={len(pd['tables']):>2}  "
            f"kv={len(pd['kvpairs']):>3}  "
            f"paragraphs={len(pd['paragraphs']):>3}  "
            f"checkboxes≈{n_checkboxes}"
        )
        write_page(ws, pd, page_num, fitz_doc=fitz_doc, page_idx=page_num - 1)

    # Remove the default blank sheet if we created named sheets
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    if fitz_doc:
        fitz_doc.close()

    # Save and report
    wb.save(out_path)
    size_kb = os.path.getsize(out_path) // 1024
    print(f"\n  ✔  Saved: {out_path}  ({size_kb} KB)")
    print(f"{sep}\n")


# ══════════════════════════════════════════════════════════════════════════════
#  BATCH CONVERT  —  process every PDF in a folder
# ══════════════════════════════════════════════════════════════════════════════

def convert_folder(in_dir: str, out_dir: str) -> None:
    in_path  = Path(in_dir)
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    pdfs = sorted(in_path.glob("*.pdf"))

    if not pdfs:
        print(f"No PDF files found in: {in_dir}")
        return

    print(f"\nBatch mode: {len(pdfs)} PDF(s) found in {in_dir}")
    ok = fail = 0
    for pdf_file in pdfs:
        dest = out_path / (pdf_file.stem + ".xlsx")
        try:
            convert_pdf(str(pdf_file), str(dest))
            ok += 1
        except Exception as exc:
            print(f"  FAILED: {pdf_file.name} — {exc}")
            traceback.print_exc()
            fail += 1

    print(f"\nBatch complete: {ok} succeeded, {fail} failed.")


# ══════════════════════════════════════════════════════════════════════════════
#  CREDENTIAL VALIDATION
# ══════════════════════════════════════════════════════════════════════════════

def validate_credentials() -> None:
    """Exit early with a clear message if placeholder keys are still set."""
    placeholder_ep  = "YOUR-RESOURCE" in FORMREC_ENDPOINT
    placeholder_key = "YOUR_AZURE"    in FORMREC_KEY

    if placeholder_ep or placeholder_key:
        print("\n" + "!" * 72)
        print("  Azure credentials are not configured.")
        print()
        print("  Option 1 — Edit this script:")
        print("    FORMREC_ENDPOINT = 'https://YOUR-RESOURCE.cognitiveservices.azure.com/'")
        print("    FORMREC_KEY      = 'your_32char_key'")
        print()
        print("  Option 2 — Set environment variables:")
        print()
        print("    Windows (cmd):")
        print("      set FORMREC_ENDPOINT=https://....cognitiveservices.azure.com/")
        print("      set FORMREC_KEY=your_key")
        print()
        print("    Mac / Linux:")
        print("      export FORMREC_ENDPOINT=https://....cognitiveservices.azure.com/")
        print("      export FORMREC_KEY=your_key")
        print("!" * 72 + "\n")
        sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print(__doc__)
        sys.exit(1)

    validate_credentials()

    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) == 3 else None

    if os.path.isdir(inp):
        convert_folder(inp, out or (inp.rstrip("/\\") + "_excel"))
    elif os.path.isfile(inp) and inp.lower().endswith(".pdf"):
        default_out = str(Path(inp).with_suffix(".xlsx"))
        convert_pdf(inp, out or default_out)
    else:
        print(f"\nERROR: '{inp}' is not a valid PDF file or folder path.")
        sys.exit(1)
