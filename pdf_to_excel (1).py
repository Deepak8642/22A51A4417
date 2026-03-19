"""
PDF → Excel Converter  (Zero System Dependencies)
===================================================
All packages installable via pip — no poppler, no system tools needed.

INSTALL (run once in your terminal / command prompt):
    pip install pdfplumber pypdf pillow openpyxl pymupdf

USAGE:
    python pdf_to_excel.py  input.pdf  output.xlsx

WHAT IT DOES:
    * Extracts text and preserves LEFT / CENTER / RIGHT alignment
    * Reproduces tables with styled headers, alternating rows, borders
    * Extracts embedded images using PyMuPDF (fitz) — no poppler needed
    * Places images at their accurate PDF coordinates in Excel
    * Each PDF page becomes one Excel sheet
"""

import sys, os, io, re
from collections import defaultdict

# Third-party (pip only) -------------------------------------------------------
try:
    import fitz                          # PyMuPDF  ->  pip install pymupdf
except ImportError:
    fitz = None

import pdfplumber                        # pip install pdfplumber
from PIL import Image as PILImage        # pip install pillow
from openpyxl import Workbook            # pip install openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# Layout constants -------------------------------------------------------------
EXCEL_COLS      = 16     # columns used to span the full page width
DEFAULT_COL_W   = 11     # Excel column width (characters)
DEFAULT_ROW_H   = 15     # Excel row height (pts)
TITLE_ROW_H     = 32
HEADING_ROW_H   = 22
TABLE_ROW_H     = 20

# Colour palette ---------------------------------------------------------------
C = dict(
    title_fg   = "1F3864",  title_bg   = "DCE6F1",
    heading_fg = "2F5597",
    body_fg    = "333333",
    hdr_bg     = "1F3864",  hdr_fg     = "FFFFFF",
    total_bg   = "2F5597",  total_fg   = "FFFFFF",
    row_even   = "EBF3FB",  row_odd    = "FFFFFF",
    border     = "BBBBBB",  border_hdr = "2F5597",
)

def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

THIN = Side(style="thin",   color=C["border"])
MED  = Side(style="medium", color=C["border_hdr"])

def make_border(top=THIN, bottom=THIN, left=THIN, right=THIN):
    return Border(top=top, bottom=bottom, left=left, right=right)

HDR_BORDER  = make_border(MED, MED, MED, MED)
DATA_BORDER = make_border()


# ==============================================================================
# IMAGE EXTRACTION  —  PyMuPDF (fitz), pure Python, no poppler
# ==============================================================================
def extract_images_pymupdf(pdf_path):
    """
    Returns a list of dicts, one per image found across all pages:
        { page, index, x0, y0, x1, y1, pil_image }
    Uses PyMuPDF which is a pure pip package — no system tools required.
    """
    if fitz is None:
        print("  [WARNING] PyMuPDF not installed — images will be skipped.")
        print("            Fix with:  pip install pymupdf")
        return []

    results = []
    doc = fitz.open(pdf_path)

    for page_num in range(len(doc)):
        page     = doc[page_num]
        img_list = page.get_images(full=True)

        for img_idx, img_info in enumerate(img_list):
            xref = img_info[0]
            try:
                base_image = doc.extract_image(xref)
                img_bytes  = base_image["image"]

                pil_img = PILImage.open(io.BytesIO(img_bytes)).convert("RGBA")

                # Get bounding box of this image on the page
                rects = page.get_image_rects(xref)
                if rects:
                    r = rects[0]
                    x0, y0, x1, y1 = r.x0, r.y0, r.x1, r.y1
                else:
                    x0, y0 = 0.0, 0.0
                    x1 = float(pil_img.width)
                    y1 = float(pil_img.height)

                results.append(dict(
                    page=page_num, index=img_idx,
                    x0=x0, y0=y0, x1=x1, y1=y1,
                    pil_image=pil_img,
                ))
            except Exception as e:
                print(f"  [WARNING] Could not extract image xref={xref}: {e}")

    doc.close()
    return results


# ==============================================================================
# TEXT ALIGNMENT DETECTION
# ==============================================================================
def detect_alignment(words, page_width, margin_frac=0.10):
    """
    Reads the x-coordinates of every word on the line and decides:
        LEFT   — text hugs the left margin
        CENTER — midpoint of the text span is near the page centre
        RIGHT  — text starts past 55% of page width

    This is pure coordinate maths — no system tools needed.
    """
    if not words:
        return "left"

    x_min   = min(float(w["x0"]) for w in words)
    x_max   = max(float(w["x1"]) for w in words)
    mid_txt = (x_min + x_max) / 2.0
    mid_pg  = page_width / 2.0
    margin  = page_width * margin_frac

    if abs(mid_txt - mid_pg) <= margin:
        return "center"
    if x_min > page_width * 0.55:
        return "right"
    return "left"


def avg_font_size(words):
    sizes = [float(w.get("size") or 10) for w in words if w.get("size")]
    return sum(sizes) / len(sizes) if sizes else 10.0


# ==============================================================================
# HELPERS
# ==============================================================================
def in_table_bbox(y, bboxes, tol=2):
    for (x0, top, x1, bot) in bboxes:
        if top - tol <= y <= bot + tol:
            return True
    return False


def pil_to_xl_image(pil_img, target_w_px, target_h_px):
    """Resize PIL image and return an openpyxl Image object."""
    target_w_px = max(target_w_px, 20)
    target_h_px = max(target_h_px, 20)

    bg = PILImage.new("RGB", pil_img.size, (255, 255, 255))
    if pil_img.mode == "RGBA":
        bg.paste(pil_img, mask=pil_img.split()[3])
    else:
        bg.paste(pil_img.convert("RGB"))

    bg = bg.resize((target_w_px, target_h_px), PILImage.LANCZOS)

    buf = io.BytesIO()
    bg.save(buf, format="PNG")
    buf.seek(0)

    xl_img        = XLImage(buf)
    xl_img.width  = target_w_px
    xl_img.height = target_h_px
    return xl_img


# ==============================================================================
# MAIN CONVERTER
# ==============================================================================
def convert(pdf_path, output_path):
    print(f"\nConverting: {pdf_path}")

    all_images = extract_images_pymupdf(pdf_path)
    print(f"  -> {len(all_images)} image(s) found across all pages")

    wb = Workbook()

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):

            # Sheet setup ------------------------------------------------------
            ws = wb.active if page_num == 0 else wb.create_sheet()
            ws.title = f"Page {page_num + 1}"

            page_w = float(page.width)
            page_h = float(page.height)

            for c in range(1, EXCEL_COLS + 2):
                ws.column_dimensions[get_column_letter(c)].width = DEFAULT_COL_W

            current_row = 1

            # Table bounding boxes ---------------------------------------------
            tbl_objects  = page.find_tables()
            table_bboxes = [tuple(t.bbox) for t in tbl_objects]

            # Group words into text lines (skip words inside tables) -----------
            words    = page.extract_words(
                keep_blank_chars=True, x_tolerance=3, y_tolerance=3
            )
            line_map = defaultdict(list)
            for w in words:
                y = float(w["top"])
                if not in_table_bbox(y, table_bboxes):
                    y_key = round(y / 4) * 4   # 4-pt bucket
                    line_map[y_key].append(w)

            # Images on this page ----------------------------------------------
            page_images = [im for im in all_images if im["page"] == page_num]

            # Build a single sorted event list ---------------------------------
            events = []
            for y_key, wds in line_map.items():
                events.append((y_key, "text", wds))
            for ti, tbl_obj in enumerate(tbl_objects):
                events.append((tbl_obj.bbox[1], "table", (ti, tbl_obj)))
            for ii, img_meta in enumerate(page_images):
                events.append((img_meta["y0"], "image", (ii, img_meta)))

            events.sort(key=lambda e: e[0])

            written_tables = set()
            written_images = set()

            for pdf_y, etype, edata in events:

                # ==============================================================
                # TEXT LINE
                # ==============================================================
                if etype == "text":
                    wds = sorted(edata, key=lambda w: float(w["x0"]))
                    line_text = " ".join(w["text"] for w in wds).strip()
                    if not line_text:
                        continue

                    # Detect alignment from word x-positions
                    alignment = detect_alignment(wds, page_w)
                    size      = avg_font_size(wds)
                    is_title  = size >= 16
                    is_head   = 11 <= size < 16

                    cell = ws.cell(row=current_row, column=1, value=line_text)
                    ws.merge_cells(
                        start_row=current_row, start_column=1,
                        end_row=current_row,   end_column=EXCEL_COLS
                    )

                    # Apply alignment — same keyword used by both PDF and Excel
                    al = Alignment(horizontal=alignment, vertical="center",
                                   wrap_text=True)

                    if is_title:
                        cell.font      = Font(name="Arial", bold=True,
                                              size=16, color=C["title_fg"])
                        cell.fill      = solid(C["title_bg"])
                        cell.alignment = al
                        ws.row_dimensions[current_row].height = TITLE_ROW_H

                    elif is_head:
                        cell.font      = Font(name="Arial", bold=True,
                                              size=12, color=C["heading_fg"])
                        cell.alignment = al
                        ws.row_dimensions[current_row].height = HEADING_ROW_H

                    else:
                        cell.font      = Font(name="Arial", size=10,
                                              color=C["body_fg"])
                        cell.alignment = al
                        ws.row_dimensions[current_row].height = DEFAULT_ROW_H

                    current_row += 1

                # ==============================================================
                # TABLE
                # ==============================================================
                elif etype == "table":
                    ti, tbl_obj = edata
                    if ti in written_tables:
                        continue
                    written_tables.add(ti)

                    tbl_data = tbl_obj.extract()
                    if not tbl_data:
                        continue

                    current_row += 1    # blank gap before table
                    num_cols = max(len(r) for r in tbl_data)

                    for r_i, row in enumerate(tbl_data):
                        is_header = (r_i == 0)
                        is_total  = (r_i == len(tbl_data) - 1)

                        for c_i in range(num_cols):
                            val  = row[c_i] if c_i < len(row) else ""
                            cell = ws.cell(
                                row=current_row, column=c_i + 1,
                                value=str(val).strip() if val else ""
                            )

                            if is_header:
                                cell.font      = Font(name="Arial", bold=True,
                                                      size=10, color=C["hdr_fg"])
                                cell.fill      = solid(C["hdr_bg"])
                                cell.alignment = Alignment(horizontal="center",
                                                           vertical="center",
                                                           wrap_text=True)
                                cell.border    = HDR_BORDER

                            elif is_total:
                                cell.font      = Font(name="Arial", bold=True,
                                                      size=10, color=C["total_fg"])
                                cell.fill      = solid(C["total_bg"])
                                cell.alignment = Alignment(horizontal="center",
                                                           vertical="center")
                                cell.border    = HDR_BORDER

                            else:
                                bg = C["row_even"] if r_i % 2 == 0 else C["row_odd"]
                                cell.font      = Font(name="Arial", size=10,
                                                      color="1A1A1A")
                                cell.fill      = solid(bg)
                                cell.alignment = Alignment(horizontal="center",
                                                           vertical="center")
                                cell.border    = DATA_BORDER

                        ws.row_dimensions[current_row].height = TABLE_ROW_H
                        current_row += 1

                    current_row += 1    # blank gap after table

                # ==============================================================
                # IMAGE — placed at accurate PDF coordinates
                # ==============================================================
                elif etype == "image":
                    ii, img_meta = edata
                    if ii in written_images:
                        continue
                    written_images.add(ii)

                    pil_img  = img_meta["pil_image"]
                    pdf_x0   = img_meta["x0"]
                    pdf_y0   = img_meta["y0"]
                    pdf_x1   = img_meta["x1"]
                    pdf_y1   = img_meta["y1"]

                    img_w_pts = max(pdf_x1 - pdf_x0, 1)
                    img_h_pts = max(pdf_y1 - pdf_y0, 1)

                    # PDF points -> pixels  (96 screen dpi / 72 pts per inch)
                    PX_PER_PT = 96 / 72
                    target_w  = max(int(img_w_pts * PX_PER_PT), 30)
                    target_h  = max(int(img_h_pts * PX_PER_PT), 20)

                    try:
                        xl_img = pil_to_xl_image(pil_img, target_w, target_h)

                        # Map PDF position (pts) -> Excel anchor (row, col)
                        anchor_row = max(1, round((pdf_y0 / page_h) * current_row) + 1)
                        anchor_col = max(1, round((pdf_x0 / page_w) * EXCEL_COLS) + 1)

                        anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
                        ws.add_image(xl_img, anchor)

                        # Reserve rows so image does not overlap text below
                        rows_needed = round(target_h / DEFAULT_ROW_H) + 2
                        end_row     = anchor_row + rows_needed
                        while current_row < end_row:
                            ws.row_dimensions[current_row].height = DEFAULT_ROW_H
                            current_row += 1

                        print(f"  [Page {page_num+1}] Image {ii+1} -> {anchor}  "
                              f"size={target_w}x{target_h}px  "
                              f"pdf_pos=({pdf_x0:.0f},{pdf_y0:.0f})")

                    except Exception as exc:
                        print(f"  [Page {page_num+1}] Image {ii+1} SKIPPED: {exc}")

    wb.save(output_path)
    print(f"\nDone!  Saved -> {output_path}\n")


# Entry point ------------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    pdf_in  = sys.argv[1]
    xls_out = sys.argv[2]

    if not os.path.exists(pdf_in):
        print(f"ERROR: File not found -> {pdf_in}")
        sys.exit(1)

    convert(pdf_in, xls_out)
