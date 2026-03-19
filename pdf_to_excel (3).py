"""
PDF -> Excel Converter  —  FINAL ACCURATE VERSION
==================================================
Works on ANY PDF: forms, reports, invoices, tax documents.

HOW IT WORKS (pixel-accurate approach):
  1. Reads every CHARACTER with its exact (x, y) position from the PDF
  2. Reads every HORIZONTAL LINE to find exact row boundaries
  3. Reads every VERTICAL LINE to find exact column boundaries
  4. Builds a precise grid from those lines
  5. Places each character's text into the correct grid cell
  6. Detects filled rectangles → section headers (black bg) / grey headers
  7. Detects text alignment (left/center/right) per cell using x-position math
  8. Extracts images via PyMuPDF and places at accurate coordinates

INSTALL (pip only — no admin / no system tools needed):
    pip install pdfplumber pymupdf pillow openpyxl

USAGE:
    python pdf_to_excel.py  input.pdf  output.xlsx
"""

import sys, os, io, re
from collections import defaultdict

try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import pdfplumber
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── styling helpers ────────────────────────────────────────────────────────────
def solid(hex6):
    return PatternFill("solid", fgColor=hex6)

def bdr(t="thin", b="thin", l="thin", r="thin",
        tc="AAAAAA", bc="AAAAAA", lc="AAAAAA", rc="AAAAAA"):
    mk = lambda s, c: Side(style=s, color=c) if s else Side(style=None)
    return Border(top=mk(t,tc), bottom=mk(b,bc), left=mk(l,lc), right=mk(r,rc))

THIN  = bdr()
THICK = bdr("medium","medium","medium","medium","2F5597","2F5597","2F5597","2F5597")
NO    = bdr(None,None,None,None)

# ── colour constants ───────────────────────────────────────────────────────────
BLACK_BG  = "000000"   # Part I / II / III headers
GREY_BG   = "D3D3D3"   # column-header grey rows
WHITE_BG  = "FFFFFF"
EVEN_BG   = "F2F2F2"
LABEL_BG  = "F5F5F5"

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 – Extract images (PyMuPDF, pure pip)
# ══════════════════════════════════════════════════════════════════════════════
def get_images(pdf_path):
    if not HAS_FITZ:
        print("  [INFO] pip install pymupdf  to include images")
        return []
    results = []
    doc = fitz.open(pdf_path)
    for pn in range(len(doc)):
        pg = doc[pn]
        for idx, info in enumerate(pg.get_images(full=True)):
            xref = info[0]
            try:
                base  = doc.extract_image(xref)
                pil   = PILImage.open(io.BytesIO(base["image"])).convert("RGBA")
                rects = pg.get_image_rects(xref)
                r     = rects[0] if rects else fitz.Rect(0,0,pil.width,pil.height)
                results.append(dict(page=pn, idx=idx,
                                    x0=r.x0, y0=r.y0, x1=r.x1, y1=r.y1,
                                    pil=pil))
            except Exception as e:
                print(f"  [WARN] img xref={xref}: {e}")
    doc.close()
    return results

def pil_to_xl(pil, wp, hp):
    wp, hp = max(wp,20), max(hp,20)
    bg = PILImage.new("RGB", pil.size, (255,255,255))
    bg.paste(pil.convert("RGB"),
             mask=pil.split()[3] if pil.mode=="RGBA" else None)
    bg = bg.resize((wp,hp), PILImage.LANCZOS)
    buf = io.BytesIO()
    bg.save(buf,"PNG"); buf.seek(0)
    xl = XLImage(buf); xl.width=wp; xl.height=hp
    return xl

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 – Build precise row/column grid from PDF lines
# ══════════════════════════════════════════════════════════════════════════════
def snap(values, tol=3):
    """Merge nearby float values within `tol` pts into a single boundary."""
    vals = sorted(set(round(v,1) for v in values))
    merged = []
    for v in vals:
        if merged and v - merged[-1] <= tol:
            merged[-1] = (merged[-1] + v) / 2   # average
        else:
            merged.append(v)
    return sorted(merged)

def build_grid(page):
    """
    Returns (row_tops, col_lefts) — sorted lists of y/x boundary positions.
    Derived from the actual horizontal and vertical lines in the PDF.
    """
    lines = page.lines
    h_ys  = [l["y0"] for l in lines if abs(l["y0"]-l["y1"]) < 2]
    v_xs  = [l["x0"] for l in lines if abs(l["x0"]-l["x1"]) < 2]

    # Always include page edges
    h_ys += [0, float(page.height)]
    v_xs += [0, float(page.width)]

    # Also harvest word top-positions as soft row hints
    words = page.extract_words(x_tolerance=2, y_tolerance=2,
                               extra_attrs=["fontname","size"])
    for w in words:
        h_ys.append(float(w["top"]) - 2)
        h_ys.append(float(w["bottom"]) + 2)

    row_tops  = snap(h_ys, tol=4)
    col_lefts = snap(v_xs, tol=4)
    return row_tops, col_lefts

def grid_index(value, boundaries):
    """Return the index of the interval [boundaries[i], boundaries[i+1]) that contains value."""
    for i in range(len(boundaries)-1):
        if boundaries[i] - 2 <= value < boundaries[i+1] + 2:
            return i
    return max(0, len(boundaries)-2)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 – Classify each filled rect
# ══════════════════════════════════════════════════════════════════════════════
def classify_rects(page):
    """
    Returns list of dicts: {top, bottom, x0, x1, kind}
      kind = 'section'  (black bg  → Part I / II …)
           = 'grey_hdr' (grey bg   → column header row)
           = 'other'
    """
    result = []
    for r in page.rects:
        color = r.get("non_stroking_color")
        if color is None:
            continue
        # Normalise colour to 0-1 floats
        if isinstance(color, (int, float)):
            g = float(color)
            color = (g, g, g)
        if not isinstance(color, (list,tuple)) or len(color) < 3:
            continue
        r0,g0,b0 = float(color[0]), float(color[1]), float(color[2])
        if r0 < 0.05 and g0 < 0.05 and b0 < 0.05:
            kind = "section"
        elif r0 > 0.6 and g0 > 0.6 and b0 > 0.6:
            kind = "grey_hdr"
        else:
            kind = "other"
        result.append(dict(top=r["top"], bottom=r["bottom"],
                           x0=r["x0"], x1=r["x1"], kind=kind))
    return result

def rect_kind_at(y, rects):
    """Return the kind of filled rect at a given y, or None."""
    for r in rects:
        if r["top"] - 2 <= y <= r["bottom"] + 2:
            return r["kind"]
    return None

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 – Detect text alignment within a cell
# ══════════════════════════════════════════════════════════════════════════════
def detect_align(words_in_cell, cell_x0, cell_x1):
    """
    Given words that fall inside one cell, decide left/center/right.
    Uses the midpoint of the text span vs the cell midpoint.
    """
    if not words_in_cell:
        return "left"
    xmin = min(float(w["x0"]) for w in words_in_cell)
    xmax = max(float(w["x1"]) for w in words_in_cell)
    mid_text = (xmin + xmax) / 2
    mid_cell = (cell_x0 + cell_x1) / 2
    cell_w   = cell_x1 - cell_x0
    if cell_w < 10:
        return "left"
    if abs(mid_text - mid_cell) <= cell_w * 0.15:
        return "center"
    if xmin > cell_x0 + cell_w * 0.55:
        return "right"
    return "left"

# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 – Assign words to grid cells
# ══════════════════════════════════════════════════════════════════════════════
def assign_words_to_cells(words, row_tops, col_lefts):
    """
    Returns dict: {(row_i, col_i): [word, word, ...]}
    """
    cells = defaultdict(list)
    for w in words:
        wy = float(w["top"]) + (float(w["bottom"]) - float(w["top"])) * 0.4
        wx = float(w["x0"]) + 2
        ri = grid_index(wy, row_tops)
        ci = grid_index(wx, col_lefts)
        cells[(ri, ci)].append(w)
    return cells

# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 – Write one page to a worksheet
# ══════════════════════════════════════════════════════════════════════════════
def write_page(ws, page, page_images, excel_max_cols=20):
    page_w = float(page.width)
    page_h = float(page.height)

    # ── get all words with font info ─────────────────────────────────────────
    words = page.extract_words(x_tolerance=2, y_tolerance=2,
                               keep_blank_chars=False,
                               extra_attrs=["fontname","size"])

    # ── build grid ───────────────────────────────────────────────────────────
    row_tops, col_lefts = build_grid(page)
    n_rows = len(row_tops) - 1
    n_cols = len(col_lefts) - 1

    # ── classify filled rectangles ───────────────────────────────────────────
    rects = classify_rects(page)

    # ── assign words to grid cells ───────────────────────────────────────────
    cell_words = assign_words_to_cells(words, row_tops, col_lefts)

    # ── set uniform column widths proportional to PDF widths ─────────────────
    for ci in range(n_cols):
        pdf_w  = col_lefts[ci+1] - col_lefts[ci]
        xl_w   = max(4, round(pdf_w / 6))          # rough pts→chars
        ws.column_dimensions[get_column_letter(ci+1)].width = xl_w

    # ── set row heights proportional to PDF heights ───────────────────────────
    for ri in range(n_rows):
        pdf_h = row_tops[ri+1] - row_tops[ri]
        ws.row_dimensions[ri+1].height = max(10, round(pdf_h * 0.85))

    # ── track which excel cells are already merged ────────────────────────────
    merged_map = {}   # (xl_row, xl_col) -> True

    # ── write each cell ───────────────────────────────────────────────────────
    for ri in range(n_rows):
        mid_y    = (row_tops[ri] + row_tops[ri+1]) / 2
        bg_kind  = rect_kind_at(mid_y, rects)
        xl_row   = ri + 1

        for ci in range(n_cols):
            xl_col  = ci + 1
            wds     = cell_words.get((ri, ci), [])
            text    = " ".join(
                " ".join(w["text"] for w in
                         sorted(cell_words.get((ri,ci),[]),
                                key=lambda w:w["x0"]))
                         .split()
            ) if wds else ""

            if not text and bg_kind is None:
                continue

            # ── detect bold/size ─────────────────────────────────────────────
            is_bold = any("Bold" in str(w.get("fontname","")) for w in wds)
            sizes   = [float(w.get("size") or 9) for w in wds if w.get("size")]
            fs      = round(sum(sizes)/len(sizes)) if sizes else 9

            # ── detect alignment ─────────────────────────────────────────────
            cx0 = col_lefts[ci]
            cx1 = col_lefts[ci+1]
            align = detect_align(wds, cx0, cx1)

            # ── style based on background kind ───────────────────────────────
            if bg_kind == "section":
                fg   = Font(name="Arial", bold=True,  size=max(fs,9), color="FFFFFF")
                fill = solid(BLACK_BG)
                al   = Alignment(horizontal="left", vertical="center", wrap_text=True)
                bd   = NO
            elif bg_kind == "grey_hdr":
                fg   = Font(name="Arial", bold=True,  size=max(fs,8), color="000000")
                fill = solid(GREY_BG)
                al   = Alignment(horizontal=align, vertical="center", wrap_text=True)
                bd   = THIN
            else:
                # normal cell — even/odd row tint
                bg_hex = EVEN_BG if ri % 2 == 0 else WHITE_BG
                fg   = Font(name="Arial", bold=is_bold, size=max(fs,8), color="111111")
                fill = solid(bg_hex)
                al   = Alignment(horizontal=align, vertical="top", wrap_text=True)
                bd   = THIN

            cell = ws.cell(row=xl_row, column=xl_col, value=text or None)
            cell.font      = fg
            cell.fill      = fill
            cell.alignment = al
            cell.border    = bd

    # ── find and merge cells that span multiple grid columns ──────────────────
    # Strategy: for each row, scan consecutive empty cells after a filled cell
    # and merge if they share the same background rect.
    # (pdfplumber's find_tables gives us merged spans directly)
    try:
        for tbl in page.find_tables():
            data = tbl.extract()
            if not data:
                continue
            bbox = tbl.bbox   # x0,top,x1,bottom

            # For each cell in the table, check if it spans multiple cols
            for r_idx, row in enumerate(data):
                # find the pdf y midpoint of this table row
                row_frac = (r_idx + 0.5) / max(len(data), 1)
                pdf_y    = bbox[1] + row_frac * (bbox[3] - bbox[1])
                xl_row   = grid_index(pdf_y, row_tops) + 1

                # collect filled cell column spans
                prev_ci  = None
                prev_xl  = None
                span_start = None

                for c_idx, val in enumerate(row):
                    col_frac = (c_idx + 0.5) / max(len(row), 1)
                    pdf_x    = bbox[0] + col_frac * (bbox[2] - bbox[0])
                    xl_col   = grid_index(pdf_x, col_lefts) + 1

                    if val is not None:
                        if span_start is None:
                            span_start = xl_col
                        prev_xl = xl_col
                    else:
                        if span_start is not None and prev_xl is not None and prev_xl > span_start:
                            try:
                                ws.merge_cells(start_row=xl_row, start_column=span_start,
                                               end_row=xl_row, end_column=prev_xl)
                            except Exception:
                                pass
                        span_start = None
                        prev_xl    = None

                if span_start and prev_xl and prev_xl > span_start:
                    try:
                        ws.merge_cells(start_row=xl_row, start_column=span_start,
                                       end_row=xl_row, end_column=prev_xl)
                    except Exception:
                        pass
    except Exception:
        pass   # merging is best-effort

    # ── place images ──────────────────────────────────────────────────────────
    PX = 96 / 72
    for img in page_images:
        w_px = max(int((img["x1"]-img["x0"])*PX), 30)
        h_px = max(int((img["y1"]-img["y0"])*PX), 20)
        try:
            xl_img = pil_to_xl(img["pil"], w_px, h_px)
            ar = grid_index(img["y0"], row_tops) + 1
            ac = grid_index(img["x0"], col_lefts) + 1
            ws.add_image(xl_img, f"{get_column_letter(ac)}{ar}")
            print(f"    img -> {get_column_letter(ac)}{ar}  ({w_px}x{h_px}px)")
        except Exception as e:
            print(f"    img skipped: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def convert(pdf_path, out_path):
    print(f"\nConverting: {pdf_path}")
    images = get_images(pdf_path)
    print(f"  {len(images)} image(s) found")

    wb = Workbook()

    with pdfplumber.open(pdf_path) as pdf:
        for pn, page in enumerate(pdf.pages):
            ws = wb.active if pn == 0 else wb.create_sheet()
            ws.title = f"Page {pn+1}"
            print(f"  [Page {pn+1}] {page.width:.0f}x{page.height:.0f} pts")

            page_imgs = [im for im in images if im["page"] == pn]
            write_page(ws, page, page_imgs)

    wb.save(out_path)
    print(f"\n  Saved -> {out_path}\n")

# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage:  python pdf_to_excel.py  input.pdf  output.xlsx")
        sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"File not found: {sys.argv[1]}"); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
