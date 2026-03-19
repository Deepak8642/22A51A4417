"""
PDF → Excel  ·  ULTIMATE VERSION
===================================
Fully accurate conversion using a hybrid approach:

  COLUMN SLOTS  = PDF vertical lines  (the actual grid columns of the form)
  ROW SLOTS     = unique char y-positions  (one visual text line = one Excel row)
  TEXT SEGMENTS = x-gap analysis within each line joins chars back into words,
                  then each segment is mapped to the correct column slot
  CELL MERGING  = segments that span multiple column slots are merged
  SECTION HDR   = black filled rect → white bold text, full-row merge
  COL HEADER    = grey filled rect  → grey bg, bold
  ALIGNMENT     = per-segment: text midpoint vs slot midpoint

INSTALL:
    pip install pdfplumber pymupdf pillow openpyxl

USAGE:
    python pdf_to_excel.py  input.pdf  output.xlsx
"""

import sys, os, io
from collections import defaultdict

try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import pdfplumber
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── tunables ───────────────────────────────────────────────────────────────────
GAP_PT       = 8     # char gap > this → new text segment
Y_BUCKET     = 1.0   # chars within this pt vertically → same visual line
SLOT_SNAP    = 5     # v-line values within this pt → merged into one slot boundary
COL_W_SCALE  = 5.5   # pts per Excel column-width character unit

# ── styling ────────────────────────────────────────────────────────────────────
def solid(h):
    return PatternFill("solid", fgColor=h)

def bdr(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)

THIN  = bdr("thin",   "CCCCCC")
THICK = bdr("medium", "555555")
NONE  = Border()

def fnt(bold=False, size=9, color="111111"):
    return Font(name="Arial", bold=bold, size=int(size), color=color)

def aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

# ══════════════════════════════════════════════════════════════════════════════
# IMAGES
# ══════════════════════════════════════════════════════════════════════════════
def get_images(pdf_path):
    if not HAS_FITZ:
        print("  [INFO] pip install pymupdf  to keep embedded images")
        return []
    out, doc = [], fitz.open(pdf_path)
    for pn in range(len(doc)):
        pg = doc[pn]
        for info in pg.get_images(full=True):
            xref = info[0]
            try:
                b   = doc.extract_image(xref)
                pil = PILImage.open(io.BytesIO(b["image"])).convert("RGBA")
                rs  = pg.get_image_rects(xref)
                r   = rs[0] if rs else fitz.Rect(0, 0, pil.width, pil.height)
                out.append(dict(page=pn, x0=r.x0, y0=r.y0,
                                x1=r.x1, y1=r.y1, pil=pil))
            except Exception as e:
                print(f"    [img] {e}")
    doc.close()
    return out

def pil_to_xl(pil, wp, hp):
    wp, hp = max(wp, 20), max(hp, 20)
    bg = PILImage.new("RGB", pil.size, (255, 255, 255))
    bg.paste(pil.convert("RGB"),
             mask=pil.split()[3] if pil.mode == "RGBA" else None)
    bg = bg.resize((wp, hp), PILImage.LANCZOS)
    buf = io.BytesIO(); bg.save(buf, "PNG"); buf.seek(0)
    xl = XLImage(buf); xl.width = wp; xl.height = hp
    return xl

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Extract column slot boundaries from PDF vertical lines
# ══════════════════════════════════════════════════════════════════════════════
def get_col_slots(page):
    """
    Returns sorted list of x-values that are the TRUE column boundaries
    of the form, derived from actual vertical lines drawn in the PDF.
    Nearby values (within SLOT_SNAP pts) are merged into one boundary.
    """
    raw = [float(l["x0"]) for l in page.lines
           if abs(l["x0"] - l["x1"]) < 1]
    raw += [0.0, float(page.width)]

    # Snap nearby values
    slots = []
    for v in sorted(set(round(x, 1) for x in raw)):
        if slots and v - slots[-1] <= SLOT_SNAP:
            slots[-1] = (slots[-1] + v) / 2
        else:
            slots.append(v)
    return sorted(slots)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Group characters into exact visual lines by y-coordinate
# ══════════════════════════════════════════════════════════════════════════════
def get_visual_lines(chars):
    """
    Returns sorted list of (y, [chars]) — one entry per visual text line.
    """
    buckets = defaultdict(list)
    for ch in chars:
        key = round(float(ch["top"]) / Y_BUCKET) * Y_BUCKET
        buckets[key].append(ch)

    lines = []
    for key in sorted(buckets):
        chs = sorted(buckets[key], key=lambda c: float(c["x0"]))
        avg_y = sum(float(c["top"]) for c in chs) / len(chs)
        lines.append((avg_y, chs))
    return lines

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Split line chars into text segments using x-gap analysis
# ══════════════════════════════════════════════════════════════════════════════
def get_segments(line_chars):
    """
    Splits characters into segments wherever the x-gap between consecutive
    chars exceeds GAP_PT.
    Returns list of (x0, x1, text, is_bold, font_size).
    """
    if not line_chars:
        return []
    chs  = sorted(line_chars, key=lambda c: float(c["x0"]))
    segs = []
    seg  = [chs[0]]

    for i in range(1, len(chs)):
        gap = float(chs[i]["x0"]) - float(chs[i-1]["x1"])
        if gap > GAP_PT:
            segs.append(seg)
            seg = []
        seg.append(chs[i])
    segs.append(seg)

    result = []
    for s in segs:
        text = "".join(c["text"] for c in s).strip()
        if not text:
            continue
        x0      = float(s[0]["x0"])
        x1      = float(s[-1]["x1"])
        is_bold = any("Bold" in str(c.get("fontname", "")) for c in s)
        sz      = float(s[0].get("size") or 9)
        result.append((x0, x1, text, is_bold, sz))
    return result

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Map a segment's x-span to column slot indices
# ══════════════════════════════════════════════════════════════════════════════
def find_slot(x, slots):
    """Return the slot index whose left boundary ≤ x < next boundary."""
    for i in range(len(slots) - 1):
        if slots[i] - 2 <= x < slots[i+1] + 2:
            return i
    return len(slots) - 2

def find_end_slot(x1, slots):
    """Return the slot index that contains x1 (right edge of segment)."""
    best = 0
    for i in range(len(slots) - 1):
        if slots[i] <= x1 + 4:
            best = i
    return best

# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Detect text alignment within a column slot
# ══════════════════════════════════════════════════════════════════════════════
def detect_align(seg_x0, seg_x1, slot_x0, slot_x1):
    mid_t = (seg_x0 + seg_x1) / 2
    mid_s = (slot_x0 + slot_x1) / 2
    sw    = slot_x1 - slot_x0
    if sw < 8:
        return "left"
    if abs(mid_t - mid_s) <= sw * 0.18:
        return "center"
    if seg_x0 > slot_x0 + sw * 0.55:
        return "right"
    return "left"

# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — Detect filled rectangle background at a y position
# ══════════════════════════════════════════════════════════════════════════════
def get_fills(page):
    fills = []
    for r in page.rects:
        c = r.get("non_stroking_color")
        if c is None: continue
        if isinstance(c, (int, float)): c = (float(c),) * 3
        if not (isinstance(c, (list, tuple)) and len(c) >= 3): continue
        r0, g0, b0 = float(c[0]), float(c[1]), float(c[2])
        if   r0 < 0.15 and g0 < 0.15 and b0 < 0.15: kind = "black"
        elif r0 > 0.50 and g0 > 0.50 and b0 > 0.50: kind = "grey"
        else: continue
        fills.append(dict(top=float(r["top"]), bottom=float(r["bottom"]),
                          x0=float(r["x0"]),   x1=float(r["x1"]),
                          kind=kind))
    return fills

def fill_at(y, fills, tol=3):
    for f in fills:
        if f["top"] - tol <= y <= f["bottom"] + tol:
            return f["kind"]
    return None

# ══════════════════════════════════════════════════════════════════════════════
# STEP 7 — Write one PDF page to one Excel worksheet
# ══════════════════════════════════════════════════════════════════════════════
def write_page(ws, page, page_images):
    pw    = float(page.width)
    ph    = float(page.height)
    chars = page.chars

    # Column slots from vertical lines
    slots   = get_col_slots(page)
    n_slots = len(slots) - 1

    # Visual lines from character y-positions
    vis_lines = get_visual_lines(chars)

    # Filled rects
    fills = get_fills(page)

    # Set column widths proportional to PDF column widths
    for ci in range(n_slots):
        pdf_w = slots[ci+1] - slots[ci]
        ws.column_dimensions[get_column_letter(ci+1)].width = max(2, round(pdf_w / COL_W_SCALE))

    # Track merged regions to avoid double-writing
    merged_cells = set()

    # Write each visual line as one Excel row
    for xl_row, (y, line_chars) in enumerate(vis_lines, start=1):

        # Row height: use font size of first char
        sz = float(line_chars[0].get("size") or 9) if line_chars else 9
        ws.row_dimensions[xl_row].height = max(11, round(sz * 1.7))

        kind = fill_at(y, fills)
        segs = get_segments(line_chars)

        if not segs:
            continue

        # ── BLACK SECTION HEADER ─────────────────────────────────────────────
        if kind == "black":
            # Join all text — section headers always span full row
            full_text = " ".join(s[2] for s in segs)
            cell = ws.cell(row=xl_row, column=1, value=full_text)
            cell.font      = fnt(bold=True, size=max(int(sz), 9), color="FFFFFF")
            cell.fill      = solid("111111")
            cell.alignment = aln("left")
            cell.border    = THICK
            ws.row_dimensions[xl_row].height = 16
            if n_slots > 1:
                try:
                    ws.merge_cells(start_row=xl_row, start_column=1,
                                   end_row=xl_row,   end_column=n_slots)
                    for c in range(1, n_slots+1):
                        merged_cells.add((xl_row, c))
                except Exception:
                    pass
            continue

        # ── GREY COLUMN HEADER ───────────────────────────────────────────────
        if kind == "grey":
            used = set()
            for (x0, x1, text, is_bold, fsz) in segs:
                ci_s = find_slot(x0, slots)
                ci_e = find_end_slot(x1, slots)
                if ci_s + 1 in used:
                    continue
                al = detect_align(x0, x1,
                                  slots[ci_s], slots[min(ci_s+1, len(slots)-1)])
                cell = ws.cell(row=xl_row, column=ci_s+1, value=text)
                cell.font      = fnt(bold=True, size=max(int(fsz), 8), color="111111")
                cell.fill      = solid("D0D0D0")
                cell.alignment = aln(al)
                cell.border    = THIN
                if ci_e > ci_s:
                    try:
                        ws.merge_cells(start_row=xl_row, start_column=ci_s+1,
                                       end_row=xl_row,   end_column=ci_e+1)
                        for c in range(ci_s+1, ci_e+2):
                            used.add(c)
                    except Exception:
                        pass
                else:
                    used.add(ci_s+1)
            ws.row_dimensions[xl_row].height = 15
            continue

        # ── NORMAL ROW ───────────────────────────────────────────────────────
        bg   = "F6F6F6" if (xl_row % 2 == 0) else "FFFFFF"
        used = set()

        for (x0, x1, text, is_bold, fsz) in segs:
            ci_s = find_slot(x0, slots)
            ci_e = find_end_slot(x1, slots)

            # Skip if already written (overlapping merge)
            if ci_s + 1 in used:
                continue

            al = detect_align(x0, x1,
                              slots[ci_s], slots[min(ci_s+1, len(slots)-1)])

            cell = ws.cell(row=xl_row, column=ci_s+1, value=text)
            cell.font      = fnt(bold=is_bold, size=max(int(fsz), 8))
            cell.fill      = solid(bg)
            cell.alignment = aln(al)
            cell.border    = THIN

            if ci_e > ci_s:
                try:
                    ws.merge_cells(start_row=xl_row, start_column=ci_s+1,
                                   end_row=xl_row,   end_column=ci_e+1)
                    for c in range(ci_s+1, ci_e+2):
                        used.add(c)
                except Exception:
                    pass
            else:
                used.add(ci_s+1)

    # ── IMAGES ────────────────────────────────────────────────────────────────
    line_ys = [y for (y, _) in vis_lines]
    PX = 96 / 72
    for img in page_images:
        wp = max(int((img["x1"] - img["x0"]) * PX), 30)
        hp = max(int((img["y1"] - img["y0"]) * PX), 20)
        try:
            xl_img = pil_to_xl(img["pil"], wp, hp)
            # Find anchor row (nearest line above img)
            ar = 1
            for i, ly in enumerate(line_ys):
                if ly <= img["y0"]:
                    ar = i + 1
            ac = find_slot(img["x0"], slots) + 1
            ws.add_image(xl_img, f"{get_column_letter(ac)}{ar}")
            print(f"    img → {get_column_letter(ac)}{ar} ({wp}×{hp}px)")
        except Exception as e:
            print(f"    img skipped: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def convert(pdf_path, out_path):
    print(f"\nConverting: {pdf_path}")
    images = get_images(pdf_path)
    print(f"  {len(images)} embedded image(s)")

    wb = Workbook()
    with pdfplumber.open(pdf_path) as pdf:
        for pn, page in enumerate(pdf.pages):
            ws = wb.active if pn == 0 else wb.create_sheet()
            ws.title = f"Page {pn+1}"
            pg_imgs  = [im for im in images if im["page"] == pn]
            print(f"  [Page {pn+1}]  {page.width:.0f}×{page.height:.0f} pts")
            write_page(ws, page, pg_imgs)

    wb.save(out_path)
    print(f"\n  ✓  Saved → {out_path}\n")

# ── CLI ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage:  python pdf_to_excel.py  input.pdf  output.xlsx")
        sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"File not found: {sys.argv[1]}"); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
