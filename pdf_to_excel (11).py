"""
PDF → Excel  DEFINITIVE CONVERTER
====================================
Works on ANY PDF form or document accurately.

Key techniques:
  1. Chars grouped by y → visual lines (y_snap=1.5pt)
  2. Each line split into segments by x-gap > 8pt
  3. Dot-leaders (. . . .) merged into parent segment
  4. Global column slots built from ALL segment x0/x1 boundaries
  5. Black wide filled rects = Part headers (white bold, full merge)
  6. Grey filled rects = column headers (grey, bold)
  7. Alignment detected per-segment (left/center/right)
  8. Row heights from font size
  9. Images via PyMuPDF

INSTALL:  pip install pdfplumber pymupdf pillow openpyxl
RUN:      python pdf_to_excel.py  input.pdf  output.xlsx
"""
import sys, os, io, re
from collections import defaultdict

try:
    import fitz; HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import pdfplumber
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── styles ─────────────────────────────────────────────────────────────────────
def solid(h): return PatternFill("solid", fgColor=h)
def mk_bdr(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)
THIN  = mk_bdr("thin",   "CCCCCC")
THICK = mk_bdr("medium", "333333")
def fnt(bold=False, size=9, color="111111"):
    return Font(name="Arial", bold=bold, size=int(max(size, 7)), color=color)
def aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

# ── images ─────────────────────────────────────────────────────────────────────
def get_images(pdf_path):
    if not HAS_FITZ: return []
    out, doc = [], fitz.open(pdf_path)
    for pn in range(len(doc)):
        pg = doc[pn]
        for info in pg.get_images(full=True):
            xref = info[0]
            try:
                b   = doc.extract_image(xref)
                pil = PILImage.open(io.BytesIO(b["image"])).convert("RGBA")
                rs  = pg.get_image_rects(xref)
                r   = rs[0] if rs else fitz.Rect(0, 0, pil.width, pil.height)
                out.append(dict(page=pn, x0=r.x0, y0=r.y0, x1=r.x1, y1=r.y1, pil=pil))
            except: pass
    doc.close(); return out

def pil_to_xl(pil, wp, hp):
    wp, hp = max(wp, 20), max(hp, 20)
    bg = PILImage.new("RGB", pil.size, (255, 255, 255))
    bg.paste(pil.convert("RGB"), mask=pil.split()[3] if pil.mode == "RGBA" else None)
    bg = bg.resize((wp, hp), PILImage.LANCZOS)
    buf = io.BytesIO(); bg.save(buf, "PNG"); buf.seek(0)
    xl = XLImage(buf); xl.width = wp; xl.height = hp; return xl

# ── char grouping ──────────────────────────────────────────────────────────────
def chars_to_lines(chars, y_snap=1.5):
    bkts = defaultdict(list)
    for ch in chars:
        key = round(float(ch["top"]) / y_snap) * y_snap
        bkts[key].append(ch)
    lines = []
    for k in sorted(bkts):
        chs = sorted(bkts[k], key=lambda c: float(c["x0"]))
        lines.append((sum(float(c["top"]) for c in chs) / len(chs), chs))
    return lines

def is_dot_only(chs):
    """True if all chars are dots or spaces — dot leader."""
    return all(c["text"].strip() in (".", "") for c in chs)

def chars_to_segs(chs, gap=8):
    """
    Split chars by x-gap into segments.
    Pure dot-leader segments are merged into the preceding text segment.
    Returns list of dicts: {x0, x1, text, bold, size}.
    """
    if not chs: return []
    chs  = sorted(chs, key=lambda c: float(c["x0"]))
    segs = []; seg = [chs[0]]
    for i in range(1, len(chs)):
        if float(chs[i]["x0"]) - float(chs[i-1]["x1"]) > gap:
            segs.append(seg); seg = []
        seg.append(chs[i])
    segs.append(seg)

    result = []
    for s in segs:
        txt = "".join(c["text"] for c in s).strip()
        if not txt: continue
        if is_dot_only(s) and result:
            # absorb dot leaders into previous segment
            result[-1]["x1"]  = float(s[-1]["x1"])
            result[-1]["text"] += " " + txt
            continue
        result.append(dict(
            x0   = float(s[0]["x0"]),
            x1   = float(s[-1]["x1"]),
            text = txt,
            bold = any("Bold" in str(c.get("fontname", "")) for c in s),
            size = float(s[0].get("size") or 9),
        ))
    return result

# ── fills ──────────────────────────────────────────────────────────────────────
def get_fills(page):
    out = []
    for r in page.rects:
        c = r.get("non_stroking_color")
        if c is None: continue
        if isinstance(c, (int, float)): c = (float(c),) * 3
        if not (isinstance(c, (list, tuple)) and len(c) >= 3): continue
        r0, g0, b0 = float(c[0]), float(c[1]), float(c[2])
        if   r0 < 0.15 and g0 < 0.15 and b0 < 0.15: kind = "black"
        elif r0 > 0.50 and g0 > 0.50 and b0 > 0.50: kind = "grey"
        else: continue
        out.append(dict(
            top=float(r["top"]), bottom=float(r["bottom"]),
            x0=float(r["x0"]),  x1=float(r["x1"]),
            kind=kind,          width=float(r["x1"]-r["x0"])
        ))
    return out

def fill_at(y, fills, page_width, tol=4):
    """
    Return fill kind at y, but ONLY for rects that span most of the page
    width (true background headers, not small checkboxes/squares).
    """
    for f in fills:
        if f["top"] - tol <= y <= f["bottom"] + tol:
            if f["width"] > page_width * 0.3:   # must be wide (>30% of page)
                return f["kind"]
    return None

# ── column slots ───────────────────────────────────────────────────────────────
def build_slots(all_line_segs, page_width, snap=5):
    xs = {0.0, float(page_width)}
    for segs in all_line_segs:
        for s in segs:
            xs.add(s["x0"]); xs.add(s["x1"])
    out = []
    for v in sorted(xs):
        if out and v - out[-1] <= snap: out[-1] = (out[-1] + v) / 2
        else: out.append(v)
    return sorted(out)

def slot_of(x, slots):
    best = 0
    for i in range(len(slots) - 1):
        if slots[i] - 2 <= x: best = i
    return best

def end_slot_of(x1, slots):
    best = 0
    for i in range(len(slots) - 1):
        if slots[i] <= x1 + 4: best = i
    return best

def detect_align(seg, sx0, sx1):
    mt = (seg["x0"] + seg["x1"]) / 2
    mc = (sx0 + sx1) / 2
    cw = sx1 - sx0
    if cw < 5: return "left"
    if abs(mt - mc) <= cw * 0.20: return "center"
    if seg["x0"] > sx0 + cw * 0.55: return "right"
    return "left"

# ── merge helper ───────────────────────────────────────────────────────────────
def safe_merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2: return
    try: ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    except: pass

# ── write one page ─────────────────────────────────────────────────────────────
def write_page(ws, page, page_images):
    pw    = float(page.width)
    chars = page.chars
    fills = get_fills(page)

    vis_lines = chars_to_lines(chars)
    all_segs  = [chars_to_segs(chs) for (_, chs) in vis_lines]
    slots     = build_slots(all_segs, pw, snap=5)
    n_slots   = len(slots) - 1

    # Column widths proportional to PDF pts
    for ci in range(n_slots):
        cw = slots[ci + 1] - slots[ci]
        ws.column_dimensions[get_column_letter(ci + 1)].width = max(2, round(cw / 5.0))

    line_ys = [y for y, _ in vis_lines]

    for xl_row, (y, _) in enumerate(vis_lines, start=1):
        segs = all_segs[xl_row - 1]
        kind = fill_at(y, fills, pw)
        sz   = segs[0]["size"] if segs else 9
        ws.row_dimensions[xl_row].height = max(10, round(sz * 1.65))

        if not segs: continue

        # ── PART HEADER (wide black rect + "Part" text) ──────────────────────
        if kind == "black":
            has_part_kw = any(
                re.search(r'\bPart\b|Summary|Recapture', s["text"], re.I)
                for s in segs
            )
            if has_part_kw:
                txt = "  " + "  ".join(s["text"] for s in segs)
                cell = ws.cell(row=xl_row, column=1, value=txt)
                cell.font      = fnt(bold=True, size=max(int(sz), 9), color="FFFFFF")
                cell.fill      = solid("111111")
                cell.alignment = aln("left")
                cell.border    = THICK
                ws.row_dimensions[xl_row].height = 15
                if n_slots > 1:
                    safe_merge(ws, xl_row, 1, xl_row, n_slots)
                continue

        # ── GREY HEADER ──────────────────────────────────────────────────────
        if kind == "grey":
            used = set()
            for seg in segs:
                ci_s = slot_of(seg["x0"], slots)
                ci_e = end_slot_of(seg["x1"], slots)
                if ci_s + 1 in used: continue
                al   = detect_align(seg, slots[ci_s], slots[min(ci_s+1, len(slots)-1)])
                cell = ws.cell(row=xl_row, column=ci_s+1, value=seg["text"])
                cell.font      = fnt(bold=True, size=max(int(seg["size"]), 8))
                cell.fill      = solid("DCDCDC")
                cell.alignment = aln(al)
                cell.border    = THIN
                if ci_e > ci_s:
                    safe_merge(ws, xl_row, ci_s+1, xl_row, ci_e+1)
                    for c in range(ci_s+1, ci_e+2): used.add(c)
                else:
                    used.add(ci_s+1)
            ws.row_dimensions[xl_row].height = 13
            continue

        # ── NORMAL ROW ────────────────────────────────────────────────────────
        bg   = "F6F6F6" if xl_row % 2 == 0 else "FFFFFF"
        used = set()
        for seg in segs:
            ci_s = slot_of(seg["x0"], slots)
            ci_e = end_slot_of(seg["x1"], slots)
            if ci_s + 1 in used: continue
            al   = detect_align(seg, slots[ci_s], slots[min(ci_s+1, len(slots)-1)])
            cell = ws.cell(row=xl_row, column=ci_s+1, value=seg["text"])
            cell.font      = fnt(bold=seg["bold"], size=max(int(seg["size"]), 8))
            cell.fill      = solid(bg)
            cell.alignment = aln(al)
            cell.border    = THIN
            if ci_e > ci_s:
                safe_merge(ws, xl_row, ci_s+1, xl_row, ci_e+1)
                for c in range(ci_s+1, ci_e+2): used.add(c)
            else:
                used.add(ci_s+1)

    # ── images ────────────────────────────────────────────────────────────────
    PX = 96 / 72
    for img in page_images:
        wp = max(int((img["x1"] - img["x0"]) * PX), 30)
        hp = max(int((img["y1"] - img["y0"]) * PX), 20)
        try:
            xl_img = pil_to_xl(img["pil"], wp, hp)
            ar = 1
            for i, ly in enumerate(line_ys):
                if ly <= img["y0"]: ar = i + 1
            ac = slot_of(img["x0"], slots) + 1
            ws.add_image(xl_img, f"{get_column_letter(ac)}{ar}")
        except Exception as e:
            print(f"    img: {e}")

# ── main ───────────────────────────────────────────────────────────────────────
def convert(pdf_path, out_path):
    print(f"\nConverting: {pdf_path}")
    images = get_images(pdf_path)
    if not images and not HAS_FITZ:
        print("  [INFO] pip install pymupdf  → images preserved")

    wb = Workbook()
    with pdfplumber.open(pdf_path) as pdf:
        for pn, page in enumerate(pdf.pages):
            ws = wb.active if pn == 0 else wb.create_sheet()
            ws.title = f"Page {pn+1}"
            pg_imgs  = [im for im in images if im["page"] == pn]
            print(f"  [Page {pn+1}] {page.width:.0f}×{page.height:.0f}  {len(page.chars)} chars")
            write_page(ws, page, pg_imgs)

    wb.save(out_path)
    print(f"\n  ✓  Saved → {out_path}\n")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python pdf_to_excel.py  input.pdf  output.xlsx")
        sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"Not found: {sys.argv[1]}"); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
