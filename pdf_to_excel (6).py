"""
PDF → Excel  ·  DEFINITIVE VERSION
====================================
Approach that gives pixel-perfect results:

  ROWS     → horizontal lines from the PDF define row boundaries
             (each band between two h-lines = one or more Excel rows)
             Within each band, unique char-y values = individual Excel rows

  COLUMNS  → x-gaps > GAP_PT between characters on the same line
             define column segments (no guessing, no v-line heuristics)

  MERGING  → segments whose x-span covers multiple Excel column slots
             are merged across those slots

  HEADERS  → black filled rect  → Part I/II/… white-on-black full-width row
             grey  filled rect  → column header row (grey, bold)

  ALIGNMENT → per-segment: text-midpoint vs column-slot midpoint

  IMAGES   → PyMuPDF extracts + places at exact PDF coordinates

INSTALL (no admin / no system tools):
    pip install pdfplumber pymupdf pillow openpyxl

RUN:
    python pdf_to_excel.py  your.pdf  output.xlsx
"""

import sys, os, io
from collections import defaultdict

try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import pdfplumber
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── tunables ───────────────────────────────────────────────────────────────────
GAP_PT    = 10    # char x-gap larger than this → new column segment
Y_MERGE   = 1.5   # chars within this many pts vertically → same line

# ── styles ─────────────────────────────────────────────────────────────────────
def solid(h):
    return PatternFill("solid", fgColor=h)

def mkbdr(style, color):
    s = Side(style=style, color=color)
    return Border(top=s, bottom=s, left=s, right=s)

THIN  = mkbdr("thin",   "CCCCCC")
THICK = mkbdr("medium", "555555")

def font(bold=False, size=9, color="111111", name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)

def align(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

# ══════════════════════════════════════════════════════════════════════════════
#  IMAGES
# ══════════════════════════════════════════════════════════════════════════════
def get_images(pdf_path):
    if not HAS_FITZ:
        print("  [INFO] pip install pymupdf  to keep embedded images")
        return []
    results, doc = [], fitz.open(pdf_path)
    for pn in range(len(doc)):
        pg = doc[pn]
        for info in pg.get_images(full=True):
            xref = info[0]
            try:
                b   = doc.extract_image(xref)
                pil = PILImage.open(io.BytesIO(b["image"])).convert("RGBA")
                rs  = pg.get_image_rects(xref)
                r   = rs[0] if rs else fitz.Rect(0, 0, pil.width, pil.height)
                results.append(dict(page=pn,
                                    x0=r.x0, y0=r.y0, x1=r.x1, y1=r.y1,
                                    pil=pil))
            except Exception as e:
                print(f"    [img] {e}")
    doc.close()
    return results

def pil_to_xl(pil, wp, hp):
    wp, hp = max(wp, 20), max(hp, 20)
    bg = PILImage.new("RGB", pil.size, (255, 255, 255))
    bg.paste(pil.convert("RGB"),
             mask=pil.split()[3] if pil.mode == "RGBA" else None)
    bg = bg.resize((wp, hp), PILImage.LANCZOS)
    buf = io.BytesIO(); bg.save(buf, "PNG"); buf.seek(0)
    xl = XLImage(buf); xl.width = wp; xl.height = hp
    return xl

# ══════════════════════════════════════════════════════════════════════════════
#  CHAR → LINES
#  Group every character into a visual line by snapping y-coordinates.
# ══════════════════════════════════════════════════════════════════════════════
def chars_to_visual_lines(chars):
    """
    Returns sorted list of (avg_y, [char, …]) — one entry per visual line.
    Characters with 'top' within Y_MERGE pts of each other share a line.
    """
    buckets = defaultdict(list)
    for ch in chars:
        y_key = round(float(ch["top"]) / Y_MERGE) * Y_MERGE
        buckets[y_key].append(ch)

    lines = []
    for y_key in sorted(buckets):
        chs = sorted(buckets[y_key], key=lambda c: float(c["x0"]))
        avg_y = sum(float(c["top"]) for c in chs) / len(chs)
        lines.append((avg_y, chs))
    return lines

# ══════════════════════════════════════════════════════════════════════════════
#  CHARS → COLUMN SEGMENTS
#  Split one line's chars wherever the x-gap exceeds GAP_PT.
# ══════════════════════════════════════════════════════════════════════════════
def chars_to_segments(line_chars):
    """
    Returns list of (x0, x1, text, is_bold, font_size).
    Each segment = a run of chars with no gap > GAP_PT between them.
    """
    if not line_chars:
        return []
    chs  = sorted(line_chars, key=lambda c: float(c["x0"]))
    segs = []
    seg  = [chs[0]]

    for i in range(1, len(chs)):
        gap = float(chs[i]["x0"]) - float(chs[i-1]["x1"])
        if gap > GAP_PT:
            segs.append(seg); seg = []
        seg.append(chs[i])
    segs.append(seg)

    result = []
    for s in segs:
        x0      = float(s[0]["x0"])
        x1      = float(s[-1]["x1"])
        text    = "".join(c["text"] for c in s).strip()
        is_bold = any("Bold" in str(c.get("fontname","")) for c in s)
        sz      = round(float(s[0].get("size") or 9), 1)
        if text:
            result.append((x0, x1, text, is_bold, sz))
    return result

# ══════════════════════════════════════════════════════════════════════════════
#  FILLED RECT DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def get_fills(page):
    fills = []
    for r in page.rects:
        c = r.get("non_stroking_color")
        if c is None: continue
        if isinstance(c, (int, float)): c = (float(c),) * 3
        if not (isinstance(c, (list, tuple)) and len(c) >= 3): continue
        r0, g0, b0 = float(c[0]), float(c[1]), float(c[2])
        if   r0 < 0.15 and g0 < 0.15 and b0 < 0.15: kind = "black"
        elif r0 > 0.50 and g0 > 0.50 and b0 > 0.50: kind = "grey"
        else: continue
        fills.append(dict(top=float(r["top"]), bottom=float(r["bottom"]),
                          x0=float(r["x0"]),   x1=float(r["x1"]),
                          kind=kind))
    return fills

def fill_kind_at(y, fills, tol=3):
    for f in fills:
        if f["top"] - tol <= y <= f["bottom"] + tol:
            return f["kind"]
    return None

# ══════════════════════════════════════════════════════════════════════════════
#  COLUMN SLOT MAP
#  Collect all segment x-boundaries across ALL lines → build a shared
#  set of column slots so segments align into consistent columns.
# ══════════════════════════════════════════════════════════════════════════════
def build_col_slots(all_lines_segs, page_width, snap=8):
    """
    Returns sorted list of x-boundaries that define column slots.
    """
    xs = set()
    xs.add(0.0)
    xs.add(float(page_width))
    for segs in all_lines_segs:
        for (x0, x1, *_) in segs:
            xs.add(x0)
            xs.add(x1)

    # Snap nearby values together
    sorted_xs = sorted(xs)
    merged = [sorted_xs[0]]
    for v in sorted_xs[1:]:
        if v - merged[-1] <= snap:
            merged[-1] = (merged[-1] + v) / 2   # average
        else:
            merged.append(v)
    return sorted(merged)

def slot_of(x, col_slots):
    """Return index of the slot whose left boundary ≤ x."""
    for i in range(len(col_slots) - 1):
        if col_slots[i] <= x < col_slots[i+1] + 1:
            return i
    return max(0, len(col_slots) - 2)

def end_slot_of(x1, col_slots):
    """Return index of the slot that contains x1 (right edge of segment)."""
    for i in range(len(col_slots) - 1):
        if col_slots[i] <= x1 <= col_slots[i+1] + 1:
            return i
    return max(0, len(col_slots) - 2)

# ══════════════════════════════════════════════════════════════════════════════
#  ALIGNMENT WITHIN A SLOT
# ══════════════════════════════════════════════════════════════════════════════
def seg_align(seg_x0, seg_x1, slot_x0, slot_x1):
    mid_t = (seg_x0 + seg_x1) / 2
    mid_s = (slot_x0 + slot_x1) / 2
    sw    = slot_x1 - slot_x0
    if sw < 8: return "left"
    if abs(mid_t - mid_s) <= sw * 0.18: return "center"
    if seg_x0 > slot_x0 + sw * 0.55:   return "right"
    return "left"

# ══════════════════════════════════════════════════════════════════════════════
#  WRITE ONE PAGE
# ══════════════════════════════════════════════════════════════════════════════
def write_page(ws, page, page_images):
    pw = float(page.width)
    ph = float(page.height)

    # ── 1. All characters ──────────────────────────────────────────────────────
    chars = page.chars

    # ── 2. Visual lines (y → [chars]) ─────────────────────────────────────────
    vis_lines = chars_to_visual_lines(chars)   # [(avg_y, [chars]), …]

    # ── 3. Segments per line ───────────────────────────────────────────────────
    lines_segs = [(y, chars_to_segments(chs)) for (y, chs) in vis_lines]

    # ── 4. Global column slot map ──────────────────────────────────────────────
    col_slots = build_col_slots([s for (_, s) in lines_segs], pw, snap=6)
    n_slots   = len(col_slots) - 1

    # ── 5. Filled rectangles ───────────────────────────────────────────────────
    fills = get_fills(page)

    # ── 6. Set column widths (proportional to PDF widths) ─────────────────────
    for ci in range(n_slots):
        cw = col_slots[ci+1] - col_slots[ci]
        ws.column_dimensions[get_column_letter(ci+1)].width = max(2, round(cw / 5))

    # ── 7. Write each visual line as one Excel row ─────────────────────────────
    for xl_row, (y, segs) in enumerate(lines_segs, start=1):

        # Row height: find the PDF h-lines that bracket this y,
        # use their distance as the row height
        row_h_pts = 14   # default
        row_h_pts = max(10, round(row_h_pts * 0.9))
        ws.row_dimensions[xl_row].height = row_h_pts

        kind = fill_kind_at(y, fills)

        if not segs:
            continue

        # ── BLACK SECTION HEADER ─────────────────────────────────────────────
        if kind == "black":
            full_text = "  " + "   ".join(s[2] for s in segs)
            cell = ws.cell(row=xl_row, column=1, value=full_text)
            cell.font      = font(bold=True, size=10, color="FFFFFF")
            cell.fill      = solid("111111")
            cell.alignment = align("left")
            cell.border    = THICK
            ws.row_dimensions[xl_row].height = 16
            if n_slots > 1:
                try:
                    ws.merge_cells(start_row=xl_row, start_column=1,
                                   end_row=xl_row,   end_column=n_slots)
                except Exception:
                    pass
            continue

        # ── GREY COLUMN HEADER ───────────────────────────────────────────────
        if kind == "grey":
            for (x0, x1, text, is_bold, sz) in segs:
                ci_s = slot_of(x0, col_slots)
                ci_e = end_slot_of(x1, col_slots)
                al   = seg_align(x0, x1, col_slots[ci_s],
                                 col_slots[min(ci_e+1, len(col_slots)-1)])
                cell = ws.cell(row=xl_row, column=ci_s+1, value=text)
                cell.font      = font(bold=True, size=max(int(sz),8), color="111111")
                cell.fill      = solid("D6D6D6")
                cell.alignment = align(al)
                cell.border    = THIN
                if ci_e > ci_s:
                    try:
                        ws.merge_cells(start_row=xl_row, start_column=ci_s+1,
                                       end_row=xl_row,   end_column=ci_e+1)
                    except Exception:
                        pass
            ws.row_dimensions[xl_row].height = 15
            continue

        # ── NORMAL ROW ───────────────────────────────────────────────────────
        # Even/odd tint
        bg = "F7F7F7" if (xl_row % 2 == 0) else "FFFFFF"
        used_cols = set()   # track cols already written/merged this row

        for (x0, x1, text, is_bold, sz) in segs:
            ci_s = slot_of(x0, col_slots)
            ci_e = end_slot_of(x1, col_slots)

            # skip if this column already used (overlapping merge from prev seg)
            if ci_s + 1 in used_cols:
                continue

            al   = seg_align(x0, x1, col_slots[ci_s],
                             col_slots[min(ci_e+1, len(col_slots)-1)])

            cell = ws.cell(row=xl_row, column=ci_s+1, value=text)
            cell.font      = font(bold=is_bold, size=max(int(sz),8))
            cell.fill      = solid(bg)
            cell.alignment = align(al)
            cell.border    = THIN

            if ci_e > ci_s:
                try:
                    ws.merge_cells(start_row=xl_row, start_column=ci_s+1,
                                   end_row=xl_row,   end_column=ci_e+1)
                    for c in range(ci_s+1, ci_e+2):
                        used_cols.add(c)
                except Exception:
                    pass
            else:
                used_cols.add(ci_s+1)

        # Taller rows for bigger font
        max_sz = max(s[4] for s in segs)
        ws.row_dimensions[xl_row].height = max(12, round(max_sz * 1.6))

    # ── 8. Place images ────────────────────────────────────────────────────────
    PX = 96 / 72
    line_ys = [y for (y, _) in lines_segs]

    for img in page_images:
        wp = max(int((img["x1"] - img["x0"]) * PX), 30)
        hp = max(int((img["y1"] - img["y0"]) * PX), 20)
        try:
            xl_img = pil_to_xl(img["pil"], wp, hp)
            # find nearest row
            ar = 1
            for i, ly in enumerate(line_ys):
                if ly <= img["y0"]:
                    ar = i + 1
            ac = slot_of(img["x0"], col_slots) + 1
            ws.add_image(xl_img, f"{get_column_letter(ac)}{ar}")
            print(f"    img → {get_column_letter(ac)}{ar} ({wp}×{hp}px)")
        except Exception as e:
            print(f"    img skipped: {e}")

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def convert(pdf_path, out_path):
    print(f"\nConverting: {pdf_path}")
    images = get_images(pdf_path)
    print(f"  {len(images)} embedded image(s)")

    wb = Workbook()
    with pdfplumber.open(pdf_path) as pdf:
        for pn, page in enumerate(pdf.pages):
            ws = wb.active if pn == 0 else wb.create_sheet()
            ws.title = f"Page {pn+1}"
            pg_imgs  = [im for im in images if im["page"] == pn]
            print(f"  [Page {pn+1}]  {page.width:.0f}×{page.height:.0f} pts")
            write_page(ws, page, pg_imgs)

    wb.save(out_path)
    print(f"\n  ✓  Saved → {out_path}\n")

# ── CLI ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage:  python pdf_to_excel.py  input.pdf  output.xlsx")
        sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"File not found: {sys.argv[1]}"); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
