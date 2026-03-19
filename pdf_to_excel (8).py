"""
PDF → Excel  ·  EXACT MATCH VERSION
=====================================
Produces an Excel that exactly mirrors the PDF layout:
- Every visual line = one Excel row
- Column positions from PDF vertical lines (exact x mapping)
- Text segments placed in correct column slot by x-coordinate lookup
- Black section headers → white bold, full-width merge
- Grey column headers → grey bg, bold
- Mixed rows (e.g. "Total" bold + "$" same line) handled correctly
- Empty field rows show borders like the PDF
- Images placed at exact coordinates

INSTALL:  pip install pdfplumber pymupdf pillow openpyxl
RUN:      python pdf_to_excel.py  input.pdf  output.xlsx
"""

import sys, os, io
from collections import defaultdict

try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

import pdfplumber
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
GAP_PT   = 8    # x-gap bigger than this → new text segment
Y_SNAP   = 1.0  # chars within this many pts vertically → same visual line

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def solid(h):
    return PatternFill("solid", fgColor=h)

def side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def bdr(t="thin", b="thin", l="thin", r="thin", c="CCCCCC"):
    s = lambda st: Side(style=st, color=c) if st else Side(style=None)
    return Border(top=s(t), bottom=s(b), left=s(l), right=s(r))

THIN  = bdr()
THICK = bdr("medium","medium","medium","medium", c="444444")
FIELD = bdr(b="thin", t=None, l=None, r=None)  # only bottom border = field line

def fnt(bold=False, size=9, color="000000"):
    return Font(name="Arial", bold=bold, size=int(max(size,7)), color=color)

def aln(h="left", wrap=True, v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ─────────────────────────────────────────────────────────────────────────────
# IMAGE HANDLING
# ─────────────────────────────────────────────────────────────────────────────
def get_images(pdf_path):
    if not HAS_FITZ:
        print("  [INFO] pip install pymupdf  to preserve images")
        return []
    out, doc = [], fitz.open(pdf_path)
    for pn in range(len(doc)):
        pg = doc[pn]
        for info in pg.get_images(full=True):
            xref = info[0]
            try:
                b   = doc.extract_image(xref)
                pil = PILImage.open(io.BytesIO(b["image"])).convert("RGBA")
                rs  = pg.get_image_rects(xref)
                r   = rs[0] if rs else fitz.Rect(0, 0, pil.width, pil.height)
                out.append(dict(page=pn, x0=r.x0, y0=r.y0,
                                x1=r.x1, y1=r.y1, pil=pil))
            except: pass
    doc.close()
    return out

def pil_to_xl(pil, wp, hp):
    wp, hp = max(wp, 20), max(hp, 20)
    bg = PILImage.new("RGB", pil.size, (255, 255, 255))
    bg.paste(pil.convert("RGB"),
             mask=pil.split()[3] if pil.mode == "RGBA" else None)
    bg = bg.resize((wp, hp), PILImage.LANCZOS)
    buf = io.BytesIO(); bg.save(buf, "PNG"); buf.seek(0)
    xl = XLImage(buf); xl.width = wp; xl.height = hp
    return xl

# ─────────────────────────────────────────────────────────────────────────────
# CORE: extract column boundaries from PDF vertical lines
# ─────────────────────────────────────────────────────────────────────────────
def get_col_boundaries(page):
    """
    Returns sorted list of x-values = TRUE column boundaries of the PDF form.
    Derived from actual vertical lines drawn in the PDF.
    """
    raw = [float(l["x0"]) for l in page.lines
           if abs(float(l["x0"]) - float(l["x1"])) < 1]
    raw += [0.0, float(page.width)]
    # cluster nearby values
    vals = sorted(set(round(x, 1) for x in raw))
    cols = [vals[0]]
    for v in vals[1:]:
        if v - cols[-1] <= 4:
            cols[-1] = (cols[-1] + v) / 2
        else:
            cols.append(v)
    return sorted(cols)

# ─────────────────────────────────────────────────────────────────────────────
# CORE: group characters into visual lines
# ─────────────────────────────────────────────────────────────────────────────
def get_visual_lines(chars):
    """
    Returns sorted list of (avg_y, [chars_sorted_by_x]).
    Each entry = one visual text row in the PDF.
    """
    buckets = defaultdict(list)
    for ch in chars:
        key = round(float(ch["top"]) / Y_SNAP) * Y_SNAP
        buckets[key].append(ch)
    lines = []
    for key in sorted(buckets):
        chs = sorted(buckets[key], key=lambda c: float(c["x0"]))
        y   = sum(float(c["top"]) for c in chs) / len(chs)
        lines.append((y, chs))
    return lines

# ─────────────────────────────────────────────────────────────────────────────
# CORE: split line chars into text segments by x-gap
# ─────────────────────────────────────────────────────────────────────────────
def get_segments(line_chars):
    """
    Splits chars into segments wherever gap > GAP_PT.
    Returns [(x0, x1, text, bold, size), ...]
    """
    if not line_chars:
        return []
    chs = sorted(line_chars, key=lambda c: float(c["x0"]))
    segs, seg = [], [chs[0]]
    for i in range(1, len(chs)):
        if float(chs[i]["x0"]) - float(chs[i-1]["x1"]) > GAP_PT:
            segs.append(seg); seg = []
        seg.append(chs[i])
    segs.append(seg)

    result = []
    for s in segs:
        text = "".join(c["text"] for c in s).strip()
        if not text: continue
        result.append((
            float(s[0]["x0"]),
            float(s[-1]["x1"]),
            text,
            any("Bold" in str(c.get("fontname","")) for c in s),
            float(s[0].get("size") or 9)
        ))
    return result

# ─────────────────────────────────────────────────────────────────────────────
# CORE: find which column slot an x-value belongs to
# ─────────────────────────────────────────────────────────────────────────────
def slot_of(x, cols):
    for i in range(len(cols)-1):
        if cols[i] - 3 <= x < cols[i+1] + 3:
            return i
    return len(cols) - 2

def end_slot_of(x1, cols):
    best = 0
    for i in range(len(cols)-1):
        if cols[i] <= x1 + 5:
            best = i
    return best

# ─────────────────────────────────────────────────────────────────────────────
# CORE: detect text alignment within a slot
# ─────────────────────────────────────────────────────────────────────────────
def get_align(sx0, sx1, col_l, col_r):
    mid_t = (sx0 + sx1) / 2
    mid_c = (col_l + col_r) / 2
    cw    = col_r - col_l
    if cw < 10: return "left"
    if abs(mid_t - mid_c) <= cw * 0.18: return "center"
    if sx0 > col_l + cw * 0.55: return "right"
    return "left"

# ─────────────────────────────────────────────────────────────────────────────
# CORE: filled rect detection
# ─────────────────────────────────────────────────────────────────────────────
def get_fills(page):
    fills = []
    for r in page.rects:
        c = r.get("non_stroking_color")
        if c is None: continue
        if isinstance(c, (int,float)): c = (float(c),)*3
        if not (isinstance(c,(list,tuple)) and len(c)>=3): continue
        r0,g0,b0 = float(c[0]),float(c[1]),float(c[2])
        if   r0 < 0.15 and g0 < 0.15 and b0 < 0.15: kind = "black"
        elif r0 > 0.50 and g0 > 0.50 and b0 > 0.50: kind = "grey"
        else: continue
        fills.append(dict(top=float(r["top"]), bottom=float(r["bottom"]),
                          x0=float(r["x0"]),   x1=float(r["x1"]), kind=kind))
    return fills

def fill_at(y, fills, tol=3):
    for f in fills:
        if f["top"]-tol <= y <= f["bottom"]+tol:
            return f["kind"]
    return None

# ─────────────────────────────────────────────────────────────────────────────
# WRITE ONE PAGE TO WORKSHEET
# ─────────────────────────────────────────────────────────────────────────────
def write_page(ws, page, page_imgs):
    pw    = float(page.width)
    ph    = float(page.height)
    chars = page.chars

    cols     = get_col_boundaries(page)   # x boundaries
    n_cols   = len(cols) - 1
    vis      = get_visual_lines(chars)    # visual rows
    fills    = get_fills(page)

    # ── Set column widths proportional to PDF column widths ──────────────────
    for ci in range(n_cols):
        w = cols[ci+1] - cols[ci]
        ws.column_dimensions[get_column_letter(ci+1)].width = max(2, round(w / 5.5))

    # ── Write each visual line ───────────────────────────────────────────────
    used_merge = set()   # (row, col) already part of a merge

    for xl_row, (y, line_chars) in enumerate(vis, start=1):

        segs  = get_segments(line_chars)
        kind  = fill_at(y, fills)

        # row height from font size
        sz_max = max((float(c.get("size") or 9) for c in line_chars), default=9)
        ws.row_dimensions[xl_row].height = max(11, round(sz_max * 1.65))

        if not segs:
            continue

        # ── BLACK SECTION HEADER ─────────────────────────────────────────────
        if kind == "black":
            text = " ".join(s[2] for s in segs)
            cell = ws.cell(row=xl_row, column=1, value=text)
            cell.font      = fnt(bold=True, size=9, color="FFFFFF")
            cell.fill      = solid("111111")
            cell.alignment = aln("left")
            cell.border    = THICK
            ws.row_dimensions[xl_row].height = 15
            if n_cols > 1:
                try:
                    ws.merge_cells(start_row=xl_row, start_column=1,
                                   end_row=xl_row, end_column=n_cols)
                    for c in range(1, n_cols+1):
                        used_merge.add((xl_row, c))
                except: pass
            continue

        # ── GREY COLUMN HEADER ROW ───────────────────────────────────────────
        if kind == "grey":
            done = set()
            for (x0, x1, text, bold, fsz) in segs:
                cs = slot_of(x0, cols)
                ce = end_slot_of(x1, cols)
                if cs+1 in done: continue
                al = get_align(x0, x1, cols[cs], cols[min(cs+1, len(cols)-1)])
                cell = ws.cell(row=xl_row, column=cs+1, value=text)
                cell.font      = fnt(bold=True, size=max(int(fsz),7), color="111111")
                cell.fill      = solid("D6D6D6")
                cell.alignment = aln(al)
                cell.border    = THIN
                if ce > cs:
                    try:
                        ws.merge_cells(start_row=xl_row, start_column=cs+1,
                                       end_row=xl_row, end_column=ce+1)
                        for c in range(cs+1, ce+2): done.add(c)
                    except: pass
                else:
                    done.add(cs+1)
            ws.row_dimensions[xl_row].height = 14
            continue

        # ── NORMAL ROW ───────────────────────────────────────────────────────
        bg   = "F5F5F5" if xl_row % 2 == 0 else "FFFFFF"
        done = set()

        for (x0, x1, text, bold, fsz) in segs:
            cs = slot_of(x0, cols)
            ce = end_slot_of(x1, cols)
            if cs+1 in done: continue

            al = get_align(x0, x1, cols[cs], cols[min(cs+1, len(cols)-1)])

            cell = ws.cell(row=xl_row, column=cs+1, value=text)
            cell.font      = fnt(bold=bold, size=max(int(fsz),7))
            cell.fill      = solid(bg)
            cell.alignment = aln(al)
            cell.border    = THIN

            if ce > cs:
                try:
                    ws.merge_cells(start_row=xl_row, start_column=cs+1,
                                   end_row=xl_row, end_column=ce+1)
                    for c in range(cs+1, ce+2): done.add(c)
                except: pass
            else:
                done.add(cs+1)

    # ── IMAGES ───────────────────────────────────────────────────────────────
    ys = [y for (y,_) in vis]
    for img in page_imgs:
        PX = 96/72
        wp = max(int((img["x1"]-img["x0"])*PX), 30)
        hp = max(int((img["y1"]-img["y0"])*PX), 20)
        try:
            xi = pil_to_xl(img["pil"], wp, hp)
            ar = next((i+1 for i,ly in enumerate(ys) if ly<=img["y0"]), 1)
            ac = slot_of(img["x0"], cols) + 1
            ws.add_image(xi, f"{get_column_letter(ac)}{ar}")
        except: pass

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def convert(pdf_path, out_path):
    print(f"\nConverting: {pdf_path}")
    imgs = get_images(pdf_path)
    print(f"  {len(imgs)} image(s) found")

    wb = Workbook()
    with pdfplumber.open(pdf_path) as pdf:
        for pn, page in enumerate(pdf.pages):
            ws = wb.active if pn == 0 else wb.create_sheet()
            ws.title = f"Page {pn+1}"
            pimgs = [im for im in imgs if im["page"] == pn]
            print(f"  [Page {pn+1}]  {page.width:.0f}x{page.height:.0f}")
            write_page(ws, page, pimgs)

    wb.save(out_path)
    print(f"\n  Done → {out_path}\n")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python pdf_to_excel.py  input.pdf  output.xlsx")
        sys.exit(1)
    if not os.path.exists(sys.argv[1]):
        print(f"Not found: {sys.argv[1]}"); sys.exit(1)
    convert(sys.argv[1], sys.argv[2])
